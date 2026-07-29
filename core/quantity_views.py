"""
AI 구조산출 자동화 — views.py
구조도면 ZIP(DWG) + 구조/건축 PDF 합본을 받아 Gemini로 수량 추출

필요 패키지:
    pip install ezdxf pdf2image pillow google-genai openpyxl
    apt-get install poppler-utils  (pdf2image 의존, macOS는 brew install poppler)
"""

import base64
import hashlib
import io
import json
import logging
import math
import os
import re
import shutil
import subprocess
import tempfile
import threading
import time
import unicodedata
import zipfile
from functools import wraps

import ezdxf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_bytes, pdfinfo_from_bytes
from PIL import Image, ImageDraw
from pypdf import PdfReader, PdfWriter

from django.conf import settings
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from google import genai
from google.genai import types

from .views import admin_required
from .quantity_calc import compute_structural_quantities, compute_massing_model, REBAR_UNIT_WEIGHT

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
#  접근 제한: 이 수량산출 도구는 요청 1건마다 Gemini API 비용이 실제로 발생한다.
#  로그인/권한 체크 없이 공개돼 있으면 누구든(봇 포함) 파일 없이/의미없는 파일로 반복
#  호출해서 크레딧을 소모시킬 수 있다 — 아직 결제 시스템을 붙이기 전(관리자 테스트 단계)
#  이므로, 실제로 비용이 드는 화면/엔드포인트는 전부 관리자(스태프/슈퍼유저)만 쓸 수
#  있게 막는다. 나중에 결제 시스템이 붙으면 이 제한을 결제 여부 체크로 바꾸면 된다.
# ─────────────────────────────────────────────
def _admin_only_json(view_func):
    """JSON을 돌려주는 API 엔드포인트용 — user_passes_test처럼 로그인 페이지로 리다이렉트하면
    fetch()가 HTML을 JSON으로 파싱하려다 깨지므로, 대신 403 JSON을 그대로 돌려준다."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not admin_required(request.user):
            return JsonResponse({"error": "관리자만 사용할 수 있는 기능입니다."}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped

# core/ai_writer.py와 동일한 방식으로 .env의 GEMINI_API_KEY를 사용합니다.
GEMINI_QUANTITY_MODEL = os.environ.get("GEMINI_QUANTITY_MODEL") or os.environ.get("GEMINI_TEXT_MODEL", "gemini-2.5-pro")


def get_gemini_client():
    """GEMINI_API_KEY가 없으면 None을 반환합니다 (import 시점에 죽지 않도록 지연 초기화)."""
    api_key = (getattr(settings, "GEMINI_API_KEY", None) or os.environ.get("GEMINI_API_KEY", "")).strip()
    if not api_key:
        return None
    return genai.Client(api_key=api_key)


# ─────────────────────────────────────────────
#  개발용 추출 결과 캐시: 팝업 UI/계산 로직을 다듬는 동안 같은 도면으로 반복 테스트하면
#  로직만 바뀌었을 뿐인데도 매번 Gemini(구조부재 추출)를 다시 호출해서 비용이 든다.
#  QTY_DEV_CACHE=1 환경변수가 켜져 있을 때만 동작하며, 기본값(꺼짐)에서는 실제 서비스
#  동작에 전혀 영향을 주지 않는다 — 같은 PDF+DWG 조합이면 이전에 저장해둔 추출 결과를
#  그대로 재사용하고(Gemini 호출 0회), 처음 보는 조합이면 평소처럼 호출한 뒤 결과를
#  로컬 임시 폴더에 저장해 다음 테스트부터 재사용한다. 실제 물량 계산(quantity_calc)과
#  검토 팝업(UI)은 캐시와 무관하게 매번 새로 돈다 — 아끼는 건 Gemini 호출뿐이다.
# ─────────────────────────────────────────────
_DEV_CACHE_ENABLED = os.environ.get("QTY_DEV_CACHE", "").strip().lower() in ("1", "true", "yes")
_DEV_CACHE_DIR = os.path.join(tempfile.gettempdir(), "cbl_qty_dev_cache")


def _dev_cache_key(pdf_bytes, dwg_data):
    h = hashlib.sha256()
    h.update(pdf_bytes or b"")
    try:
        h.update(json.dumps(dwg_data, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8"))
    except Exception:
        pass
    return h.hexdigest()


def _dev_cache_load(cache_key):
    path = os.path.join(_DEV_CACHE_DIR, cache_key + ".json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        data["notes"] = ["[개발 캐시] 이전에 같은 도면으로 추출한 결과를 그대로 재사용했습니다 — Gemini를 다시 호출하지 않았습니다."] + list(data.get("notes") or [])
        return data
    except Exception:
        return None


def _dev_cache_save(cache_key, result):
    try:
        os.makedirs(_DEV_CACHE_DIR, exist_ok=True)
        with open(os.path.join(_DEV_CACHE_DIR, cache_key + ".json"), "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False)
    except Exception:
        pass


# ─────────────────────────────────────────────
#  진행률 표시: api_run_quantity가 한 번의 긴 HTTP 요청 안에서 여러 단계(입면/건축분석/
#  구조부재 배치추출)를 거치는 동안, 프론트엔드가 별도로 폴링해서 "지금 어디까지 됐는지"를
#  볼 수 있게 하는 아주 단순한 인메모리 진행상황 저장소.
#  - 로컬 단일 프로세스 개발 서버 기준으로 설계했다(멀티 워커/멀티 서버 배포에서는 각
#    워커가 자기 메모리만 보므로 폴링이 다른 워커로 튀면 못 볼 수 있음 — 지금 용도엔 충분).
#  - job_id는 프론트엔드가 요청 시작 전에 만들어서 업로드와 함께 보내고, 그 값으로 폴링한다.
#  - 오래된 항목이 계속 쌓이지 않도록, 조회/기록 시마다 일정 시간(_PROGRESS_TTL_SEC)이 지난
#    항목은 같이 청소한다.
# ─────────────────────────────────────────────
_PROGRESS_STORE = {}
_PROGRESS_LOCK = threading.Lock()
_PROGRESS_TTL_SEC = 60 * 30  # 30분 이상 조회 없는 job은 정리


def _progress_cleanup_locked():
    cutoff = time.time() - _PROGRESS_TTL_SEC
    stale = [jid for jid, v in _PROGRESS_STORE.items() if v.get("_updated_at", 0) < cutoff]
    for jid in stale:
        _PROGRESS_STORE.pop(jid, None)


def _progress_set(job_id, stage, current, total, label="", stage_index=1, total_stages=1):
    """job_id가 없으면(프론트가 진행률 표시를 요청하지 않은 옛 클라이언트 등) 조용히 무시한다.
    stage_index/total_stages는 "건축분석→입면검토→구조추출" 같은 여러 단계 중 지금 몇 번째
    단계인지를 나타내고, current/total은 그 단계 안에서의 진행(구조추출이면 배치 번호)이다.
    프론트엔드는 두 값을 합쳐 전체 퍼센티지를 계산한다."""
    if not job_id:
        return
    with _PROGRESS_LOCK:
        _PROGRESS_STORE[job_id] = {
            "stage": stage, "current": current, "total": max(total, 1), "label": label,
            "stage_index": stage_index, "total_stages": max(total_stages, 1),
            "_updated_at": time.time(),
        }
        _progress_cleanup_locked()


def _progress_clear(job_id):
    if not job_id:
        return
    with _PROGRESS_LOCK:
        _PROGRESS_STORE.pop(job_id, None)


def _progress_get(job_id):
    with _PROGRESS_LOCK:
        _progress_cleanup_locked()
        return _PROGRESS_STORE.get(job_id)


# ─────────────────────────────────────────────
#  결과 저장소: api_run_quantity는 이제 Gemini 작업을 백그라운드 스레드로 던져놓고
#  즉시 응답한다(진행률 폴링 저장소와 같은 이유 — 응답을 안 보내고 몇 분~몇십 분씩
#  연결을 붙잡고 있으면, 그 사이 브라우저/OS/공유기 어딘가의 유휴 커넥션 타임아웃에
#  걸려 "네트워크 오류: Load failed"로 끊길 수 있다. 이때 서버는 클라이언트가
#  끊긴 걸 알 방법이 없어 Gemini 호출을 계속 실행하므로, 돈은 나가는데 사용자는
#  결과를 영영 못 받는 상황이 생긴다 — 실제로 재현된 문제다).
#  백그라운드 스레드가 끝나면 최종 결과(성공/실패 모두)를 여기에 job_id로 저장해두고,
#  프론트엔드는 기존에 쓰던 진행률 폴링(api_quantity_progress)을 그대로 계속 돌리다가
#  done=true가 오면 같이 담겨온 결과를 쓴다. 폴링 연결이 중간에 한두 번 끊겨도 다음
#  폴링에서 다시 시도하면 그만이라, 몇 초짜리 요청이 반복되는 구조가 몇십 분짜리
#  요청 하나보다 훨씬 안전하다.
# ─────────────────────────────────────────────
_RESULT_STORE = {}
_RESULT_LOCK = threading.Lock()
_RESULT_TTL_SEC = 60 * 60  # 1시간 — 사용자가 자리를 비웠다 돌아와도 결과를 받을 수 있게 넉넉히


def _result_cleanup_locked():
    cutoff = time.time() - _RESULT_TTL_SEC
    stale = [jid for jid, v in _RESULT_STORE.items() if v.get("_created_at", 0) < cutoff]
    for jid in stale:
        _RESULT_STORE.pop(jid, None)


def _result_set(job_id, payload):
    if not job_id:
        return
    with _RESULT_LOCK:
        stored = dict(payload)
        stored["_created_at"] = time.time()
        _RESULT_STORE[job_id] = stored
        _result_cleanup_locked()


def _result_get(job_id):
    if not job_id:
        return None
    with _RESULT_LOCK:
        _result_cleanup_locked()
        return _RESULT_STORE.get(job_id)


# ─────────────────────────────────────────────
#  취소 요청: 사용자가 분석 도중 "취소" 버튼을 누르면, 그 job_id를 이 집합에 넣어둔다.
#  api_run_quantity가 각 단계(건축분석/입면검토/구조부재 배치추출)를 시작하기 직전에
#  이 집합을 확인해서, 이미 취소 요청이 온 경우 그 단계부터는 건너뛰고 지금까지
#  모인 결과만 반환한다. 이미 Gemini에 보낸(진행 중인) 배치 1개는 어차피 끝까지
#  가야 하지만, 아직 시작 안 한 나머지 배치는 호출 자체를 안 해서 비용을 아낀다.
#  진행률 저장소와 동일하게 로컬 단일 프로세스 개발 서버 기준의 단순 구현이다.
# ─────────────────────────────────────────────
_CANCEL_STORE = set()
_CANCEL_LOCK = threading.Lock()


def _cancel_request(job_id):
    if not job_id:
        return
    with _CANCEL_LOCK:
        _CANCEL_STORE.add(job_id)


def _is_cancelled(job_id):
    if not job_id:
        return False
    with _CANCEL_LOCK:
        return job_id in _CANCEL_STORE


def _cancel_clear(job_id):
    if not job_id:
        return
    with _CANCEL_LOCK:
        _CANCEL_STORE.discard(job_id)


# ─────────────────────────────────────────────
#  도면 확인(검토) 단계: 구조 부재 추출(Gemini)까지만 끝내고 실제 물량 계산은 바로
#  하지 않는다 — 대신 추출된 members를 job_id로 임시 저장해두고, 채팅창에 "도면 확인"
#  버튼을 띄운다. 사용자가 색칠된 도면 팝업을 보고 틀린 부재를 체크/수정한 뒤 "진행"을
#  누르면, 그 수정사항(corrections)과 함께 job_id로 이 저장소에서 members를 꺼내와
#  보정한 다음에야 실제 물량 계산(compute_structural_quantities)을 실행한다.
#  진행률/취소 저장소와 동일하게 로컬 단일 프로세스 기준의 단순 인메모리 구현이다.
#  리뷰는 사람이 직접 보고 판단하는 단계라 진행률 저장소(30분)보다 TTL을 넉넉하게 둔다.
# ─────────────────────────────────────────────
_EXTRACTION_STORE = {}
_EXTRACTION_LOCK = threading.Lock()
_EXTRACTION_TTL_SEC = 60 * 60 * 2  # 2시간 — 도면 검토는 시간이 걸릴 수 있어 넉넉하게


def _extraction_cleanup_locked():
    cutoff = time.time() - _EXTRACTION_TTL_SEC
    stale = [jid for jid, v in _EXTRACTION_STORE.items() if v.get("_created_at", 0) < cutoff]
    for jid in stale:
        _EXTRACTION_STORE.pop(jid, None)


def _extraction_store_set(job_id, members, elevation_data):
    if not job_id:
        return
    with _EXTRACTION_LOCK:
        _EXTRACTION_STORE[job_id] = {
            "members": members, "elevation_data": elevation_data, "_created_at": time.time(),
        }
        _extraction_cleanup_locked()


def _extraction_store_pop(job_id):
    if not job_id:
        return None
    with _EXTRACTION_LOCK:
        _extraction_cleanup_locked()
        return _EXTRACTION_STORE.pop(job_id, None)


def _extraction_store_get(job_id):
    """pop과 달리 저장소에서 지우지 않고 읽기만 한다. 확정 계산(api_quantity_confirm_review)이
    이 값으로 계산을 시도했다가 실패하면 그대로 재시도할 수 있어야 하므로(Gemini 추출을
    처음부터 다시 하지 않아도 되게), 계산이 실제로 성공했을 때만 호출부가 별도로
    _extraction_store_pop을 불러서 지우게 한다."""
    if not job_id:
        return None
    with _EXTRACTION_LOCK:
        _extraction_cleanup_locked()
        return _EXTRACTION_STORE.get(job_id)


# ─────────────────────────────────────────────
#  확인 절차 상태머신 (review_id) — 지피티 독립 검토에서 지적된 실제 구멍을 막기 위해
#  추가했다: api_run_quantity가 confirmed_general_spec 없이도 바로 본 산출을 시작할 수
#  있었고, 있어도 브라우저가 보내는 JSON을 서버가 그대로 신뢰했다 — API를 직접 두드리면
#  개요/구조일반사항 확인을 생략하거나 임의의 구조일반사항을 주입할 수 있었다는 뜻이다.
#
#  이제는 프론트가 흐름을 시작할 때 review_id 하나를 만들어 모든 단계(개요확인→
#  개요확정→구조일반사항확인→구조일반사항확정→지하주차장평면도확인→확정→본추출→
#  부재검토→계산) 내내 재사용하고, 서버가 이 review_id로 "지금까지 실제로 어디까지
#  확정됐는지"를 기억한다. 각 단계 엔드포인트는 자기 앞 단계가 서버 기록상 확정돼
#  있는지 확인한 뒤에만 진행하고, 아니면 409를 돌려준다 — 클라이언트가 무슨 값을
#  같이 보내든 서버 기록이 이긴다(general_spec도 서버가 들고 있는 값을 쓰지, 요청에
#  실려온 값을 신뢰하지 않는다).
#
#  progress 폴링용 job_id(_PROGRESS_STORE/_RESULT_STORE 키)와는 별개의 개념이다 —
#  job_id는 호출마다 새로 만들어도 되는 일회성 폴링 토큰이고, review_id는 세션
#  전체에서 하나만 만들어 계속 재사용하는 "이 확인 절차가 지금 어디까지 왔는지"의
#  단일 진실 공급원이다.
# ─────────────────────────────────────────────
_REVIEW_STORE = {}
_REVIEW_LOCK = threading.Lock()
_REVIEW_TTL_SEC = 60 * 60 * 3  # 3시간 — 개요/구조일반사항/지하주차장 확인을 여러 번 왕복해도 넉넉하게

REVIEW_STAGE_ORDER = ["overview", "general_spec", "drawing_coordination", "basement_plan"]
MEMBER_REBAR_CHECK_KEYS = (
    "columns", "walls", "beams", "slabs", "foundations", "parking",
)


def _empty_member_rebar_check_state():
    """4단계 부재별 판독의 독립 캐시 슬롯. 실제 판독 endpoint는 후속 작업에서 연결한다."""
    return {
        key: {"status": "locked", "result": None, "error": None, "updated_at": None}
        for key in MEMBER_REBAR_CHECK_KEYS
    }


def _sha256_bytes(value):
    return hashlib.sha256(value or b"").hexdigest() if value is not None else None


def _review_file_hashes(structural_pdf_bytes=None, architectural_pdf_bytes=None,
                        structural_zip_bytes=None, architectural_zip_bytes=None):
    return {
        "structural_pdf": _sha256_bytes(structural_pdf_bytes),
        "architectural_pdf": _sha256_bytes(architectural_pdf_bytes),
        "structural_zip": _sha256_bytes(structural_zip_bytes),
        "architectural_zip": _sha256_bytes(architectural_zip_bytes),
    }


def _canonical_review_id(user_id, file_hashes):
    """로그인 사용자와 업로드 파일 내용에 종속된, 클라이언트가 위조할 수 없는 상태 키."""
    material = json.dumps(
        {"user_id": str(user_id), "files": file_hashes},
        sort_keys=True, separators=(",", ":"),
    ).encode("utf-8")
    return "qtyrev-" + hashlib.sha256(material).hexdigest()


def _matching_uploaded_files(record, current_hashes):
    expected = (record or {}).get("_file_hashes") or {}
    for key, current in current_hashes.items():
        if current is not None and expected.get(key) != current:
            return False
    return True


def _review_cleanup_locked():
    cutoff = time.time() - _REVIEW_TTL_SEC
    stale = [rid for rid, v in _REVIEW_STORE.items() if v.get("_updated_at", 0) < cutoff]
    for rid in stale:
        _REVIEW_STORE.pop(rid, None)


def _review_ensure(review_id, user_id=None, file_hashes=None):
    """review_id에 해당하는 상태 기록이 없으면 새로 만들고, 있으면 그대로 반환한다."""
    with _REVIEW_LOCK:
        _review_cleanup_locked()
        rec = _REVIEW_STORE.get(review_id)
        if rec is None:
            rec = {
                "overview": None, "overview_confirmed": False,
                "overview_page_detection": None,
                "general_spec": None, "general_spec_confirmed": False,
                "drawing_coordination": None, "drawing_coordination_confirmed": False,
                "basement_plan": None, "basement_plan_confirmed": False,
                "member_rebar_checks": _empty_member_rebar_check_state(),
                "extraction_started": False,
                "_user_id": str(user_id) if user_id is not None else None,
                "_file_hashes": dict(file_hashes or {}),
                "_created_at": time.time(), "_updated_at": time.time(),
            }
            _REVIEW_STORE[review_id] = rec
        else:
            rec.setdefault("member_rebar_checks", _empty_member_rebar_check_state())
        return rec


def _review_get(review_id):
    if not review_id:
        return None
    with _REVIEW_LOCK:
        _review_cleanup_locked()
        return _REVIEW_STORE.get(review_id)


def _review_update(review_id, **fields):
    """review_id 기록에 필드를 병합한다. 기록이 없으면(만료됐거나 애초에 없으면) 아무것도
    하지 않고 False를 반환한다 — 호출부가 "세션을 찾을 수 없음"으로 처리해야 한다."""
    with _REVIEW_LOCK:
        rec = _REVIEW_STORE.get(review_id)
        if rec is None:
            return False
        rec.update(fields)
        rec["_updated_at"] = time.time()
        return True


def _review_require_stage(review_id, required_confirmed_field, stage_label):
    """review_id가 required_confirmed_field(예: "general_spec_confirmed")까지 확정된
    상태인지 확인한다. 통과하면 (record, None)을, 실패하면 (None, JsonResponse) 를
    반환한다 — 호출부는 두 번째 값이 있으면 그대로 return하면 된다."""
    rec = _review_get(review_id)
    if rec is None:
        return None, JsonResponse({
            "error": "확인 절차 세션을 찾을 수 없습니다 — 시간이 너무 지났거나 잘못된 review_id입니다. "
                     "개요 확인부터 다시 진행해 주세요.",
        }, status=404)
    if not rec.get(required_confirmed_field):
        return None, JsonResponse({
            "error": f"{stage_label} 확정이 되지 않았습니다 — 이전 단계를 먼저 확인/확정해 주세요.",
        }, status=409)
    return rec, None


def _review_reset_confirmations_from(review_id, stage):
    """stage 및 그 이후 단계(REVIEW_STAGE_ORDER 기준)의 *_confirmed 플래그를 전부 False로
    되돌린다. 지피티 독립 검토에서 지적된 문제: 개요/구조일반사항을 다시 확인하거나
    "아니요, 수정할게요"로 재검토한 뒤에도 이미 확정(confirmed=True)돼 있던 플래그가
    그대로 남아있었다 — 새로 받아온 데이터가 이전에 사람이 눈으로 보고 확정한 데이터와
    다를 수 있는데도, 서버 기록상으로는 여전히 "확정됨"으로 남아 있었다는 뜻이다.
    이제 overview-check/overview-revise/basement-plan-check가 데이터를 새로 쓸 때마다
    그 단계부터 이후 모든 단계의 확정을 무효화해서, 사용자가 매번 새 데이터를 다시
    "예, 맞습니다"로 확인해야만 본 추출까지 갈 수 있게 한다."""
    try:
        idx = REVIEW_STAGE_ORDER.index(stage)
    except ValueError:
        return
    reset_fields = {f"{s}_confirmed": False for s in REVIEW_STAGE_ORDER[idx:]}
    _review_update(review_id, **reset_fields)


# 체크리스트에 보여주고, 사용자가 고칠 수 있게 허용할 카테고리별 핵심 필드.
# 처음엔 치수/개수/철근규격만 추렸었는데, 정작 철근량 계산에 제일 크게 영향을 주는
# rebar_spacing_m(간격)·main_rebar_count(개수) 등이 빠져있으면 검토 화면에서 고쳐봤자
# 계산에 반영이 안 되는 항목이 생겨서 — 각 카테고리 스키마의 실제 계산용 필드를 전부 포함시킨다.
# (mark/zone/section/bbox/openings처럼 목록형이거나 식별용인 필드만 편집 대상에서 제외)
_MEMBER_SUMMARY_FIELDS = {
    "foundations": [
        "length_m", "width_m", "thickness_m", "count",
        "rebar_size", "rebar_spacing_m",
        "dowel_bar_size", "dowel_bar_count", "dowel_has_hook",
    ],
    "columns": [
        "width_m", "depth_m", "height_m", "count",
        "main_rebar_size", "main_rebar_count",
        "tie_rebar_size", "tie_spacing_m", "has_hook",
    ],
    "beams": [
        "width_m", "depth_m", "length_m", "count",
        "main_rebar_size", "main_rebar_count",
        "stirrup_size", "stirrup_spacing_m", "has_hook", "is_top_bar",
    ],
    "slabs": [
        "area_m2", "thickness_m", "count",
        "rebar_size", "rebar_spacing_m", "has_hook", "is_top_bar", "is_deck_slab",
    ],
    "walls": [
        "length_m", "height_m", "thickness_m", "count",
        "rebar_size", "rebar_spacing_m", "has_hook", "end_condition", "is_single_face",
    ],
    "stairs": [
        "width_m", "length_m", "thickness_m", "count",
        "rebar_size", "rebar_spacing_m",
        "distribution_rebar_size", "distribution_rebar_spacing_m",
        "is_top_bar", "has_hook",
    ],
}

# 위 필드 중 참/거짓 값인 것들 — 검토 팝업에서 텍스트 입력 대신 체크박스로 보여주기 위한 목록.
_MEMBER_BOOLEAN_FIELDS = {
    "has_hook", "is_top_bar", "is_deck_slab", "dowel_has_hook", "is_single_face",
}

# extract_structural_members()가 남기는 notes 중, "이 결과는 온전하지 않다"는 신호로 볼 수 있는
# 문구들 — 이런 note가 하나라도 있으면 검토 팝업에서 "일부만 인식됨"을 눈에 띄게 알려야 한다.
# 그냥 notes 목록 맨 아래 묻혀 있으면 사용자가 놓치고 "확정" 취급해버릴 위험이 크다(실제로
# 지적받은 문제 — Gemini 응답이 잘리거나 배치가 실패해도 지금까지는 조용히 부분 결과로
# 진행할 수 있었다).
_INCOMPLETE_EXTRACTION_MARKERS = [
    "도중에 잘려", "렌더링 중 오류가 발생했습니다", "취소를 요청하여", "JSON 파싱 오류",
]


def _extraction_incomplete_reasons(members):
    """members["notes"]를 훑어서 추출이 불완전했다고 볼 수 있는 사유만 골라 돌려준다.
    빈 리스트면 완전하다고 간주한다(마커에 해당하는 note가 하나도 없었다는 뜻)."""
    reasons = []
    for note in members.get("notes") or []:
        if not isinstance(note, str):
            continue
        if any(marker in note for marker in _INCOMPLETE_EXTRACTION_MARKERS):
            reasons.append(note)
    return reasons


def _build_review_checklist(members):
    """members 딕셔너리를 프론트엔드 검토 팝업의 왼쪽 체크리스트용 평평한 리스트로 변환한다.
    각 항목은 review_id(카테고리:인덱스)로 식별되며, 이 review_id는 나중에 corrections에서
    "이 항목을 지워라/이 필드를 이 값으로 바꿔라"를 지정할 때 그대로 사용된다."""
    checklist = []
    for cat_key, label in _CATEGORY_KEY_TO_LABEL.items():
        for i, it in enumerate(members.get(cat_key) or []):
            if not isinstance(it, dict):
                continue
            fields = {}
            for f in _MEMBER_SUMMARY_FIELDS.get(cat_key, []):
                if f in it:
                    fields[f] = it.get(f)
            bbox = it.get("bbox") if isinstance(it.get("bbox"), dict) else None
            rebar_layers = it.get("rebar_layers")
            checklist.append({
                "review_id": f"{cat_key}:{i}",
                "category": label,
                "mark": it.get("mark") or "",
                "zone": it.get("zone") or "",
                "section": it.get("section") or "",
                "fields": fields,
                # 필드 중 어떤 게 참/거짓 값인지 알려줘서, 프론트가 그 필드만 텍스트 입력 대신
                # 체크박스로 그릴 수 있게 한다 (has_hook 같은 걸 "true"라고 글자로 치게 하면 실수하기 쉬움).
                "bool_fields": sorted(f for f in fields if f in _MEMBER_BOOLEAN_FIELDS),
                "bbox_page": bbox.get("page") if bbox else None,
                "dedup_flag": it.get("_dedup_flag") or None,
                # 세분화 배근(rebar_layers)은 role/size/spacing_m/count 등을 담은 리스트라
                # 위 fields(스칼라 값 전용)와 형태가 달라 따로 내려준다 — 프론트는 이걸
                # 읽기전용 요약 + 편집용 JSON 텍스트영역으로 보여준다.
                "rebar_layers": rebar_layers if isinstance(rebar_layers, list) else None,
            })
    return checklist


_VALID_REBAR_SIZES = set(REBAR_UNIT_WEIGHT.keys())
_VALID_WALL_END_CONDITIONS = {"일자형", "T자형", "모서리"}
_MAX_REASONABLE_NUMERIC = 100000  # 오타로 0을 몇 개 더 붙인 값 등을 걸러내는 상한 — 실제 치수/간격/개수 범위를 크게 벗어남


def _validate_correction_field(field, value):
    """검토 팝업에서 들어온 수정값 1개가 계산에 안전하게 들어가도 되는 값인지 확인한다.
    유효하면 (True, 정제된 값), 무효하면 (False, 사유 문자열)을 돌려준다.

    이게 없으면(원래 없었음) 수정 팝업에 count=-1 같은 값을 넣었을 때 그대로 계산에
    들어가 콘크리트/거푸집 물량이 음수로 나오는 문제가 실제로 있었다 — 치수/간격/개수
    필드는 양수인지, 철근 규격은 실제 존재하는 규격인지 정도는 최소한으로 걸러낸다."""
    if field in _MEMBER_BOOLEAN_FIELDS:
        return True, bool(value)

    if field.endswith("_size"):
        if value in _VALID_REBAR_SIZES:
            return True, value
        return False, f"{field}={value!r}: 허용되지 않는 철근 규격입니다"

    if field == "end_condition":
        if not value or value in _VALID_WALL_END_CONDITIONS:
            return True, value
        return False, f"end_condition={value!r}: 일자형/T자형/모서리 중 하나여야 합니다"

    # 나머지(길이/두께/간격/개수 등)는 전부 숫자 필드 — 음수·NaN·비정상적으로 큰 값을 막는다.
    try:
        num = float(value)
    except (TypeError, ValueError):
        return False, f"{field}={value!r}: 숫자가 아닙니다"
    if not math.isfinite(num) or num <= 0:
        return False, f"{field}={value!r}: 0보다 큰 숫자여야 합니다"
    if num > _MAX_REASONABLE_NUMERIC:
        return False, f"{field}={value!r}: 비정상적으로 큰 값입니다"
    if field == "count" or field.endswith("_count"):
        if num != int(num):
            return False, f"{field}={value!r}: 정수(개수)여야 합니다"
        return True, int(num)
    return True, num


_VALID_REBAR_LAYER_ROLES = {
    "주근", "후프", "타이", "스터럽", "수직근", "수평근",
    "단부보강근", "모서리보강근", "교차부보강근", "개구부보강근", "배력근",
}


def _validate_rebar_layers_correction(value):
    """검토 팝업에서 세분화 배근(rebar_layers) 전체를 사용자가 JSON으로 고쳐 넣었을 때
    계산에 안전하게 들어가도 되는 형태인지 검증한다. 유효하면 (True, 정제된 리스트),
    무효하면 (False, 사유 문자열)을 돌려준다.

    rebar_layers는 스칼라 값이 아니라 role/size/spacing_m/count 등을 담은 딕셔너리의
    리스트라 _validate_correction_field(숫자·철근규격·참거짓만 다룸)로는 검증할 수 없어서
    따로 만들었다. 항목 하나하나의 최종 방어는 quantity_calc._valid_rebar_layers가 실제
    계산 시점에도 한 번 더 하지만(그 함수는 안전을 위해 조용히 걸러내고 경고만 남김),
    검토 화면에서 미리 걸러 알려줘야 사용자가 "저장했는데 왜 반영이 안 되지"라고
    헷갈리지 않는다."""
    if value in (None, ""):
        return True, []
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (TypeError, ValueError):
            return False, "세분화배근: JSON 형식이 아닙니다"
    if value == []:
        return True, []
    if not isinstance(value, list):
        return False, "세분화배근: 목록(JSON 배열) 형태여야 합니다"
    cleaned = []
    for i, layer in enumerate(value):
        if not isinstance(layer, dict):
            return False, f"세분화배근[{i}]: 각 항목은 객체({{...}}) 형태여야 합니다"
        role = layer.get("role")
        if not role or role not in _VALID_REBAR_LAYER_ROLES:
            return False, f"세분화배근[{i}]: role={role!r}은 허용되지 않는 배근 종류입니다"
        size = layer.get("size")
        if size is not None and size not in _VALID_REBAR_SIZES:
            return False, f"세분화배근[{i}]: size={size!r}은 허용되지 않는 철근 규격입니다"
        for numf in ("spacing_m", "zone_length_m"):
            v = layer.get(numf)
            if v is not None:
                try:
                    v = float(v)
                except (TypeError, ValueError):
                    return False, f"세분화배근[{i}].{numf}={layer.get(numf)!r}: 숫자가 아닙니다"
                if not math.isfinite(v) or v <= 0:
                    return False, f"세분화배근[{i}].{numf}={layer.get(numf)!r}: 0보다 큰 숫자여야 합니다"
                layer[numf] = v
        if layer.get("count") is not None:
            try:
                c = float(layer.get("count"))
            except (TypeError, ValueError):
                return False, f"세분화배근[{i}].count={layer.get('count')!r}: 숫자가 아닙니다"
            if not math.isfinite(c) or c <= 0 or c != int(c):
                return False, f"세분화배근[{i}].count={layer.get('count')!r}: 0보다 큰 정수(개수)여야 합니다"
            layer["count"] = int(c)
        cleaned.append(layer)
    return True, cleaned


def _sanitize_raw_members(members):
    """Gemini가 추출한 원본 부재 데이터를 실제 계산(compute_structural_quantities)에
    넣기 전에 마지막으로 한 번 더 검증한다.

    검토 팝업에서 사용자가 직접 고친 필드는 _apply_member_corrections가
    _validate_correction_field로 이미 검증하지만, 사용자가 건드리지 않은 나머지 필드
    (실무에서는 대부분의 필드가 여기 해당한다)는 Gemini 원본값이 그대로 계산에
    들어간다 — count=0이 `it.get("count", 1) or 1` 폴백 때문에 조용히 1로 바뀌거나,
    count=-2 같은 값이 그대로 들어가 콘크리트/거푸집/철근량이 통째로 음수로 계산되는
    문제가 실제로 재현/확인됐다. 이 함수는 카테고리별 계산용 필드(_MEMBER_SUMMARY_FIELDS)를
    전부 훑어서 무효한 값(음수/0/NaN/비정수 개수 등)이 있으면 그 필드만 None으로 비운다
    (항목 전체를 지우지는 않는다 — 필드 하나가 이상하다고 나머지 정상 값까지 버릴 필요는
    없어서다). 필드가 None이 되면 quantity_calc.py에 이미 있는 "치수 누락시 계산 제외"
    "철근정보 없으면 정보없음 처리" 로직이 그대로 안전하게 이어받는다."""
    cleaned = json.loads(json.dumps(members))
    rejected_notes = []
    for cat_key, label in _CATEGORY_KEY_TO_LABEL.items():
        for it in cleaned.get(cat_key) or []:
            if not isinstance(it, dict):
                continue
            mark = it.get("mark") or "무명"
            for field in _MEMBER_SUMMARY_FIELDS.get(cat_key, []):
                if field not in it or it[field] is None:
                    continue
                ok, result = _validate_correction_field(field, it[field])
                if ok:
                    it[field] = result
                else:
                    it[field] = None
                    rejected_notes.append(f"{label} {mark}: 원본 추출값 {result} — 계산에 반영하지 않고 비웠습니다.")
    if rejected_notes:
        cleaned["notes"] = list(cleaned.get("notes") or []) + rejected_notes
    return cleaned


def _apply_member_corrections(members, corrections):
    """검토 팝업에서 받은 corrections(제거/수정 목록)를 members에 반영한다.
    review_id 형식은 "카테고리:인덱스"이며, _build_review_checklist가 만든 것과
    반드시 같은 순서/인덱스 기준이어야 한다(그 사이에 members가 바뀌면 안 됨).

    Returns: (corrected_members, rejected_notes) — rejected_notes는 값이 유효하지
    않아 반영하지 않은 수정 건에 대한 설명 리스트다(원래 값은 그대로 유지됨).
    무효한 값을 조용히 버리기만 하면 사용자가 "수정했는데 왜 그대로지"라고 헷갈릴
    수 있어서, 호출부(api_quantity_confirm_review)가 이 리스트를 결과 warnings에 같이 붙인다.

    사용자가 고치지 않은 나머지 필드(Gemini 원본값)는 마지막에 _sanitize_raw_members로
    한 번 더 걸러서 반환한다 — 그래서 이 함수의 반환값은 "사용자 수정 반영 + 원본값
    안전성 검증"까지 끝난, 곧바로 계산에 넣어도 되는 상태다."""
    corrected = json.loads(json.dumps(members))  # 얕은 복사로는 중첩 dict가 공유돼서 깊은 복사 사용
    removals = set()
    edits = {}
    for c in (corrections or []):
        if not isinstance(c, dict):
            continue
        rid = c.get("review_id")
        if not rid:
            continue
        action = c.get("action")
        if action == "remove":
            removals.add(rid)
        elif action == "edit":
            edits[rid] = c.get("fields") or {}

    rejected_notes = []
    for cat_key in _CATEGORY_KEY_TO_LABEL.keys():
        items = corrected.get(cat_key) or []
        new_items = []
        for i, it in enumerate(items):
            rid = f"{cat_key}:{i}"
            if rid in removals:
                continue
            if rid in edits and isinstance(it, dict):
                allowed = set(_MEMBER_SUMMARY_FIELDS.get(cat_key, []))
                label = _CATEGORY_KEY_TO_LABEL.get(cat_key, cat_key)
                mark = it.get("mark") or "무명"
                for k, v in edits[rid].items():
                    if k == "rebar_layers":
                        # rebar_layers는 스칼라 값이 아니라 리스트라 _MEMBER_SUMMARY_FIELDS
                        # (allowed)에 들어있지 않다 — 여기서 따로 처리한다.
                        ok, result = _validate_rebar_layers_correction(v)
                        if ok:
                            if result:
                                it["rebar_layers"] = result
                            else:
                                it.pop("rebar_layers", None)
                        else:
                            rejected_notes.append(f"{label} {mark}: {result} — 이 수정은 반영하지 않고 기존 값을 유지했습니다.")
                        continue
                    if k not in allowed:
                        continue
                    ok, result = _validate_correction_field(k, v)
                    if ok:
                        it[k] = result
                    else:
                        rejected_notes.append(f"{label} {mark}: {result} — 이 수정은 반영하지 않고 기존 값을 유지했습니다.")
            new_items.append(it)
        corrected[cat_key] = new_items

    # 사용자가 건드리지 않은 나머지 필드(Gemini 원본값)까지 마지막으로 한 번 더 검증한다.
    sanitized = _sanitize_raw_members(corrected)
    sanitize_notes = sanitized.get("notes") or []
    original_note_count = len(corrected.get("notes") or [])
    # _sanitize_raw_members가 새로 덧붙인 note만 rejected_notes에 합쳐서 반환한다
    # (원래 있던 notes는 corrected에 이미 있으니 중복으로 다시 붙이지 않는다).
    new_sanitize_notes = sanitize_notes[original_note_count:] if len(sanitize_notes) > original_note_count else []
    return sanitized, rejected_notes + new_sanitize_notes


def _extract_text_from_gemini_response(response):
    """core/ai_writer.py의 동명 함수와 동일한 방식으로 응답 텍스트를 안전하게 추출합니다."""
    text = getattr(response, "text", None)
    if text:
        return str(text).strip()

    pieces = []
    for candidate in getattr(response, "candidates", []) or []:
        content = getattr(candidate, "content", None)
        for part in getattr(content, "parts", []) or []:
            part_text = getattr(part, "text", None)
            if part_text:
                pieces.append(str(part_text))
    return "\n".join(piece for piece in pieces if piece).strip()


def _gemini_response_diagnostics(response):
    """Gemini 응답 본문을 저장하지 않고 종료 사유·토큰·후보 길이만 진단한다."""
    candidates = list(getattr(response, "candidates", None) or [])
    candidate_details = []
    for index, candidate in enumerate(candidates):
        content = getattr(candidate, "content", None)
        parts = list(getattr(content, "parts", None) or [])
        text_length = sum(
            len(str(part_text))
            for part in parts
            for part_text in [getattr(part, "text", None)]
            if part_text
        )
        finish_reason = getattr(candidate, "finish_reason", None)
        candidate_details.append({
            "index": index,
            "finish_reason": (
                getattr(finish_reason, "name", None)
                or (str(finish_reason) if finish_reason is not None else None)
            ),
            "part_count": len(parts),
            "text_length": text_length,
        })

    usage = getattr(response, "usage_metadata", None)

    def usage_value(name):
        if usage is None:
            return None
        if isinstance(usage, dict):
            return usage.get(name)
        return getattr(usage, name, None)

    prompt_feedback = getattr(response, "prompt_feedback", None)
    block_reason = getattr(prompt_feedback, "block_reason", None)
    return {
        "candidate_count": len(candidates),
        "candidates": candidate_details,
        "prompt_block_reason": (
            getattr(block_reason, "name", None)
            or (str(block_reason) if block_reason is not None else None)
        ),
        "usage": {
            "prompt_token_count": usage_value("prompt_token_count"),
            "candidates_token_count": usage_value("candidates_token_count"),
            "thoughts_token_count": usage_value("thoughts_token_count"),
            "total_token_count": usage_value("total_token_count"),
        },
    }


def _try_repair_truncated_json(raw_clean):
    """
    Gemini 응답이 max_output_tokens 한도에 걸려 중간에 잘렸을 때, 마지막으로 완전히 끝난
    항목까지만 남기고 열린 배열/객체를 강제로 닫아서 부분 복구를 시도한다.
    (전체 파싱 실패로 데이터를 통째로 버리는 것보다, 끝부분 몇 개 항목만 못 읽더라도
    앞부분 데이터는 건지는 게 낫기 때문 — 실제로 큰 프로젝트에서 이 문제가 발생했음)
    실패하면 None을 반환해서 호출부가 기존처럼 에러 처리하게 한다.
    """
    depth_stack = []
    in_str = False
    escape = False
    safe_cutoffs = []  # (그 지점까지의 인덱스, 그 시점의 열린 괄호 스택)
    for i, ch in enumerate(raw_clean):
        if escape:
            escape = False
            continue
        if ch == "\\" and in_str:
            escape = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch in "{[":
            depth_stack.append(ch)
        elif ch in "}]":
            if depth_stack:
                depth_stack.pop()
            safe_cutoffs.append((i + 1, list(depth_stack)))
        elif ch == "," and depth_stack:
            safe_cutoffs.append((i, list(depth_stack)))

    for cutoff_idx, stack_snapshot in reversed(safe_cutoffs[-200:]):
        candidate = raw_clean[:cutoff_idx].rstrip().rstrip(",")
        closing = "".join("]" if c == "[" else "}" for c in reversed(stack_snapshot))
        attempt = candidate + closing
        try:
            return json.loads(attempt)
        except json.JSONDecodeError:
            continue
    return None


def _code_matches_filename(code: str, filename_upper: str) -> bool:
    """
    도면 코드가 파일명에 존재하는지 확인한다.

    실제 현장 도면은 'S-001~002', 'S-111~130'처럼 여러 도면번호를 구간(~)으로
    묶어 파일명을 짓는 경우가 많다. 단순 substring 매칭("S-002" in filename)은
    "S-001~002"에서 "S-002"를 놓치므로, 구간 표기를 해석해서 코드가 그 구간에
    포함되는지까지 확인한다.
    """
    if code in filename_upper:
        return True

    m = re.match(r"^([A-Z]+)-(\d+)$", code)
    if not m:
        return False
    prefix, num_str = m.group(1), m.group(2)
    num = int(num_str)

    for fm in re.finditer(re.escape(prefix) + r"-(\d+)(?:~(\d+))?", filename_upper):
        low = int(fm.group(1))
        high = int(fm.group(2)) if fm.group(2) else low
        if low <= num <= high:
            return True
    return False


def _normalize_for_match(s):
    """한글 NFC/NFD·전각 기호·공백 차이를 없앤 파일명/텍스트 매칭 전처리."""
    value = unicodedata.normalize("NFKC", str(s or ""))
    value = value.replace("～", "~").replace("∼", "~")
    return re.sub(r"[\s,_()（）\\[\\]{}]+", "", value).upper()


def _filename_matches_keywords(filename_norm, keywords, exclude_keywords=None):
    """정규화된 파일명(filename_norm)에 keywords 중 하나라도 포함되고,
    exclude_keywords는 하나도 포함되지 않으면 True."""
    if not any(_normalize_for_match(kw) in filename_norm for kw in keywords):
        return False
    if exclude_keywords and any(_normalize_for_match(kw) in filename_norm for kw in exclude_keywords):
        return False
    return True


# ─────────────────────────────────────────────
#  필수 구조도면 목록 (도면번호: 도면명)
# ─────────────────────────────────────────────
REQUIRED_STRUCTURAL = {
    "S-001": "구조설계개요 및 시방서",
    "S-002": "사용자재표 (콘크리트/철근 규격)",
    "S-101": "기초평면도",
    "S-102": "기초상세도",
    "S-103": "지하층 골조평면도 (지하주차장 구조 포함)",
    "S-104": "지하외벽 배근도",
    "S-105": "기준층 골조평면도 (반복되는 표준층)",
    "S-106": "지하주차장 램프 평면도/입단면도",
    "S-201": "기둥배근도",
    "S-202": "기둥일람표",
    "S-203": "보배근도",
    "S-204": "슬래브 배근도 (층별)",
    "S-301": "전단벽 배근도",
    "S-302": "계단 배근도",
    "S-401": "지붕층 골조평면도",
    "S-402": "옥탑층 골조평면도",
    "S-501": "부재일람표 요약본",
}

# 2026-07-27 1차 지적: "웬만한거 다있으면 넘어가... 개요 구조일반사항 기초평면도
# 벽체 기둥일람표 주요한거 있으면 그냥 넘어가" — 15개 도면 전부를 엄격히 대조하니
# 실제로는 큰 문제 없는 프로젝트에서도 매번 같은 몇 개(S-203/S-204/S-302 등)가
# "누락"으로 뜨면서 업로드 흐름을 방해했다. 이후 계산에 실제로 꼭 필요한 핵심
# 항목만 "누락 도면" 경고 대상으로 좁혔다.
#
# 2026-07-27 2차 지적: "지하주차장 평면도 각동 각층 평면도 기준층 평면도 램프
# 평입단면도 등 안필요해?? 그런거를 확인하라고" — 첫 좁힌 목록(S-001/002/101/202/301)이
# 실제 물량산출(층별 기둥/보/슬래브/벽체 부재 추출)에 정작 핵심인 지하층/기준층
# 골조평면도와 램프 도면을 빠뜨리고 있었다. 이 세 가지(지하층 골조평면도,
# 기준층 골조평면도, 램프 평입단면도)를 핵심 목록에 추가했었다.
#
# 2026-07-27 3차(가장 중요한) 지적: "처음에 할때 구조평면도 하면 거기 캐드에 다들어있어" +
# 실제 프로젝트(부천 현장)의 구조/건축 폴더 파일 목록을 직접 받아본 결과, 지금까지 쓰던
# S-001~S-501 코드 체계는 전부 내가 임의로 지어낸 것이었고 실제 설계사무소 도면번호
# 규칙과 전혀 안 맞았다 (이 프로젝트는 예: 구조일반사항=S-011~022, 동기초 구조평면도=
# S-201~202, 기둥일람표=S-301, 벽체일람표=S-311~324, 지하주차장 구조평면도=S-401~403,
# 동(기준층 포함) 구조평면도=S-111~130 사용 — 내가 가정한 번호와 겹치는 게 거의 없었다).
# 도면번호 규칙은 설계사무소마다 다르고 업계 표준이 없어서, 코드 기반 매칭 자체가
# 근본적으로 신뢰할 수 없는 방식이었다. 그래서 도면번호가 아니라 파일명에 실제로 적힌
# 한글 도면명(예: "기초", "지하주차장", "기둥일람표")으로 매칭하는 방식으로 전면 교체한다
# — 이미 낱장 PDF 파일명 힌트(_classify_filename_hint)에서 검증된 것과 같은 접근이다.
# 주의: "각동 각층" 전체(동별로 기준층이 다 다른 경우 각 동마다 몇 장씩)까지는 파일명
# 규칙만으로 일반화해서 자동 판별하기 어렵다 — 최소 그 종류의 도면이 한 장이라도
# 있는지는 확인하지만, 동별 층별 전수 커버리지까지 보장하진 못한다는 한계는 있다
# (사용자에게 그대로 알려야 함).
_STRUCTURAL_CRITICAL_ITEMS = [
    {
        "key": "general_spec",
        "name": "구조일반사항 / 구조설계개요",
        "filename_keywords": ["구조일반사항", "구조설계개요", "설계개요"],
        "content_keywords": ["구조설계개요", "설계기준", "구조일반사항", "내진설계", "허용지내력", "설계하중"],
    },
    {
        "key": "foundation_plan",
        "name": "기초(동기초) 구조평면도",
        "filename_keywords": ["기초"],
        "content_keywords": ["기초", "FOOTING", "PILE", "매트기초", "독립기초", "줄기초"],
    },
    {
        "key": "basement_parking_plan",
        "name": "지하주차장 구조평면도",
        "filename_keywords": ["지하주차장", "지하층"],
        "content_keywords": ["지하주차장", "골조평면도", "주차", "지하"],
    },
    {
        "key": "typical_floor_plan",
        "name": "동(기준층 포함) 구조평면도",
        "filename_keywords": ["구조평면도", "골조평면도"],
        "exclude_keywords": ["기초", "지하주차장", "지하층"],
        "content_keywords": ["구조평면도", "골조평면도", "기준층"],
    },
    {
        "key": "column_schedule",
        "name": "기둥 일람표",
        "filename_keywords": ["기둥일람표"],
        "content_keywords": ["기둥", "COLUMN", "일람표"],
    },
    {
        "key": "wall_schedule",
        "name": "벽체(전단벽) 일람표",
        "filename_keywords": ["벽체일람표", "전단벽일람표", "전단벽"],
        "content_keywords": ["전단벽", "내력벽", "WALL", "벽체", "일람표"],
    },
]

# 구버전 코드에서 참조하던 이름 — 항목 개수/설명 문구에 쓰기 위해 남겨둔다.
REQUIRED_STRUCTURAL_CRITICAL = [item["key"] for item in _STRUCTURAL_CRITICAL_ITEMS]

REQUIRED_ARCHITECTURAL = {
    "A-000": "건축 일반사항 / 범례",
    "A-001": "배치도",
    "A-101": "1층 평면도",
    "A-102": "2층 평면도",
    "A-103": "지붕층 평면도",
    "A-201": "입면도 (정면/배면/측면)",
    "A-301": "단면도",
    "A-401": "창호도",
    "A-501": "내부마감표",
    "A-601": "계단 상세도",
}

CAD_PRECHECK_STRUCTURAL_ITEMS = [
    {"key": "general_spec", "name": "구조일반사항 / 구조설계개요",
     "filename_any": ["구조일반사항", "구조설계개요", "설계개요"],
     "content_any": ["구조일반사항", "구조설계개요", "설계기준", "내진설계"]},
    {"key": "building_structure_plan", "name": "동·단위세대·기준층 구조평면도",
     "filename_any": ["구조평면도", "골조평면도"],
     "filename_exclude": ["동기초", "주차장", "기초"],
     "content_any": ["구조평면도", "골조평면도", "기준층", "단위세대"]},
    {"key": "building_foundation_plan", "name": "동기초 구조평면도",
     "filename_all_groups": [["기초"], ["동", "주동", "아파트"]],
     "filename_exclude": ["주차장"],
     "content_any": ["기초", "FOUNDATION", "FOOTING"]},
    {"key": "parking_foundation_plan", "name": "지하주차장 기초 구조평면도",
     "filename_all_groups": [["주차장"], ["기초"]],
     "content_any": ["주차장", "기초", "FOUNDATION", "FOOTING"]},
    {"key": "parking_structure_plan", "name": "지하주차장 구조평면도",
     "filename_all_groups": [["주차장"], ["구조평면도", "골조평면도"]],
     "filename_exclude": ["기초"],
     "content_any": ["지하주차장", "구조평면도", "골조평면도"]},
    {"key": "building_column_schedule", "name": "주동 기둥일람표",
     "filename_all_groups": [["기둥"], ["일람표", "배근도"], ["아파트", "주동", "단위세대", "동"]],
     "filename_exclude": ["주차장"],
     "content_any": ["기둥", "COLUMN", "일람표"]},
    {"key": "parking_column_schedule", "name": "지하주차장 기둥일람표",
     "filename_all_groups": [["주차장"], ["기둥"], ["일람표", "배근도"]],
     "content_any": ["주차장", "기둥", "COLUMN", "일람표"]},
    {"key": "building_wall_schedule", "name": "주동 벽체·전단벽·지하외벽 일람표/배근도",
     "filename_all_groups": [["벽체", "전단벽", "지하외벽"], ["일람표", "배근도"]],
     "filename_exclude": ["주차장"],
     "content_any": ["벽체", "전단벽", "지하외벽", "WALL"]},
    {"key": "parking_wall_schedule", "name": "지하주차장 벽체·지하외벽 일람표/배근도",
     "filename_all_groups": [["주차장"], ["벽체", "전단벽", "지하외벽"], ["일람표", "배근도"]],
     "content_any": ["주차장", "벽체", "지하외벽", "WALL"]},
    {"key": "building_beam_schedule", "name": "주동 보 일람표",
     "filename_all_groups": [["보일람표", "보 일람표", "보배근도", "보 배근도"],
                             ["아파트", "주동", "단위세대", "동"]],
     "filename_exclude": ["주차장"],
     "content_any": ["보일람표", "보 일람표", "BEAM", "배근"]},
    {"key": "parking_beam_schedule", "name": "지하주차장 보 일람표",
     "filename_all_groups": [["주차장"], ["보일람표", "보 일람표", "보배근도", "보 배근도"]],
     "content_any": ["주차장", "보일람표", "BEAM", "배근"]},
    {"key": "building_slab_rebar", "name": "주동 슬래브 배근도",
     "filename_all_groups": [["슬래브"], ["배근도", "일람표"], ["아파트", "주동", "단위세대", "동"]],
     "filename_exclude": ["주차장"],
     "content_any": ["슬래브", "SLAB", "배근"]},
    {"key": "parking_slab_rebar", "name": "지하주차장 슬래브 배근도",
     "filename_all_groups": [["주차장"], ["슬래브"], ["배근도", "일람표"]],
     "content_any": ["주차장", "슬래브", "SLAB", "배근"]},
]

CAD_PRECHECK_ARCHITECTURAL_ITEMS = [
    {"key": "building_plans", "name": "주동 평면도",
     "filename_all_groups": [["평면도"], ["주동", "동평면"]],
     "filename_exclude": ["구조", "주차장", "코아", "코어", "부대시설"],
     "content_any": ["주동", "동평면", "FLOOR PLAN"]},
    {"key": "unit_plans", "name": "단위세대 평면도",
     "filename_all_groups": [["평면도"], ["단위세대"]],
     "filename_exclude": ["구조", "주차장", "코아", "코어", "부대시설"],
     "content_any": ["단위세대", "평면도", "FLOOR PLAN"]},
    {"key": "building_elevations", "name": "주동 입면도",
     "filename_any": ["동입면도", "주동입면도", "주동 입면도"],
     "content_any": ["입면도", "ELEVATION"]},
    {"key": "building_sections", "name": "주동 단면도",
     "filename_all_groups": [["단면도"], ["주동", "동단면"]],
     "filename_exclude": ["코아", "코어", "주차장", "부대시설"],
     "content_any": ["단면도", "SECTION"]},
    {"key": "core_plans", "name": "코어 확대평면도",
     "filename_all_groups": [["코아", "코어"], ["평면도"]],
     "content_any": ["코아", "코어", "CORE", "평면도", "PLAN"]},
    {"key": "core_sections", "name": "코어 단면도",
     "filename_all_groups": [["코아", "코어"], ["단면도"]],
     "content_any": ["코아", "코어", "CORE", "단면도", "SECTION"]},
    {"key": "parking_plans", "name": "지하주차장 평면도",
     "filename_all_groups": [["주차장"], ["평면도"]],
     "filename_exclude": ["구조", "기초", "경사로", "램프"],
     "content_any": ["지하주차장", "평면도", "PARKING"]},
    {"key": "parking_sections", "name": "지하주차장 종·횡단면도",
     "filename_all_groups": [["주차장"], ["단면도", "종단면", "횡단면", "종횡"]],
     "content_any": ["주차장", "종단면", "횡단면", "SECTION"]},
    {"key": "parking_ramps", "name": "주차장 램프/경사로 평면·단면·상세도",
     "filename_all_groups": [["램프", "경사로"], ["평면", "단면", "상세"]],
     "content_any": ["램프", "경사로", "RAMP"]},
    {"key": "amenity_drawings", "name": "부대시설 평면·입면·단면도",
     "filename_all_groups": [["부대시설", "경비실"], ["평", "입", "단면"]],
     "content_any": ["부대시설", "경비실", "평면도", "입면도", "단면도"],
     "content_components": [["평면도", "PLAN"], ["입면도", "ELEVATION"], ["단면도", "SECTION"]]},
    {"key": "finish_schedule", "name": "실내재료마감표",
     "filename_any": ["실내재료마감표", "재료마감표"],
     "content_any": ["재료마감표", "FINISH"]},
    {"key": "window_schedule", "name": "창호일람표",
     "filename_any": ["창호일람표", "창호도"],
     "content_any": ["창호일람표", "창호도", "WINDOW"]},
    {"key": "overview_reference", "name": "사업개요·동별개요·면적산출표 (참고/검산)",
     "filename_any": ["사업개요", "동별개요", "면적산출", "면적 산출"],
     "content_any": ["사업개요", "동별개요", "건축면적", "연면적"], "reference": True},
]

CAD_PRECHECK_ALLOWED_EXTENSIONS = {".dwg", ".dxf"}
CAD_PRECHECK_EXCLUDED_EXTENSIONS = {".bak", ".dwl", ".dwl2", ".cdc", ".txt", ".err"}


def _zip_name_quality(value):
    """깨진 ZIP 이름보다 한글/NFC 이름을 우선하기 위한 보수적인 품질 점수."""
    hangul = sum(
        "\uac00" <= char <= "\ud7a3" or "\u1100" <= char <= "\u11ff"
        or "\u3130" <= char <= "\u318f"
        for char in value
    )
    controls = sum(unicodedata.category(char).startswith("C") for char in value)
    mojibake = sum(char in "ßäàåÇÑ⌐⌫⌂╡╕│┤¡" for char in value)
    replacements = value.count("\ufffd")
    return (hangul, -(controls + replacements), -mojibake)


def _decode_zip_member_name(zip_info):
    """ZIP UTF-8 플래그를 존중하고, 비플래그 CP437 오해석만 안전하게 복구한다."""
    raw_name = str(zip_info.filename or "")
    normalized_raw = unicodedata.normalize("NFC", raw_name)
    if zip_info.flag_bits & 0x800:
        return {
            "decoded_name": normalized_raw,
            "raw_name": raw_name,
            "decode_method": "utf8_flag",
        }

    try:
        original_bytes = raw_name.encode("cp437")
    except UnicodeEncodeError:
        return {
            "decoded_name": normalized_raw,
            "raw_name": raw_name,
            "decode_method": "unchanged",
        }

    candidates = []
    for encoding, method in (
        ("utf-8", "legacy_cp437_to_utf8"),
        ("cp949", "legacy_cp437_to_cp949"),
        ("euc-kr", "legacy_cp437_to_cp949"),
    ):
        try:
            candidate = unicodedata.normalize("NFC", original_bytes.decode(encoding))
        except UnicodeDecodeError:
            continue
        candidates.append((
            candidate,
            "unchanged" if candidate == normalized_raw else method,
        ))
        if encoding == "cp949":
            break

    if not candidates:
        return {
            "decoded_name": normalized_raw,
            "raw_name": raw_name,
            "decode_method": "failed",
        }

    decoded_name, decode_method = max(
        candidates,
        key=lambda item: (_zip_name_quality(item[0]), item[1] == "legacy_cp437_to_utf8"),
    )
    return {
        "decoded_name": decoded_name,
        "raw_name": raw_name,
        "decode_method": decode_method,
    }


def _is_macos_zip_metadata(member_path):
    parts = str(member_path or "").replace("\\", "/").split("/")
    basename = parts[-1] if parts else ""
    return (
        any(part.upper() == "__MACOSX" for part in parts)
        or basename.startswith("._")
        or basename in {".DS_Store", ".localized"}
    )


# ─────────────────────────────────────────────
#  뷰: 메인 페이지
# ─────────────────────────────────────────────
@user_passes_test(admin_required)
def quantity_main(request):
    """AI 구조산출 자동화 메인 페이지 — 관리자(스태프/슈퍼유저)만 접근 가능.
    Gemini 호출 비용이 실제로 발생하는 도구라 결제 시스템이 붙기 전까지는 공개하지 않는다."""
    user_display = request.user.username if request.user.is_authenticated else "게스트"
    return render(request, "core/quantity_main.html", {
        "user_display": user_display,
        "required_structural": REQUIRED_STRUCTURAL,
        "required_architectural": REQUIRED_ARCHITECTURAL,
    })


def _cad_source_folder(upload_name, member_path):
    parts = [
        _normalize_for_match(part)
        for part in str(member_path or "").replace("\\", "/").split("/")[:-1]
    ]
    upload_hint = _normalize_for_match(os.path.splitext(upload_name or "")[0])
    if any("XREF" in part for part in parts) or "XREF" in upload_hint:
        return "XRef"
    if any("구조" in part for part in parts) or "구조" in upload_hint:
        return "구조"
    if any("건축" in part for part in parts) or "건축" in upload_hint:
        return "건축"
    return "루트/기타"


def _cad_item_matches(record, item):
    filename_norm = _normalize_for_match(record["filename"])
    if any(_normalize_for_match(word) in filename_norm for word in item.get("filename_exclude", [])):
        return False
    if item.get("filename_any") and not any(
        _normalize_for_match(word) in filename_norm for word in item["filename_any"]
    ):
        return False
    for alternatives in item.get("filename_all_groups", []):
        if not any(_normalize_for_match(word) in filename_norm for word in alternatives):
            return False
    return True


def _collect_cad_precheck_inventory(uploaded_files):
    """모든 업로드와 ZIP 하위 경로를 열거한다. 내용 파싱 상한은 이 단계에 적용하지 않는다."""
    records = []
    excluded = []
    zip_stats = []
    scan_errors = []
    seen_uploads = set()
    accepted_upload_names = []

    for upload_index, uploaded in enumerate(uploaded_files):
        upload_name = unicodedata.normalize("NFC", os.path.basename(uploaded.name or "upload"))
        upload_role = getattr(uploaded, "_quantity_cad_role", None)
        raw = uploaded.read()
        upload_digest = hashlib.sha256(raw).hexdigest()
        upload_key = (upload_name, len(raw), upload_digest)
        if upload_key in seen_uploads:
            continue
        seen_uploads.add(upload_key)
        accepted_upload_names.append(upload_name)
        ext = os.path.splitext(upload_name)[1].lower()

        if ext in CAD_PRECHECK_ALLOWED_EXTENSIONS:
            records.append({
                "upload_name": upload_name, "path": upload_name, "raw_path": upload_name,
                "filename": upload_name, "decode_method": "unchanged",
                "source_folder": _cad_source_folder(upload_name, upload_name),
                "extension": ext, "data": raw, "upload_index": upload_index,
                "upload_role": upload_role,
            })
            continue
        if ext != ".zip":
            excluded.append({"upload": upload_name, "path": upload_name, "reason": "지원하지 않는 업로드 형식"})
            continue

        try:
            with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                infos = zf.infolist()
                zip_stats.append({"name": upload_name, "entry_count": len(infos)})
                logger.info(
                    "quantity_cad_precheck_zip upload=%s total_entries=%s",
                    upload_name, len(infos),
                )
                decode_counts = {}
                cad_log_count = 0
                for info in infos:
                    if info.is_dir():
                        continue
                    decoded = _decode_zip_member_name(info)
                    raw_member_path = decoded["raw_name"].replace("\\", "/")
                    member_path = decoded["decoded_name"].replace("\\", "/")
                    decode_method = decoded["decode_method"]
                    decode_counts[decode_method] = decode_counts.get(decode_method, 0) + 1
                    if _is_macos_zip_metadata(member_path) or _is_macos_zip_metadata(raw_member_path):
                        reason = "macOS 메타데이터"
                        excluded.append({
                            "upload": upload_name, "path": member_path,
                            "raw_path": raw_member_path, "reason": reason,
                        })
                        logger.info(
                            "quantity_cad_precheck_excluded upload=%s path=%s reason=%s",
                            upload_name, member_path, reason,
                        )
                        continue
                    member_ext = os.path.splitext(member_path)[1].lower()
                    if member_ext not in CAD_PRECHECK_ALLOWED_EXTENSIONS:
                        reason = (
                            "기본도면 제외 확장자" if member_ext in CAD_PRECHECK_EXCLUDED_EXTENSIONS
                            else "CAD 도면이 아닌 파일"
                        )
                        excluded.append({"upload": upload_name, "path": member_path, "reason": reason})
                        logger.info(
                            "quantity_cad_precheck_excluded upload=%s path=%s reason=%s",
                            upload_name, member_path, reason,
                        )
                        continue
                    try:
                        member_data = zf.read(info)
                    except Exception as exc:
                        scan_errors.append(f"{upload_name}:{member_path}: {str(exc)[:160]}")
                        continue
                    source_folder = _cad_source_folder(upload_name, member_path)
                    records.append({
                        "upload_name": upload_name,
                        "path": member_path,
                        "raw_path": raw_member_path,
                        "filename": unicodedata.normalize("NFC", os.path.basename(member_path)),
                        "decode_method": decode_method,
                        "source_folder": source_folder,
                        "extension": member_ext,
                        "data": member_data,
                        "upload_index": upload_index,
                        "upload_role": upload_role,
                    })
                    if cad_log_count < 30:
                        logger.info(
                            "quantity_cad_precheck_cad upload_name=%s raw_path=%s "
                            "decoded_path=%s decode_method=%s source_folder=%s extension=%s",
                            upload_name, raw_member_path, member_path, decode_method,
                            source_folder, member_ext,
                        )
                        cad_log_count += 1
                logger.info(
                    "quantity_cad_precheck_decode_summary upload=%s method_counts=%s "
                    "cad_path_logs=%s",
                    upload_name, decode_counts, cad_log_count,
                )
        except (zipfile.BadZipFile, OSError, RuntimeError) as exc:
            scan_errors.append(f"{upload_name}: {str(exc)[:160]}")
            logger.warning(
                "quantity_cad_precheck_zip_failed upload=%s error=%s",
                upload_name, str(exc)[:200],
            )

    unique_records = []
    duplicates = []
    by_content = {}
    for record in records:
        digest = hashlib.sha256(record["data"]).hexdigest()
        duplicate_key = (_normalize_for_match(record["filename"]), digest)
        location = {
            "upload_name": record["upload_name"], "path": record["path"],
            "source_folder": record["source_folder"],
        }
        if duplicate_key in by_content:
            canonical = by_content[duplicate_key]
            canonical.setdefault("duplicate_locations", []).append(location)
            duplicates.append({"filename": record["filename"], "locations": [
                {"upload_name": canonical["upload_name"], "path": canonical["path"],
                 "source_folder": canonical["source_folder"]},
                location,
            ]})
            continue
        record["content_sha256"] = digest
        record["duplicate_locations"] = []
        by_content[duplicate_key] = record
        unique_records.append(record)

    folder_counts = {"구조": 0, "건축": 0, "XRef": 0, "루트/기타": 0}
    for record in records:
        folder_counts[record["source_folder"]] = folder_counts.get(record["source_folder"], 0) + 1
    logger.info(
        "quantity_cad_precheck_inventory uploads=%s upload_names=%s cad_count=%s "
        "folder_counts=%s excluded_count=%s scan_error_count=%s",
        len(seen_uploads), accepted_upload_names, len(records),
        folder_counts, len(excluded), len(scan_errors),
    )
    return {
        "records": unique_records, "all_cad_count": len(records), "zip_stats": zip_stats,
        "folder_counts": folder_counts, "excluded": excluded, "duplicates": duplicates,
        "scan_errors": scan_errors, "scan_complete": not scan_errors,
        "upload_count": len(seen_uploads), "upload_names": accepted_upload_names,
    }


def _parse_precheck_candidates(candidate_records):
    """체크리스트 후보를 먼저 배치하고, 그 후보에만 DWG 파싱 안전 상한을 적용한다."""
    selected = candidate_records[:DWG_ZIP_PARSE_MAX_FILES]
    capped = len(candidate_records) > len(selected)
    if capped:
        logger.warning(
            "quantity_cad_precheck_parse_capped candidate_count=%s cap=%s",
            len(candidate_records), DWG_ZIP_PARSE_MAX_FILES,
        )
    if not selected:
        return {}, set(), capped

    buffer = io.BytesIO()
    synthetic_to_record = {}
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for index, record in enumerate(selected):
            synthetic_path = f"{index:04d}/{record['path'].lstrip('/')}"
            zf.writestr(synthetic_path, record["data"])
            synthetic_to_record[synthetic_path] = record
    try:
        parsed_raw = parse_dwg_from_zip(buffer.getvalue())
    except Exception as exc:
        logger.warning("quantity_cad_precheck_content_parse_failed error=%s", str(exc)[:200])
        parsed_raw = {}

    parsed_by_digest = {}
    attempted = set()
    for synthetic_path, record in synthetic_to_record.items():
        attempted.add(record["content_sha256"])
        info = parsed_raw.get(synthetic_path)
        if info is None:
            info = {"error": "CAD 내용 파싱 결과를 찾지 못했습니다."}
        parsed_by_digest[record["content_sha256"]] = info
    return parsed_by_digest, attempted, capped


def _build_cad_precheck(uploaded_files):
    inventory = _collect_cad_precheck_inventory(uploaded_files)
    all_items = CAD_PRECHECK_STRUCTURAL_ITEMS + CAD_PRECHECK_ARCHITECTURAL_ITEMS
    candidates_by_key = {
        item["key"]: [record for record in inventory["records"] if _cad_item_matches(record, item)]
        for item in all_items
    }
    prioritized = []
    seen_digests = set()
    for item in all_items:
        for record in candidates_by_key[item["key"]]:
            digest = record["content_sha256"]
            if digest not in seen_digests:
                prioritized.append(record)
                seen_digests.add(digest)
    parsed, attempted, parse_capped = _parse_precheck_candidates(prioritized)

    def classify(item):
        candidates = candidates_by_key[item["key"]]
        files = [{
            "filename": record["filename"], "path": record["path"],
            "raw_path": record.get("raw_path", record["path"]),
            "decode_method": record.get("decode_method", "unchanged"),
            "upload_name": record["upload_name"], "source_folder": record["source_folder"],
            "duplicate_locations": record.get("duplicate_locations") or [],
        } for record in candidates]
        if not candidates:
            status = "missing" if inventory["scan_complete"] else "scan_incomplete"
            reason = (
                "모든 업로드와 하위 폴더에서 파일명 후보를 찾지 못했습니다."
                if status == "missing" else "일부 업로드를 열거하지 못해 파일 없음으로 단정할 수 없습니다."
            )
        else:
            status = "candidate_unverified"
            primary_candidates = [
                record for record in candidates if record["source_folder"] != "XRef"
            ]
            reason = (
                "XRef에서 후보를 찾았지만 본 도면의 기본도면을 대신할 수 없어 확인이 필요합니다."
                if not primary_candidates
                else "파일은 찾았지만 실제 내용 확인이 필요합니다."
            )
            parsed_texts = []
            for record in primary_candidates:
                info = parsed.get(record["content_sha256"])
                if not info or "error" in info:
                    continue
                parsed_texts.extend(info.get("texts") or [])
            text_norm = _normalize_for_match(" ".join(parsed_texts))
            content_found = any(
                _normalize_for_match(word) in text_norm for word in item["content_any"]
            )
            components_found = all(
                any(_normalize_for_match(word) in text_norm for word in alternatives)
                for alternatives in item.get("content_components", [])
            )
            if content_found and components_found:
                status = "confirmed"
                reason = "파일명 후보와 CAD 본문을 확인했습니다."
        logger.info(
            "quantity_cad_precheck_item key=%s candidates=%s status=%s parse_results=%s",
            item["key"], [record["path"] for record in candidates], status,
            [
                "not_attempted" if record["content_sha256"] not in attempted
                else "failed" if "error" in parsed.get(record["content_sha256"], {})
                else "parsed"
                for record in candidates
            ],
        )
        return {
            "key": item["key"], "name": item["name"], "status": status,
            "reference": bool(item.get("reference")), "files": files, "reason": reason,
        }

    structural = [classify(item) for item in CAD_PRECHECK_STRUCTURAL_ITEMS]
    architectural = [classify(item) for item in CAD_PRECHECK_ARCHITECTURAL_ITEMS]
    final_statuses = [item["status"] for item in structural + architectural]
    final_status = (
        "scan_incomplete" if not inventory["scan_complete"]
        else "missing" if "missing" in final_statuses
        else "candidate_unverified" if "candidate_unverified" in final_statuses
        else "confirmed"
    )
    logger.info(
        "quantity_cad_precheck_complete final_status=%s parse_capped=%s",
        final_status, parse_capped,
    )
    return {
        "structural_checklist": structural,
        "architectural_checklist": architectural,
        "scan": {
            "complete": inventory["scan_complete"], "status": final_status,
            "upload_count": inventory["upload_count"], "upload_names": inventory["upload_names"],
            "zip_stats": inventory["zip_stats"], "cad_count": inventory["all_cad_count"],
            "folder_counts": inventory["folder_counts"], "excluded": inventory["excluded"],
            "duplicates": inventory["duplicates"], "errors": inventory["scan_errors"],
            "cad_files": [{
                "filename": record["filename"], "path": record["path"],
                "raw_path": record.get("raw_path", record["path"]),
                "decode_method": record.get("decode_method", "unchanged"),
                "upload_name": record["upload_name"], "source_folder": record["source_folder"],
                "duplicate_locations": record.get("duplicate_locations") or [],
            } for record in inventory["records"]],
            "parse_candidate_count": len(prioritized),
            "parse_limit": DWG_ZIP_PARSE_MAX_FILES, "parse_capped": parse_capped,
        },
    }


def _collect_request_cad_uploads(request):
    """신규 다중 필드와 기존 단일 필드를 합치되 동일 업로드 객체는 한 번만 반환한다."""
    uploaded_files = []
    seen_objects = set()
    for field in ("cad_files", "zip_file", "structural_zip", "architectural_zip", "cad_file"):
        for uploaded in request.FILES.getlist(field):
            if field in ("zip_file", "structural_zip", "cad_file"):
                setattr(uploaded, "_quantity_cad_role", "structural")
            elif field == "architectural_zip":
                setattr(uploaded, "_quantity_cad_role", "architectural")
            if id(uploaded) in seen_objects:
                continue
            seen_objects.add(id(uploaded))
            uploaded_files.append(uploaded)
    return uploaded_files


def _ordered_cad_records(records, checklist):
    ordered = []
    seen = set()
    for item in checklist:
        for record in records:
            digest = record["content_sha256"]
            if digest not in seen and _cad_item_matches(record, item):
                ordered.append(record)
                seen.add(digest)
    for record in records:
        digest = record["content_sha256"]
        if digest not in seen:
            ordered.append(record)
            seen.add(digest)
    return ordered


def _cad_records_to_zip(records):
    if not records:
        return None
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for index, record in enumerate(records):
            upload_dir = os.path.splitext(record["upload_name"])[0] or f"upload_{index}"
            member_path = f"{upload_dir}/{record['path'].lstrip('/')}"
            zf.writestr(member_path, record["data"])
    return buffer.getvalue()


def _merge_uploaded_cad_sets(uploaded_files):
    """여러 ZIP/DWG/DXF를 구조·건축 통합 ZIP으로 합친다.

    모든 경로를 먼저 열거한 뒤 체크리스트 후보를 ZIP 앞쪽에 배치하므로 실제 파서의
    60개 안전 상한이 적용돼도 기본도면 후보가 일반 참고도면 뒤에서 잘리지 않는다.
    """
    if not uploaded_files:
        return None, None, {"upload_count": 0, "cad_count": 0}
    inventory = _collect_cad_precheck_inventory(uploaded_files)
    structural_records = []
    architectural_records = []
    for record in inventory["records"]:
        name_norm = _normalize_for_match(record["filename"])
        is_structural = (
            record.get("upload_role") == "structural"
            or record["source_folder"] == "구조" or name_norm.startswith("S-")
            or any(_cad_item_matches(record, item) for item in CAD_PRECHECK_STRUCTURAL_ITEMS)
        )
        is_architectural = (
            record.get("upload_role") == "architectural"
            or record["source_folder"] == "건축" or name_norm.startswith("A-")
            or any(_cad_item_matches(record, item) for item in CAD_PRECHECK_ARCHITECTURAL_ITEMS)
        )
        if is_structural:
            structural_records.append(record)
        if is_architectural:
            architectural_records.append(record)
        if not is_structural and not is_architectural:
            # 기존 structural_zip/cad_file의 분류 불가능한 CAD도 조용히 버리지 않는다.
            structural_records.append(record)

    structural_records = _ordered_cad_records(
        structural_records, CAD_PRECHECK_STRUCTURAL_ITEMS,
    )
    architectural_records = _ordered_cad_records(
        architectural_records, CAD_PRECHECK_ARCHITECTURAL_ITEMS,
    )
    logger.info(
        "quantity_cad_merge_for_run uploads=%s total_cad=%s structural=%s architectural=%s "
        "scan_complete=%s",
        inventory["upload_count"], inventory["all_cad_count"],
        len(structural_records), len(architectural_records), inventory["scan_complete"],
    )
    return (
        _cad_records_to_zip(structural_records),
        _cad_records_to_zip(architectural_records),
        {
            "upload_count": inventory["upload_count"],
            "upload_names": inventory["upload_names"],
            "cad_count": inventory["all_cad_count"],
            "structural_count": len(structural_records),
            "architectural_count": len(architectural_records),
            "scan_complete": inventory["scan_complete"],
            "scan_errors": inventory["scan_errors"],
        },
    )


# ─────────────────────────────────────────────
#  API: CAD ZIP/DWG/DXF 기본도면 사전검토
# ─────────────────────────────────────────────
@require_POST
@_admin_only_json
def api_check_zip(request):
    uploaded_files = _collect_request_cad_uploads(request)
    if not uploaded_files:
        return JsonResponse({"error": "검토할 CAD ZIP/DWG/DXF 파일이 없습니다."}, status=400)

    logger.info(
        "quantity_cad_precheck_received upload_count=%s upload_names=%s",
        len(uploaded_files), [uploaded.name for uploaded in uploaded_files],
    )
    return JsonResponse(_build_cad_precheck(uploaded_files))


# ─────────────────────────────────────────────
#  DWG → DXF 변환: ODA File Converter 연동
#  (core/views.py의 cblcad_dwg_to_dxf_api 와 동일한 탐색 방식을 재사용)
# ─────────────────────────────────────────────
def _find_oda_converter():
    """서버에 설치된 ODA File Converter 실행 파일 경로를 찾는다. 없으면 None."""
    candidates = [
        "/Applications/ODAFileConverter.app/Contents/MacOS/ODAFileConverter",
        "/Applications/ODA File Converter.app/Contents/MacOS/ODAFileConverter",
        "/Applications/ODAFileConverter 26.10.app/Contents/MacOS/ODAFileConverter",
        "/Applications/ODAFileConverter 25.12.app/Contents/MacOS/ODAFileConverter",
        "/Applications/ODAFileConverter 24.12.app/Contents/MacOS/ODAFileConverter",
        shutil.which("ODAFileConverter"),
        shutil.which("ODAFileConverter.exe"),
    ]
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return None


def _convert_dwg_folder_to_dxf(dwg_dir, out_dir, timeout=90):
    """
    dwg_dir 안의 모든 .dwg 파일을 ODA File Converter로 일괄 변환해 out_dir에 저장.
    변환기가 없거나 실패하면 (False, 메시지) 반환.
    """
    import subprocess

    converter = _find_oda_converter()
    if not converter:
        return False, "ODA_NOT_FOUND"

    cmd = [converter, dwg_dir, out_dir, "ACAD2004", "DXF", "0", "1"]
    try:
        subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=timeout)
    except Exception as e:
        return False, f"CONVERT_EXCEPTION: {e}"

    return True, "OK"


# ─────────────────────────────────────────────
#  DWG/DXF 파싱: ezdxf로 레이어/텍스트/치수/도형(길이·면적·블록개수) 추출
# ─────────────────────────────────────────────
def _polygon_area(pts):
    """슐레이스 공식으로 폐곡선 면적(부호 없음) 계산"""
    area = 0.0
    n = len(pts)
    for i in range(n):
        x1, y1 = pts[i]
        x2, y2 = pts[(i + 1) % n]
        area += x1 * y2 - x2 * y1
    return abs(area) / 2.0


def _polygon_centroid(pts):
    """폐곡선 도심(centroid) 계산 — point-in-polygon 판정용 대표점"""
    n = len(pts)
    A = 0.0
    cx = 0.0
    cy = 0.0
    for i in range(n):
        x0, y0 = pts[i]
        x1, y1 = pts[(i + 1) % n]
        cross = x0 * y1 - x1 * y0
        A += cross
        cx += (x0 + x1) * cross
        cy += (y0 + y1) * cross
    A *= 0.5
    if abs(A) < 1e-9:
        return (sum(p[0] for p in pts) / n, sum(p[1] for p in pts) / n)
    return (cx / (6 * A), cy / (6 * A))


def _point_in_polygon(pt, poly):
    """레이캐스팅 알고리즘으로 점이 폐곡선 내부에 있는지 판정"""
    x, y = pt
    inside = False
    n = len(poly)
    j = n - 1
    for i in range(n):
        xi, yi = poly[i]
        xj, yj = poly[j]
        if ((yi > y) != (yj > y)) and (x < (xj - xi) * (y - yi) / (yj - yi + 1e-12) + xi):
            inside = not inside
        j = i
    return inside


def _split_outer_and_holes(polys):
    """
    같은 레이어의 폐곡선 리스트를 콘크리트 외곽선(outer)과 그 안에 뚫린
    개구부(hole)로 분리한다. 별도 레이어 이름 규칙에 의존하지 않고,
    순수 기하학적 포함관계(centroid가 더 큰 폴리곤 안에 있는지)만으로 판단한다.
    "선(콘크리트 외곽선)만 따서" 개구부를 자동으로 인식하기 위한 로직.
    Returns: (outer_area_sum, hole_area_sum, hole_count)
    """
    if not polys:
        return 0.0, 0.0, 0
    # 면적 큰 순서로 정렬 — 더 큰 폴리곤이 outer 후보
    sorted_polys = sorted(polys, key=lambda p: p[1], reverse=True)
    is_hole = [False] * len(sorted_polys)

    for i, (pts_i, area_i) in enumerate(sorted_polys):
        centroid_i = _polygon_centroid(pts_i)
        for j in range(i):
            if is_hole[j]:
                continue
            pts_j, area_j = sorted_polys[j]
            if area_j <= area_i:
                continue
            if _point_in_polygon(centroid_i, pts_j):
                is_hole[i] = True
                break

    outer_area = sum(a for (_, a), hole in zip(sorted_polys, is_hole) if not hole)
    hole_area = sum(a for (_, a), hole in zip(sorted_polys, is_hole) if hole)
    hole_count = sum(1 for hole in is_hole if hole)
    return outer_area, hole_area, hole_count


def _extract_dxf_quantities(doc):
    """ezdxf Document에서 레이어별 도형 통계(길이/면적/블록 개수) + 텍스트/치수를 뽑는다.

    닫힌 폴리곤(벽체/슬래브 콘크리트 외곽선 등)은 같은 레이어 안에서 다른 폴리곤을
    감싸고 있으면 그 안쪽 폴리곤을 "개구부(hole)"로 인식해 콘크리트 순면적에서
    자동으로 빼준다 (레이어 이름 규칙에 의존하지 않는 순수 기하 판정).
    """
    msp = doc.modelspace()

    layer_names = [l.dxf.name for l in doc.layers]
    texts = []
    dimensions = []
    block_counts = {}
    layer_stats = {}  # layer -> {"length": float, "count": int}
    closed_polys = {}  # layer -> [(points, area), ...]

    def stat(layer):
        return layer_stats.setdefault(layer, {"length": 0.0, "count": 0})

    for entity in msp:
        dxftype = entity.dxftype()
        layer = getattr(entity.dxf, "layer", "0")
        s = stat(layer)
        s["count"] += 1

        try:
            if dxftype == "LINE":
                sp, ep = entity.dxf.start, entity.dxf.end
                s["length"] += ((ep.x - sp.x) ** 2 + (ep.y - sp.y) ** 2) ** 0.5
            elif dxftype in ("LWPOLYLINE", "POLYLINE"):
                pts = list(entity.get_points("xy")) if dxftype == "LWPOLYLINE" else [
                    (v.dxf.location.x, v.dxf.location.y) for v in entity.vertices
                ]
                length = 0.0
                for i in range(len(pts) - 1):
                    length += ((pts[i + 1][0] - pts[i][0]) ** 2 + (pts[i + 1][1] - pts[i][1]) ** 2) ** 0.5
                is_closed = getattr(entity, "is_closed", False)
                if is_closed and len(pts) >= 3:
                    length += ((pts[0][0] - pts[-1][0]) ** 2 + (pts[0][1] - pts[-1][1]) ** 2) ** 0.5
                    area = _polygon_area(pts)
                    if area > 0:
                        closed_polys.setdefault(layer, []).append((pts, area))
                s["length"] += length
            elif dxftype == "HATCH":
                # 근사 면적(바운딩 박스 기반 - 정밀하지 않음, 참고용)
                try:
                    bbox = entity.paths[0].source_boundary_objects
                except Exception:
                    pass
            elif dxftype == "INSERT":
                block_counts[entity.dxf.name] = block_counts.get(entity.dxf.name, 0) + 1
            elif dxftype in ("TEXT", "MTEXT"):
                t = entity.dxf.text if dxftype == "TEXT" else entity.text
                if t and t.strip():
                    texts.append(t.strip()[:200])
            elif dxftype == "DIMENSION":
                val = entity.dxf.actual_measurement
                dimensions.append(round(val, 2))
        except Exception:
            pass

    # 레이어별 폐곡선을 outer(콘크리트 외곽선)/hole(개구부)로 분리해서
    # net_closed_area(개구부 제외 순면적)를 계산한다.
    opening_stats = {
        layer: _split_outer_and_holes(polys)
        for layer, polys in closed_polys.items()
    }

    # 도면 단위가 mm인지 m인지 알 수 없으므로 원시값(도면 단위) 그대로 반환하고
    # 어떤 단위인지는 AI가 텍스트/치수 정보를 보고 판단하도록 note를 남긴다.
    layer_summary = {}
    for layer, v in layer_stats.items():
        outer_area, hole_area, hole_count = opening_stats.get(layer, (0.0, 0.0, 0))
        layer_summary[layer] = {
            "entity_count": v["count"],
            "total_length": round(v["length"], 2),
            "net_closed_area": round(outer_area - hole_area, 2),
            "opening_area": round(hole_area, 2),
            "opening_count": hole_count,
        }

    return {
        "layers": layer_names[:40],
        "layer_geometry": layer_summary,
        "block_counts": block_counts,
        "texts": texts[:120],
        "dimensions": dimensions[:120],
        "unit_note": "length/area 값은 도면 원본 단위(대부분 mm) 기준 원시값입니다. 실제 축척/단위는 texts, dimensions, 도면 제목을 참고해 판단하세요.",
        "opening_note": "net_closed_area는 같은 레이어 안에서 다른 폐곡선에 완전히 둘러싸인 폴리곤을 개구부(hole)로 자동 인식해 뺀 순면적입니다. opening_count가 0보다 크면 해당 레이어(주로 벽체/슬래브 콘크리트 외곽선)에 뚫린 구멍이 있다는 뜻이니, walls/slabs의 openings 필드를 채울 때 이 값을 실제 도면 이미지와 대조해 참고하세요.",
    }


# 2026-07-27: parse_dwg_from_zip은 원래 REQUIRED_STRUCTURAL/REQUIRED_ARCHITECTURAL의
# 임의로 지어낸 도면번호(target_codes)에 매칭되는 파일만 골라 파싱했다. 그 코드 체계가
# 실제 프로젝트 번호 규칙과 전혀 안 맞는다는 게 확인된 이상, api_run_quantity 본 산출
# 파이프라인에서 이 함수를 호출할 때마다 매칭되는 파일이 0개라서 dwg_data가 항상 빈
# 딕셔너리로 나왔을 가능성이 크다(= 구조/건축 PDF 기반 Gemini 추출은 정상 진행되지만,
# ZIP으로 올린 DWG의 기하 데이터가 보조 검증용으로 전혀 반영되지 않고 조용히 버려졌다).
# 도면번호는 사무소마다 달라 신뢰할 수 없으므로, keywords가 주어지면 파일명 키워드로
# 필터링하고(핵심 항목 확인용), keywords=None이면 ZIP 안의 모든 dwg/dxf를 대상으로
# 한다(본 산출 파이프라인의 보조 기하 데이터 추출용 — 특정 목록에만 의존하지 않는다).
# 파일이 지나치게 많은 ZIP에서 변환/파싱이 오래 걸리는 것을 막기 위해 개수 상한을 둔다.
DWG_ZIP_PARSE_MAX_FILES = 60


def parse_dwg_from_zip(zip_bytes, keywords=None):
    """
    ZIP 바이트에서 DWG/DXF 파일을 찾아 파싱한다.
    keywords가 None이면 ZIP 안의 모든 .dwg/.dxf 파일이 대상(최대 DWG_ZIP_PARSE_MAX_FILES개).
    keywords가 주어지면 파일명(공백 무시, 대소문자 무시)에 그 중 하나라도 포함된 파일만
    대상으로 한다.
    DWG(바이너리 AutoCAD 포맷)는 ezdxf가 직접 읽지 못하므로,
    서버에 ODA File Converter가 설치돼 있으면 자동으로 DXF로 변환 후 파싱한다.
    Returns: { filename: { layers, layer_geometry, block_counts, texts, dimensions, ... } | {"error": ...} }
    """
    result = {}
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))

    matched_members = []
    for info in zf.infolist():
        decoded = _decode_zip_member_name(info)
        member = decoded["decoded_name"].replace("\\", "/")
        if _is_macos_zip_metadata(member) or _is_macos_zip_metadata(decoded["raw_name"]):
            continue
        ext = os.path.splitext(member)[1].lower()
        if ext not in (".dwg", ".dxf"):
            continue
        if keywords is None:
            matched_members.append((info, member))
            continue
        basename_norm = _normalize_for_match(os.path.basename(member))
        if any(_normalize_for_match(kw) in basename_norm for kw in keywords):
            matched_members.append((info, member))

    if not matched_members:
        return result

    if len(matched_members) > DWG_ZIP_PARSE_MAX_FILES:
        logger.warning(
            "quantity_parse_dwg_from_zip_capped total=%s cap=%s",
            len(matched_members), DWG_ZIP_PARSE_MAX_FILES,
        )
        matched_members = matched_members[:DWG_ZIP_PARSE_MAX_FILES]

    with tempfile.TemporaryDirectory(prefix="cbl_qty_") as work_dir:
        dwg_dir = os.path.join(work_dir, "dwg_in")
        dxf_out_dir = os.path.join(work_dir, "dxf_out")
        os.makedirs(dwg_dir, exist_ok=True)
        os.makedirs(dxf_out_dir, exist_ok=True)

        # 1차: 원본 그대로 풀어두기 (파일명 충돌 방지를 위해 인덱스 접두어 사용)
        member_to_local = {}
        dwg_present = False
        for idx, (info, member) in enumerate(matched_members):
            ext = os.path.splitext(member)[1].lower()
            local_name = f"{idx:03d}_{os.path.basename(member)}"
            local_path = os.path.join(dwg_dir if ext == ".dwg" else work_dir, local_name)
            with open(local_path, "wb") as f:
                f.write(zf.read(info))
            member_to_local[member] = (ext, local_path, local_name)
            if ext == ".dwg":
                dwg_present = True

        oda_ok, oda_msg = (False, "SKIPPED")
        if dwg_present:
            oda_ok, oda_msg = _convert_dwg_folder_to_dxf(dwg_dir, dxf_out_dir)

        for _, member in matched_members:
            ext, local_path, local_name = member_to_local[member]
            # 전체 ZIP 경로를 키로 유지해 구조/건축/XRef의 동명 파일이 서로 덮어쓰지 않게 한다.
            out_name = member

            if ext == ".dxf":
                dxf_path = local_path
            else:
                # 변환된 DXF 찾기 (확장자만 dxf로 바뀌고 파일명은 동일)
                converted_name = os.path.splitext(local_name)[0] + ".dxf"
                dxf_path = os.path.join(dxf_out_dir, converted_name)
                if not oda_ok:
                    result[out_name] = {
                        "error": (
                            "DWG 파일은 서버에 설치된 ODA File Converter가 있어야 자동 변환됩니다. "
                            f"(사유: {oda_msg}) DXF로 내보내서 다시 업로드하거나, 서버에 ODA File Converter를 설치해 주세요."
                        )
                    }
                    continue
                if not os.path.exists(dxf_path):
                    result[out_name] = {"error": "DWG → DXF 변환에 실패했습니다 (변환 결과 파일 없음)."}
                    continue

            try:
                doc = ezdxf.readfile(dxf_path)
                result[out_name] = _extract_dxf_quantities(doc)
            except Exception as e:
                result[out_name] = {"error": str(e)}

    return result


def _check_critical_content(zip_bytes):
    """_STRUCTURAL_CRITICAL_ITEMS 각 항목에 대해, 파일명에 그 도면 종류를 나타내는
    한글 키워드가 있는 파일을 찾는 것에서 그치지 않고, 실제로 그 DWG/DXF를 열어
    (parse_dwg_from_zip 재사용) 예상 키워드가 도면 텍스트에 실제로 있는지까지 확인한다.
    도면번호(S-001 등)가 아니라 파일명에 실제로 적힌 한글 도면명으로 매칭한다 —
    번호 체계는 사무소마다 달라 코드 기반 매칭은 신뢰할 수 없다는 게 확인됐다.

    Returns: {key: {"exists": bool, "content_verified": bool|None, "reason": str}}
    - exists=False: 파일명 매칭 자체가 안 됨(파일이 없음)
    - content_verified=True: 파일을 열어봤고 예상 키워드를 실제로 찾음
    - content_verified=False: 파일은 열었지만 예상 키워드를 못 찾음(엉뚱한 내용일 가능성)
    - content_verified=None: 파일은 있는데 열어보지 못함(DWG인데 ODA 미설치 등) — 이
      경우 "확인 안 됨"을 "확인됨"으로 속이지 않고 정직하게 이유를 남긴다."""
    try:
        # ZIP 안의 모든 dwg/dxf를 한 번만 파싱해서 각 항목이 재사용한다
        # (항목별로 다시 열지 않아 변환/파싱 비용을 아낀다).
        parsed = parse_dwg_from_zip(zip_bytes)
    except Exception as e:
        parsed = {}
        logger.warning("quantity_check_zip_content_parse_failed error=%s", str(e)[:200])

    result = {}
    for item in _STRUCTURAL_CRITICAL_ITEMS:
        key = item["key"]
        fname_kw = item["filename_keywords"]
        excl_kw = item.get("exclude_keywords") or []
        matched_entries = [
            (fname, info) for fname, info in parsed.items()
            if _filename_matches_keywords(_normalize_for_match(os.path.basename(fname)), fname_kw, excl_kw)
        ]
        if not matched_entries:
            result[key] = {"exists": False, "content_verified": False, "reason": "파일명에서 이 도면 종류로 보이는 파일을 찾지 못했습니다"}
            continue

        keywords = item["content_keywords"]
        verified = False
        conversion_failed = False
        fail_reason = ""
        for fname, info in matched_entries:
            if "error" in info:
                conversion_failed = True
                fail_reason = info["error"]
                continue
            all_text = " ".join(info.get("texts") or []).upper().replace(" ", "")
            if any(kw.upper().replace(" ", "") in all_text for kw in keywords):
                verified = True
                break

        if verified:
            result[key] = {"exists": True, "content_verified": True, "reason": ""}
        elif conversion_failed:
            result[key] = {"exists": True, "content_verified": None, "reason": fail_reason}
        else:
            result[key] = {
                "exists": True, "content_verified": False,
                "reason": "파일은 찾았지만 도면 텍스트에서 예상 키워드를 찾지 못했습니다 — 실제 내용을 확인해 주세요",
            }
    return result


# ─────────────────────────────────────────────
#  PDF → 이미지 변환
# ─────────────────────────────────────────────
# 합본 PDF가 수십~백 페이지인 실제 프로젝트 도면집을 감안해, 앞 5장만 보내던
# 이전 한도를 넉넉하게 올려둔다. 입력(이미지) 토큰 자체는 80장을 넣어도 Gemini 2.5 Pro의
# 1M 토큰 한도에 여유가 있다 — 문제는 출력(OUTPUT) 쪽이다: 부재가 많은 대형 프로젝트를
# 한 번에 몰아넣으면 응답 JSON이 max_output_tokens(65536)를 넘어 잘리거나, 요청 자체가
# 타임아웃/오류로 실패해서 전체 결과가 0건이 되는 문제가 실제로 발생했다. 그래서 구조 부재
# 추출(extract_structural_members)은 EXTRACTION_BATCH_PAGE_SIZE 단위로 나눠 여러 번
# 호출한 뒤 결과를 합치도록 배치 처리한다 — 아래 관련 코드 참고.
MAX_PDF_PAGES_TO_GEMINI = 80

# convert_from_bytes()는 요청한 페이지 전부를 한 번에 렌더링해서 메모리에 PIL Image로
# 들고 있는다. A1/A0 같은 대형 도면을 150dpi로 렌더링하면 장당 30~50MB 이상이 될 수 있어,
# 79페이지짜리 프로젝트를 한꺼번에 렌더링하면 수 GB 메모리가 순식간에 필요해져 서버가
# 죽거나(MemoryError) 매우 느려질 수 있다. 어차피 image_to_jpeg_bytes()에서 최종적으로
# 1536px로 축소하므로, 150dpi로 뽑을 필요 없이 120dpi 정도로도 충분하다(메모리 약 36% 절감).
PDF_RENDER_DPI = 120
OVERVIEW_TABLE_RENDER_DPI = 220
OVERVIEW_TABLE_IMAGE_MAX_SIZE = (3000, 3000)


def pdf_to_images(pdf_bytes, max_pages=MAX_PDF_PAGES_TO_GEMINI, dpi=PDF_RENDER_DPI):
    """PDF 바이트를 PIL Image 리스트로 변환 (최대 max_pages 페이지).
    구조 부재 추출(extract_structural_members)은 이 함수를 쓰지 않고 배치 단위로 직접
    페이지 범위를 지정해서 렌더링한다(메모리 절감) — 아래 _render_pdf_page_range 참고."""
    images = convert_from_bytes(pdf_bytes, dpi=dpi, first_page=1, last_page=max_pages)
    return images


def _render_pdf_page_range(pdf_bytes, first_page, last_page, dpi=PDF_RENDER_DPI, timeout=None):
    """PDF의 지정된 페이지 범위만 렌더링한다. 대형 PDF를 배치 단위로 나눠 필요한
    페이지만 그때그때 렌더링하고 곧바로 버려서, 전체 페이지를 한꺼번에 메모리에
    올려두지 않게 하기 위함이다."""
    return convert_from_bytes(
        pdf_bytes, dpi=dpi, first_page=first_page, last_page=last_page,
        timeout=timeout,
    )


# ─────────────────────────────────────────────
#  일람표/구조일반사항 페이지 자동 감지
#  — 실제 프로젝트에서 확인된 문제: 보/계단 등의 "마크"는 평면도(N번째 배치)에 있는데,
#  그 마크의 실제 치수가 담긴 부재일람표는 도면집 앞쪽(다른 배치)에 몰려있는 경우가 많다.
#  배치는 서로 독립적으로 Gemini에 보내지므로, 평면도만 본 배치는 그 일람표 내용을
#  전혀 모른 채로 "마크만 확인, 치수 미확인"으로 남게 된다. 이걸 완화하기 위해
#  일람표/구조일반사항으로 보이는 페이지를 텍스트 키워드로 미리 찾아서, 모든 배치의
#  이미지에 함께 끼워 넣는다(자기 배치 범위와 겹치는 페이지는 중복이라 제외).
# ─────────────────────────────────────────────
_SCHEDULE_PAGE_KEYWORDS = [
    "일람표", "SCHEDULE", "구조일반사항", "GENERAL NOTE", "이음길이표", "정착길이표",
    "사용재료", "재료강도", "표준상세",
]
SCHEDULE_PAGE_MAX_CHECK = 40  # 일람표는 관례상 도면집 앞쪽에 몰려있어 앞부분만 훑는다
SCHEDULE_PAGE_MAX_RESULTS = 6  # 배치마다 끼워넣을 페이지 수를 제한해서 토큰 증가폭을 억제


def _detect_schedule_pages(pdf_bytes, total_pages,
                            max_check_pages=SCHEDULE_PAGE_MAX_CHECK,
                            max_results=SCHEDULE_PAGE_MAX_RESULTS):
    """pdftotext(poppler-utils)로 페이지별 텍스트를 뽑아 일람표/구조일반사항 키워드가
    많이 나오는 페이지를 찾는다. CAD에서 뽑은 PDF가 벡터 텍스트를 유지하고 있으면 잘
    동작하고, 완전히 래스터(이미지) PDF면 텍스트가 안 잡혀 후보가 0개로 나올 수 있다
    — 이 경우 예외 없이 그냥 빈 리스트를 반환한다(배치 주입 기능만 조용히 꺼짐,
    나머지 추출 파이프라인은 정상 동작).
    Returns: 페이지 번호 리스트(오름차순, 최대 max_results개, 키워드 히트 많은 순 우선)."""
    check_upto = min(total_pages, max_check_pages)
    if check_upto <= 0:
        return []
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(pdf_bytes)
            tmp_path = f.name

        scored = []
        for page_num in range(1, check_upto + 1):
            try:
                result = subprocess.run(
                    ["pdftotext", "-f", str(page_num), "-l", str(page_num), tmp_path, "-"],
                    capture_output=True, timeout=10,
                )
                text = result.stdout.decode("utf-8", errors="ignore").upper()
            except Exception:
                continue
            # CAD 표제란은 글자 사이를 벌려서(자소 사이 공백) 그리는 경우가 있어, pdftotext가
            # 그 간격을 단어 구분 공백으로 착각해 "일 람 표"처럼 뽑아낼 수 있다 — 공백을 제거한
            # 버전으로도 같이 대조해서 이런 경우를 놓치지 않게 한다.
            text_nospace = text.replace(" ", "").replace("\t", "")
            hits = sum(1 for kw in _SCHEDULE_PAGE_KEYWORDS if kw.upper() in text or kw.upper() in text_nospace)
            if hits > 0:
                scored.append((hits, page_num))

        scored.sort(key=lambda x: (-x[0], x[1]))
        return sorted(p for _, p in scored[:max_results])
    except Exception:
        return []
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass


# ─────────────────────────────────────────────
#  구조일반사항 전용 3단계
#  프로젝트 개요와 분리해 구조 PDF의 목차/제목/텍스트로 후보를 먼저 좁힌 뒤,
#  선택된 페이지만 고해상도로 판독한다. 파일명은 후보 힌트일 뿐 값의 근거로 쓰지 않는다.
# ─────────────────────────────────────────────
GENERAL_NOTES_TIMEOUT_SEC = 120
GENERAL_NOTES_MAX_TEXT_SCAN_PAGES = 40
GENERAL_NOTES_MAX_SELECTED_PAGES = 12
GENERAL_NOTES_RENDER_DPI = 180
GENERAL_NOTES_LOCATOR_DPI = 100
GENERAL_NOTES_LOCATOR_MAX_PAGE = 24
GENERAL_NOTES_LOCATOR_BATCH_SIZE = 6
_GENERAL_NOTES_TERMS = {
    "title": ("구조일반사항", "구조설계개요", "GENERAL NOTES", "GENERAL NOTE"),
    "material": ("구조재료", "콘크리트", "철근", "피복두께"),
    "detail": ("정착", "이음", "갈고리", "HOOK", "내진"),
}


def _general_notes_log(event, level=logging.INFO, **payload):
    safe = {"event": event}
    safe.update(payload)
    logger.log(
        level,
        "quantity_general_notes %s",
        json.dumps(safe, ensure_ascii=False, sort_keys=True, default=str),
    )


def _general_notes_page_candidates(pdf_bytes, total_pages, page_hints=None,
                                   max_scan_pages=GENERAL_NOTES_MAX_TEXT_SCAN_PAGES):
    """텍스트 레이어와 낱장 파일명 힌트로 구조일반사항 후보를 점수화한다.

    전체 페이지를 이미지로 보내지 않는다. 텍스트 레이어가 없는 페이지는 후보로
    확정하지 않으며, 파일명 힌트는 고해상도 판독 대상으로만 올린다.
    """
    limit = min(int(total_pages or 0), int(max_scan_pages))
    hint_pages = {
        int(page)
        for page, info in (page_hints or {}).items()
        if "general_spec" in (info.get("hints") or set())
    }
    tmp_path = None
    rows = []
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as handle:
            handle.write(pdf_bytes or b"")
            tmp_path = handle.name
        for page in range(1, limit + 1):
            text = ""
            text_error = None
            try:
                result = subprocess.run(
                    ["pdftotext", "-f", str(page), "-l", str(page), tmp_path, "-"],
                    capture_output=True, timeout=10,
                )
                if result.returncode == 0:
                    text = result.stdout.decode("utf-8", errors="ignore")
                else:
                    text_error = f"pdftotext_returncode_{result.returncode}"
            except Exception as exc:
                text_error = type(exc).__name__
            normalized = re.sub(r"\s+", "", text).upper()
            title_hits = [
                term for term in _GENERAL_NOTES_TERMS["title"]
                if re.sub(r"\s+", "", term).upper() in normalized
            ]
            material_hits = [
                term for term in _GENERAL_NOTES_TERMS["material"]
                if re.sub(r"\s+", "", term).upper() in normalized
            ]
            detail_hits = [
                term for term in _GENERAL_NOTES_TERMS["detail"]
                if re.sub(r"\s+", "", term).upper() in normalized
            ]
            drawing_number_set = {
                f"S-{int(match):03d}"
                for match in re.findall(r"\bS\s*[-–—]\s*0?(\d{2,3})\b", text, flags=re.I)
            }
            for start, end in re.findall(
                r"S\s*[-–—]\s*0?(\d{2,3})\s*[~～]\s*(?:S\s*[-–—]\s*)?0?(\d{2,3})",
                text, flags=re.I,
            ):
                first, last = int(start), int(end)
                if first <= last:
                    drawing_number_set.update(
                        f"S-{number:03d}" for number in range(first, last + 1)
                    )
            drawing_numbers = sorted(drawing_number_set)
            drawing_list_markers = (
                "도면목록" in normalized
                or "DRAWINGLIST" in normalized
                or all(marker in normalized for marker in ("도면번호", "도면명", "비고"))
            )
            is_drawing_list = drawing_list_markers or len(drawing_numbers) >= 4
            score = (
                len(title_hits) * 8
                + len(material_hits) * 3
                + len(detail_hits) * 2
                + len(drawing_numbers) * 5
                + (10 if page in hint_pages else 0)
            )
            reasons = []
            if title_hits:
                reasons.append("제목:" + ",".join(title_hits))
            if drawing_numbers:
                reasons.append("도면번호:" + ",".join(drawing_numbers))
            if material_hits:
                reasons.append("재료:" + ",".join(material_hits))
            if detail_hits:
                reasons.append("상세:" + ",".join(detail_hits))
            if page in hint_pages:
                reasons.append("업로드 파일명 힌트")
            selected = score > 0 and not is_drawing_list
            rows.append({
                "page": page,
                "score": score,
                "selected": selected,
                "page_type": "drawing_list" if is_drawing_list else (
                    "general_notes_text_candidate" if selected else "other"
                ),
                "reasons": reasons,
                "drawing_numbers": drawing_numbers,
                "text_available": bool(text.strip()),
                "text_error": text_error,
                "rejection_reason": "drawing_list_not_content" if is_drawing_list else (
                    None if selected else "general_notes_markers_not_found"
                ),
            })
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
    ranked = sorted(
        (row for row in rows if row["selected"]),
        key=lambda row: (-row["score"], row["page"]),
    )[:GENERAL_NOTES_MAX_SELECTED_PAGES]
    selected_pages = sorted(row["page"] for row in ranked)
    expected_numbers = sorted({
        number for row in rows if row["page_type"] == "drawing_list"
        for number in row["drawing_numbers"]
    })
    drawing_list_pages = [row["page"] for row in rows if row["page_type"] == "drawing_list"]
    return {
        "total_pages": int(total_pages or 0),
        "scan_range": [1, limit] if limit else [],
        "pages": rows,
        "selected_pages": selected_pages,
        "drawing_list_pages": drawing_list_pages,
        "expected_drawing_numbers": expected_numbers,
        "text_used": any(row["text_available"] for row in rows),
        "image_fallback": bool(selected_pages),
        "complete": bool(selected_pages),
    }


GENERAL_NOTES_LOCATOR_PROMPT = """구조 PDF 페이지 종류를 짧은 JSON으로 분류하세요.
각 페이지마다 pdf_page, page_type, drawing_number, drawing_title, is_general_notes,
confidence, evidence_terms(최대 6개)를 반환하세요. 도면목록은 page_type=drawing_list,
is_general_notes=false입니다. 단일 구조 도면번호 타이틀블록, 구조일반사항/구조설계개요/
GENERAL NOTES 제목, 또는 콘크리트·철근·피복·정착·이음 표식이 강한 실제 내용 페이지만
is_general_notes=true로 하세요. 페이지 순서만 보고 도면번호를 추정하지 마세요."""


def _classify_general_notes_image_pages(pdf_bytes, page_numbers, job_id=None):
    """텍스트가 없는 초기 페이지를 저해상도 vision으로 실제 분류한다."""
    client = get_gemini_client()
    if client is None or not page_numbers:
        return []
    decisions = []
    for offset in range(0, len(page_numbers), GENERAL_NOTES_LOCATOR_BATCH_SIZE):
        batch = page_numbers[offset:offset + GENERAL_NOTES_LOCATOR_BATCH_SIZE]
        contents = [f"요청 PDF 페이지: {batch}"]
        rendered_pages = []
        for page in batch:
            images = _render_pdf_page_range(
                pdf_bytes, page, page, dpi=GENERAL_NOTES_LOCATOR_DPI, timeout=20,
            )
            if not images:
                continue
            rendered_pages.append(page)
            contents.extend([
                f"[PDF 실제 {page}페이지]",
                types.Part.from_bytes(
                    data=image_to_jpeg_bytes(images[0], max_size=(1400, 1400)),
                    mime_type="image/jpeg",
                ),
            ])
        if not rendered_pages:
            continue
        _general_notes_log("vision_locator_call", job_id=job_id, pages=rendered_pages)
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=GENERAL_NOTES_LOCATOR_PROMPT,
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=4096,
                thinking_config=types.ThinkingConfig(thinking_budget=512),
            ),
        )
        raw = _extract_text_from_gemini_response(response)
        try:
            parsed = json.loads(raw.replace("```json", "").replace("```", "").strip())
        except json.JSONDecodeError:
            parsed = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        items = parsed.get("pages") if isinstance(parsed, dict) else parsed
        for item in items or []:
            if not isinstance(item, dict):
                continue
            try:
                page = int(item.get("pdf_page"))
            except (TypeError, ValueError):
                continue
            if page not in rendered_pages:
                continue
            number = re.sub(r"\s+", "", str(item.get("drawing_number") or "").upper())
            number_match = re.fullmatch(r"S[-–—]?0?(\d{2,3})", number)
            normalized_number = (
                f"S-{int(number_match.group(1)):03d}" if number_match else None
            )
            title = str(item.get("drawing_title") or "")
            terms = [str(value) for value in (item.get("evidence_terms") or [])[:6]]
            marker_text = " ".join([title] + terms).upper()
            strong_marker = any(term in marker_text for term in (
                "구조일반사항", "구조설계개요", "GENERAL NOTE", "콘크리트", "철근",
                "피복", "정착", "이음",
            ))
            number_valid = bool(normalized_number)
            is_drawing_list = str(item.get("page_type") or "") == "drawing_list"
            accepted = bool(
                item.get("is_general_notes") and not is_drawing_list
                and (number_valid or strong_marker)
            )
            decision = {
                "pdf_page": page,
                "page_type": item.get("page_type") or "other",
                "drawing_number": normalized_number,
                "drawing_title": title,
                "is_general_notes": accepted,
                "confidence": item.get("confidence"),
                "evidence_terms": terms,
            }
            decisions.append(decision)
            _general_notes_log("vision_locator_result", job_id=job_id, **decision)
    return decisions


def _merge_general_notes_page_candidates(scan, vision_decisions):
    """텍스트 후보와 vision으로 확인한 실제 내용 페이지만 합쳐 최대 12장을 고른다."""
    selected = set(scan.get("selected_pages") or [])
    mapping = {}
    for item in vision_decisions or []:
        if not item.get("is_general_notes"):
            continue
        page = item.get("pdf_page")
        if not isinstance(page, int):
            continue
        selected.add(page)
        number = item.get("drawing_number")
        if number and number not in mapping:
            mapping[number] = page
    drawing_list_pages = set(scan.get("drawing_list_pages") or [])
    selected.difference_update(drawing_list_pages)
    return sorted(selected)[:GENERAL_NOTES_MAX_SELECTED_PAGES], mapping


def _general_notes_has_values(result):
    return bool(
        (result or {}).get("basic_info")
        or any((result or {}).get(key) for key in (
            "concrete_materials", "rebar_materials", "cover_requirements",
            "anchorage_splice_requirements",
        ))
        or any(
            row.get("source_type") != "user_confirmed"
            for row in ((result or {}).get("quantity_notes") or [])
        )
    )


# ─────────────────────────────────────────────
#  낱장 파일명 기반 개요/구조일반사항 페이지 힌트
#  — 실제 현장 CAD는 도면 파일명 자체에 "A-015,016 사업개요,동별개요.dwg"처럼 그 시트에
#  뭐가 있는지 이미 명확히 적혀 있는 경우가 많다(2026-07-27 사용자 지적: "캐드에 별도로
#  있잖아"). 이 경우 Vision이 페이지 내용을 보고 추측하는 것보다 파일명에 이미 적힌
#  정답을 그대로 신뢰하는 게 훨씬 안전하다 — 사용자가 낱장 파일들을 여러 개 선택해서
#  올리면, 이 힌트가 사업개요 로케이터(Vision 추측)보다 우선한다.
# ─────────────────────────────────────────────
_FILENAME_HINT_KEYWORDS = {
    "overview": ["사업개요", "건축개요", "설계개요", "개요"],
    "area_table": ["동별개요", "동별면적", "층별면적", "면적표", "동별현황"],
    "general_spec": ["구조일반사항", "구조설계개요", "구조개요", "일반사항"],
}


def _classify_filename_hint(filename):
    """파일명(확장자 제외)에서 어떤 종류의 페이지인지 힌트를 뽑는다. 한 파일이 여러
    종류에 동시에 해당할 수 있다(예: "사업개요,동별개요"가 한 시트에 같이 있는 경우)
    — 매칭된 종류를 전부 집합으로 반환한다. 매칭 없으면 빈 집합."""
    name = os.path.splitext(str(filename or ""))[0]
    name_nospace = name.replace(" ", "").replace("_", "")
    hints = set()
    for hint_type, keywords in _FILENAME_HINT_KEYWORDS.items():
        for kw in keywords:
            if kw in name or kw in name_nospace:
                hints.add(hint_type)
                break
    return hints


def _merge_uploaded_pdfs(uploaded_files):
    """여러 개의 업로드된 PDF 파일(각각 .name/.read()를 가진 Django UploadedFile 또는
    (파일명, bytes) 튜플)을 페이지 순서대로 하나의 PDF로 합치고, 합쳐진 PDF의 각
    페이지 번호(1부터)가 원래 어느 파일에서 왔는지 매핑을 함께 반환한다.

    Returns: (merged_pdf_bytes 또는 None, page_source_map) — page_source_map은
    {페이지번호: {"filename": 원본파일명, "hints": {"overview", ...}}} 형태(힌트가
    없는 파일의 페이지는 매핑에서 제외).

    파일이 1개뿐이면 재인코딩 없이 그 파일의 원본 바이트를 그대로 반환한다 — 지금까지
    "합본 PDF 1개 업로드" 방식으로 잘 동작하던 경로의 화질/바이트를 그대로 보존하기
    위함이다. 유효한 PDF가 아닌 파일(예: 변환 안 된 원본 .dwg를 실수로 선택한 경우)은
    조용히 건너뛴다."""
    items = []
    for f in uploaded_files:
        if isinstance(f, tuple):
            name, data = f
        else:
            name, data = getattr(f, "name", "unknown.pdf"), f.read()
        if not data:
            continue
        items.append((name, data))

    if not items:
        return None, {}

    if len(items) == 1:
        name, data = items[0]
        hints = _classify_filename_hint(name)
        page_source_map = {}
        if hints:
            try:
                total = len(PdfReader(io.BytesIO(data)).pages)
            except Exception:
                total = 0
            for p in range(1, total + 1):
                page_source_map[p] = {"filename": name, "hints": hints}
        return data, page_source_map

    writer = PdfWriter()
    page_source_map = {}
    page_counter = 0
    for name, data in items:
        try:
            reader = PdfReader(io.BytesIO(data))
        except Exception:
            logger.warning("quantity_multi_pdf_merge_skip_invalid file=%s", name)
            continue
        hints = _classify_filename_hint(name)
        for page in reader.pages:
            writer.add_page(page)
            page_counter += 1
            if hints:
                page_source_map[page_counter] = {"filename": name, "hints": hints}

    if page_counter == 0:
        return None, {}

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue(), page_source_map


# ─────────────────────────────────────────────
#  지하주차장 각층평면도 페이지 자동 감지
#  — 본 추출(비용 큼) 전에 지하주차장 평면도만 먼저 가볍게 읽어서 부재(벽/보/슬래브/계단)
#  색칠 미리보기를 보여주고 사람이 인식이 맞는지 확인받기 위해, 그 페이지를 먼저 찾는다.
#  "지하/B1/주차장" 같은 층 키워드만으로는 구조일반사항 등 다른 페이지도 걸릴 수 있어서,
#  "평면도/PLAN" 키워드가 함께 있는 페이지만 후보로 삼는다(오탐 방지).
# ─────────────────────────────────────────────
_BASEMENT_PLAN_LEVEL_KEYWORDS = [
    "지하", "B1", "B2", "B3", "B4", "지하1층", "지하2층", "지하3층", "주차장",
]
_BASEMENT_PLAN_TYPE_KEYWORDS = ["평면도", "PLAN"]
BASEMENT_PLAN_PAGE_MAX_CHECK = 60  # 지하주차장 평면도는 도면집 전반에 흩어져 있을 수 있어 조금 더 넓게 훑는다
BASEMENT_PLAN_PAGE_MAX_RESULTS = 6  # 지하층 수만큼(보통 1~4장) + 여유분


def _detect_basement_plan_pages(pdf_bytes, total_pages,
                                 max_check_pages=BASEMENT_PLAN_PAGE_MAX_CHECK,
                                 max_results=BASEMENT_PLAN_PAGE_MAX_RESULTS):
    """_detect_schedule_pages와 동일한 방식(pdftotext 페이지별 텍스트 스캔)으로 지하주차장
    각층평면도로 보이는 페이지를 찾는다. 레이스터(이미지) PDF라 텍스트가 안 잡히면 빈
    리스트를 반환한다 — 이 경우 프론트가 이 확인 단계를 조용히 건너뛰고 바로 본 추출로
    진행한다(기능 실패가 전체 파이프라인을 막으면 안 됨)."""
    check_upto = min(total_pages, max_check_pages)
    if check_upto <= 0:
        return []
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            f.write(pdf_bytes)
            tmp_path = f.name

        scored = []
        for page_num in range(1, check_upto + 1):
            try:
                result = subprocess.run(
                    ["pdftotext", "-f", str(page_num), "-l", str(page_num), tmp_path, "-"],
                    capture_output=True, timeout=10,
                )
                text = result.stdout.decode("utf-8", errors="ignore").upper()
            except Exception:
                continue
            text_nospace = text.replace(" ", "").replace("\t", "")
            has_type = any(kw.upper() in text or kw.upper() in text_nospace for kw in _BASEMENT_PLAN_TYPE_KEYWORDS)
            level_hits = sum(1 for kw in _BASEMENT_PLAN_LEVEL_KEYWORDS if kw.upper() in text or kw.upper() in text_nospace)
            if has_type and level_hits > 0:
                scored.append((level_hits, page_num))

        scored.sort(key=lambda x: (-x[0], x[1]))
        return sorted(p for _, p in scored[:max_results])
    except Exception:
        return []
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def image_to_jpeg_bytes(img: Image.Image, max_size=(1536, 1536)) -> bytes:
    """PIL Image → JPEG 바이트 (Gemini Vision 입력용)"""
    img = img.convert("RGB")
    img.thumbnail(max_size, Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=88)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  구조: Gemini는 "부재 리스트만 읽기" (계산은 quantity_calc.py가 결정론적으로 수행)
# ─────────────────────────────────────────────
MEMBER_EXTRACTION_SYSTEM_PROMPT = """당신은 구조도면을 판독하는 전문가입니다.
도면(부재일람표, 배근도, 평면도, 구조일반사항)에서 보이는 부재의 치수와 철근 정보를
"있는 그대로" 읽어서 구조화된 데이터로만 추출하세요.

절대로 물량(부피, 중량, 면적)을 계산하지 마세요. 계산은 별도 시스템이 합니다.
당신의 역할은 도면에 적힌 치수/규격/개수를 정확히 옮겨 적는 것뿐입니다.

구조일반사항(구조설계개요) 도면이 있으면 반드시 아래 항목을 general_spec에 채워주세요:
- 레미콘(콘크리트) 설계기준강도 Fck (MPa) — 매우 중요: 부재 종류별로 값이 다르게 표기된 경우
  (예: "아파트 전부재(기초 제외) 30MPa", "기초/주차장 전부재 24MPa")에도 절대로 null로 남기지
  마세요. concrete_fck_mpa에는 가장 일반적인 대표값(보통 기초를 제외한 일반부재 값)을 반드시
  채우고, 추가로 general_spec.concrete_fck_table에 기초/기둥/보/슬래브/전단벽/계단 각각에 대해
  개별 행으로 풀어서 채우세요. 표에 "전부재" 처럼 여러 부재가 묶여 표기돼 있으면 해당하는
  개별 카테고리 각각에 같은 값으로 행을 만드세요. 예:
  [{"category":"기초","fck_mpa":24},{"category":"기둥","fck_mpa":30},{"category":"보","fck_mpa":30},
   {"category":"슬래브","fck_mpa":30},{"category":"전단벽","fck_mpa":30},{"category":"계단","fck_mpa":30}]
  모든 부재가 같은 Fck면 표 없이 concrete_fck_mpa 하나만 채워도 됩니다.
- 철근 강종 (SD400, SD500 등) — 매우 중요: 철근 지름별로 강종/항복강도(fy)가 다르게 표기된 경우
  (예: "D10 이하 SD500(fy=500)", "D16 이상 SD500S(fy=500, 내진용)")에도 절대로 null로 남기지
  마세요. rebar_grade에는 가장 일반적인 대표 강종을 반드시 채우고, 추가로
  general_spec.rebar_grade_table에 지름 상한 기준으로 행을 만드세요 (bar_size_max는 그 지름
  "이하"에 적용된다는 뜻이며 mm 숫자만 적으세요, 가장 큰 지름 구간은 bar_size_max를 999로
  채우세요). 예: [{"bar_size_max":10,"grade":"SD500","fy_mpa":500},
  {"bar_size_max":999,"grade":"SD500S","fy_mpa":500}]. 모든 지름이 같은 강종이면 표 없이
  rebar_grade 하나만 채워도 됩니다.
- 철근 이음등급 (A급/B급) — 도면에 명시된 값을 lap_splice_class에 그대로 채우세요(참고용 표기이며,
  실제 이음길이 계산은 항상 B급 기준으로 고정되어 있으니 이 필드 때문에 표를 왜곡해서 옮기지 마세요).
- 철근 이음길이표 (직경별 이음길이, m 단위) — 표가 있으면 반드시 그대로 옮겨 적으세요(lap_splice_table).
  표에 상부근/하부근이 따로 나뉘어 있으면 각 행에 position을 "상부" 또는 "하부"로 채우고,
  구분이 없으면 position은 생략하세요. 표에 A급/B급 이음길이가 각각 다른 열(컬럼)로 나뉘어
  있으면, 반드시 두 값을 각각 별도의 행으로 만들고 그 행에 splice_class를 "A" 또는 "B"로
  채우세요(예: 같은 D25/하부근이라도 A급 행 하나, B급 행 하나를 따로 만드세요). 표에 등급 구분
  없이 값이 하나뿐이면 splice_class는 생략하세요. 근사치를 만들어내지 말고, 표가 없으면 빈
  배열로 두세요. (이 도구는 실제 계산 시 B급 값만 사용하고 A급 값은 참고용으로만 저장합니다.)
- 철근 정착길이표(정착장, 직경별 m 단위) — 표가 있으면 반드시 그대로 옮겨 적으세요(anchorage_table).
  표준갈고리 정착길이표라면 각 행에 "hook": true를 채우고, 직선철근 정착길이표라면 "hook": false로
  채우세요(생략 시 false로 간주). 근사치를 만들어내지 말고, 표가 없으면 빈 배열로 두세요.
- 피복두께(콘크리트 표면부터 철근까지의 최소 거리, mm 단위) — 부재별로 다르게 표기돼 있으면
  그 중 가장 얇은 값을 general_spec.cover_thickness_mm에 채우세요. 확인 안 되면 null로 두세요.
- 우마근(Chair Bar, 슬래브 상하 2단 배근을 지지하는 보조철근) 규격과 높이 — 구조일반사항이나
  슬래브 상세도에 "우마근 HD13" 같은 표기와 상하 철근 사이 높이(H, mm 또는 m)가 함께 있으면
  general_spec.chair_bar_size, general_spec.chair_bar_height_m에 채우세요. 표기가 없으면 (즉,
  슬래브가 단일 배근이거나 우마근 언급이 없으면) 둘 다 null로 두세요 — 절대 추정하지 마세요.

기초(foundations)는 기초일람표 또는 기초상세도에 저판 하부 배근(철근 규격, 간격)이 표기되어
있으면 rebar_size(예: "D16")와 rebar_spacing_m(예: 0.2)을 반드시 채우세요. 도면에서 확인이
안 되면 null로 두고 notes에 "기초 배근 정보 미확인"이라고 남기세요 — 절대 임의로 추정해서
채우지 마세요.

기초상세도에 기둥/벽체가 기초에 연결되는 도웰바(Dowel Bar, 기초 철근과 별도로 표기된 수직
연결철근)가 표기돼 있으면 dowel_bar_size, dowel_bar_count(가닥수)를 채우고, 표준갈고리로
정착되면 dowel_has_hook을 true로 채우세요. 도면에 표기가 없으면 null로 두고 임의로 추정하지
마세요 (도웰바는 기둥 주근을 연장한 것일 수도 있고 별도 철근일 수도 있으므로 반드시 도면
근거가 있을 때만 채우세요).

기둥/보/벽체/슬래브의 주철근(main_rebar 또는 rebar)에 대해, 정착 방식을 판단할 수 있으면
아래 두 필드를 채우세요 (판단 근거가 없으면 둘 다 null로 두세요):
- has_hook: 부재 단부(기둥 하부, 보/슬래브 단부, 벽체 하부)의 정착이 90°/180° 표준갈고리로
  표시돼 있으면 true, 직선 정착이면 false
- is_top_bar (보/슬래브만 해당): 그 철근이 부재 상부에 위치한 철근(부모멘트근, 상부근)이면
  true, 하부(정모멘트근, 하부근)이면 false. 기둥/벽체 수직근에는 이 필드를 채우지 마세요
  (해당 없음).

벽체(walls)는 평면도에서 벽체 단부(끝나는 지점)의 형태를 판단할 수 있으면 end_condition을
채우세요: 벽체가 다른 벽체와 만나지 않고 끝나면 "일자형", T자로 다른 벽체와 만나면 "T자형",
직각으로 꺾이는 모서리면 "모서리". 판단이 안 되면 null로 두세요 (양쪽 단부 조건이 다르면
더 보강량이 많은 쪽을 기준으로 채우세요).

벽체(walls)의 배근이 벽 두께 방향으로 한 겹(단면, 편측)만 있는지 아니면 앞뒤 두 겹(양면)
모두 있는지 판단할 수 있으면 is_single_face를 채우세요: 배근도/단면 상세에 "단면" 표기가
있거나, 벽 두께 방향 단면도에 철근이 한 줄만 그려져 있으면 true. "양면" 표기가 있거나 벽
두께 방향에 철근이 두 줄(전후면) 그려져 있으면 false. 대부분의 구조 전단벽(두께 180mm 이상)은
양면 배근이 표준이므로, 도면에서 명확히 "단면"이라고 확인되지 않는 이상 임의로 true로
채우지 마세요 — 판단 근거가 없으면 null로 두세요 (null이면 안전하게 양면으로 간주해서
계산합니다).

슬래브(slabs)는 데크플레이트(DECK, 철제 영구 거푸집)를 사용하는 슬래브면 is_deck_slab을
true로 채우세요. 일반 합판/유로폼 거푸집을 쓰면 false, 확인이 안 되면 null로 두세요.

부재일람표(기둥일람표, 보일람표)가 있으면 그것을 최우선 근거로 삼으세요.
이번 요청에 여러 장의 도면 이미지가 함께 주어집니다 — 어떤 이미지는 평면도(부재 마크만
보임)이고, 어떤 이미지는 부재일람표/구조일반사항(마크별 치수·철근 정보가 표로 정리됨)일 수
있습니다. 평면도에서 어떤 부재 마크(예: "S2W1", "C1")를 봤는데 그 마크의 치수를 같은
이미지에서 확인할 수 없으면, 함부로 "정보 없음"으로 넘기지 말고 반드시 이번 요청에 함께
포함된 다른 이미지들(특히 일람표/구조일반사항으로 보이는 이미지)을 먼저 확인해서 그 마크와
일치하는 행이 있는지 찾아보세요. 다른 이미지에서 찾았으면 그 치수를 사용하고 note에 "N번째
이미지의 일람표에서 치수 확인"처럼 남기고, 그래도 못 찾았을 때만 null로 두고 "일람표 미확인"
이라고 남기세요.
배근도나 평면도에서만 확인되는 값은 근사치일 수 있음을 notes에 남기세요.
도면에서 확인할 수 없는 값은 null로 남기고 생략하지 마세요.
단위는 전부 미터(m) 기준으로 환산해서 넣으세요 (mm로 적혀 있으면 나눠서 변환).
철근 규격은 D10, D13, D16, D19, D22, D25, D29, D32, D35, D38 표기법으로 통일하세요.

각 부재 항목(기초/기둥/보/슬래브/전단벽/계단)에 대해, 그 부재가 실제로 배치되어 보이는
평면도/배치도 이미지 위의 위치를 bbox 필드에 채워주세요:
{"page": <그 부재가 보이는 이미지 위에 표시된 "[도면 N페이지]"의 N>, "box_2d": [ymin, xmin, ymax, xmax]}
box_2d는 그 페이지 이미지 기준 0~1000 정규화 좌표입니다(좌상단이 (0,0), 우하단이 (1000,1000),
순서는 [ymin, xmin, ymax, xmax]). 부재일람표나 구조일반사항처럼 "표"만 있고 실제 배치 위치가
안 보이는 이미지에는 bbox를 채우지 마세요 — 반드시 그 부재가 평면도/배치도 상에 실제로
그려진 그림을 기준으로 위치를 표시하세요. 같은 마크가 여러 위치에 반복되면(예: 기준층 기둥이
여러 개소) 그 중 하나만 대표로 표시하면 됩니다. 위치를 확신할 수 없으면 bbox 필드 자체를
생략하세요 — 대충 찍어서 채우지 마세요(잘못된 위치 표시보다는 빈 값이 낫습니다).

벽체(walls)와 슬래브(slabs)는 배근도/평면도에 표시된 개구부(문, 계단실, EV실, 설비 샤프트,
덕트 등 콘크리트가 뚫린 구멍)를 openings 배열에 반드시 채우세요. 개구부가 전혀 없으면
openings를 빈 배열([])로 두세요 — null이나 생략은 금지입니다 (계산 시스템이 "정보 없음"과
"개구부 없음"을 구분해서 과다산출 경고를 다르게 표시하기 때문입니다).
DWG 파싱 데이터의 layer_geometry에 있는 opening_area/opening_count는 같은 레이어의
폐곡선(콘크리트 외곽선) 안에 기하학적으로 뚫려 있는 폴리곤을 자동 인식한 값입니다.
이 값이 0보다 크면 실제 도면 이미지에서 해당 벽체/슬래브의 개구부 치수를 찾아 openings에
반영하고, 폭·높이를 읽을 수 없으면 note에 "치수 확인 필요"라고 남기세요.

기초/기둥/보/슬래브/벽체/계단 각 항목에는 zone(층/구역) 필드도 채우세요. 그 부재가 어느 층에
속하는지 도면 제목이나 표기(예: "지하1층 구조평면도", "1층 골조평면도", "기준층(2~15층)
평면도", "옥탑층 골조평면도")를 근거로 최대한 구체적인 "층" 단위로 적으세요:
- 지하층은 "지하1층", "지하2층"처럼 층수를 명시하세요.
- 지상층은 "1층", "2층"처럼 층수를 명시하세요.
- 여러 층이 완전히 동일한 평면(기준층)으로 표기돼 있으면 "기준층(2~15층)"처럼 범위로 묶어
  하나의 zone으로 적고, 그 범위에 포함된 실제 층수를 floor_repeat_count에 채우세요 (예:
  2~15층이 전부 동일하면 floor_repeat_count=14). 이때 count/rebar_count 등 개수 필드는
  "그 평면도 1개 층 기준"의 개수여야 합니다 — floor_repeat_count와 곱해져서 전체 층수만큼의
  물량이 계산되므로, 절대 직접 층수를 곱해서 채우지 마세요.
- 어느 층인지 도면에서 판단하기 어려우면(공용부/전체 개요 등) zone을 "미상"으로 적으세요.
- 반복이 없어 그 zone이 정확히 한 층만 가리키면 floor_repeat_count는 1로 두거나 생략하세요
  (생략 시 1로 간주합니다).
이 값들은 층별 물량 집계와 구역별 철근콘크리트비(철근 누락 여부 점검용) 계산에 모두 쓰입니다.

같은 항목에 section(구간/동) 필드도 채우세요 — 프로젝트가 여러 동/구간(예: "101동", "102동",
"지하주차장", "기전실", "정화조", "주민공동시설", "관리사무소", "경비실")으로 나뉘어 있으면
그 부재가 속한 동/구간명을 도면 제목이나 도면목록표(예: "101동 지하2층 구조평면도" → "101동")
에서 그대로 옮겨 적으세요. 프로젝트 전체가 단일 동/단일 건물이면 그 건물명(또는 "본동")을
적거나, 구분할 필요가 없다고 판단되면 생략해도 됩니다(생략 시 "미상"으로 처리됩니다). zone은
"그 동 안에서의 층"만 나타내고 section은 "어느 동인지"를 나타내므로 서로 다른 동의 같은
층수(예: 101동 1층과 102동 1층)가 섞이지 않도록 반드시 함께 채우세요.

계단(stairs)은 배근도/평면도/단면도에 표기된 경사길이(계단참 포함 전체 사판 길이, length_m),
폭(width_m), 계단판 두께(thickness_m)를 채우고, 경사방향 주근(rebar_size/rebar_spacing_m)과
폭방향 배력근(distribution_rebar_size/distribution_rebar_spacing_m)을 각각 채우세요.
계단은 별도 부재로, slabs 배열에 포함하지 말고 반드시 stairs 배열에 넣으세요.

세분화 배근(rebar_layers) — 기둥/보/슬래브/전단벽/계단은 위에서 설명한 "대표 철근 1세트"
필드(main_rebar_size 등) 대신, 도면에 실제 배근이 방향/위치/구간별로 나뉘어 표시돼 있으면
그 각각을 rebar_layers 배열의 별도 항목으로 채우세요. 이 필드는 선택사항입니다 — 도면에서
구분이 안 보이면 비워두고 기존처럼 대표 필드만 채우면 됩니다(그 경우 계산은 기존 방식대로
동작합니다). rebar_layers를 채우면 그게 대표 필드보다 우선 사용되니, 채울 거면 그 부재의
배근을 빠짐없이 항목으로 나열해야 합니다(일부만 적으면 나머지 배근이 누락된 것으로 계산에서
빠집니다).
각 항목 형식: {"role": "...", "position": "...", "direction": "...", "strip": "...",
"zone": "...", "size": "D13", "spacing_m": 0.2, "count": 8, "has_hook": false, "note": "..."}
(size/spacing_m 또는 size/count 중 그 배근에 맞는 것만 채우고 나머지는 생략)
- role: 그 철근의 역할. 기둥="주근"/"후프"/"타이", 보="주근"/"스터럽", 슬래브="주근",
  전단벽="수직근"/"수평근"/"단부보강근"/"모서리보강근"/"교차부보강근"/"개구부보강근",
  계단="주근"/"배력근"
- position: "상부"/"하부"(보/슬래브/계단 주근) 또는 "수직"/"수평"(전단벽, role로 이미
  구분되면 생략 가능). 해당 없으면 생략.
- direction: 슬래브 주근의 배근 방향 — "X" 또는 "Y". 슬래브가 아니면 생략.
- strip: 슬래브 주근이 주열대(기둥 위 폭이 좁은 구간, 배근이 더 촘촘함)와 중간대(그 사이
  구간)로 나뉘어 있으면 "주열대" 또는 "중간대". 구분이 없으면 생략(전체 스팬에 동일 간격).
- zone: 구간 구분 — 보/기둥은 "단부"(부재 끝, 응력이 커서 배근이 촘촘한 구간) 또는
  "중앙부", 계단은 "계단참"(휴게참 슬래브 구간), 그 외 부재는 생략. 기둥 "후프"가 단부/
  중앙부로 간격이 다르면 두 항목으로 나눠 각각 zone과 spacing_m을 채우세요.
  주근(기둥 MAIN BAR가 여러 그룹, 예: 모서리근/중간근으로 지름이나 개수가 다르면) 항목을
  여러 개로 나눠 role="주근", note에 "모서리"/"중간" 등 구분을 적으세요.
- 기둥 "주근" 항목은 spacing_m 대신 count(그 그룹의 가닥수)를 채우세요(간격 배근이 아니라
  개수 배근이므로). 보/슬래브/전단벽/계단의 주근·배력근·수직수평근은 spacing_m(간격)을
  채우세요. 스터럽/후프/타이는 spacing_m을 채우세요.
- zone_length_m: zone이 "단부"인 항목(보 주근 감소배근, 기둥/보 단부 스터럽·후프 구간)에는
  그 구간의 부재길이방향 총 길이(m, 도면에 표기된 값 그대로)를 채우세요. 모르면 생략해도
  되며(생략 시 계산이 부재길이의 25%로 근사합니다), 아는 값이 있으면 반드시 채워서 근사
  오차를 줄이세요.
- has_hook: 그 배근의 정착이 표준갈고리면 true, 직선이면 false, 판단 안 되면 생략.

반드시 아래와 같은 키를 가진 JSON 객체 하나만 반환하세요. 다른 텍스트는 절대 포함하지 마세요.
값을 모르면 null, 해당 부재가 없으면 빈 배열([])로 두세요. 예시(형식 참고용, 실제 값 아님):
{
  "foundations": [{"mark": "F1", "length_m": 2.0, "width_m": 2.0, "thickness_m": 0.6, "count": 4, "rebar_size": "D16", "rebar_spacing_m": 0.2, "dowel_bar_size": "D25", "dowel_bar_count": 8, "dowel_has_hook": false, "zone": "지하1층", "floor_repeat_count": 1, "section": "101동", "bbox": {"page": 5, "box_2d": [120, 200, 260, 340]}}],
  "columns": [{"mark": "C1", "width_m": 0.5, "depth_m": 0.5, "height_m": 3.2, "count": 12, "main_rebar_size": "D25", "main_rebar_count": 8, "tie_rebar_size": "D10", "tie_spacing_m": 0.2, "has_hook": false, "zone": "기준층(2~15층)", "floor_repeat_count": 14, "section": "101동", "bbox": {"page": 12, "box_2d": [300, 410, 380, 490]}, "rebar_layers": [{"role": "주근", "size": "D25", "count": 4, "note": "모서리"}, {"role": "주근", "size": "D22", "count": 4, "note": "중간"}, {"role": "후프", "size": "D10", "spacing_m": 0.1, "zone": "단부"}, {"role": "후프", "size": "D10", "spacing_m": 0.2, "zone": "중앙부"}]}],
  "beams": [{"mark": "G1", "width_m": 0.4, "depth_m": 0.6, "length_m": 6.0, "count": 10, "main_rebar_size": "D22", "main_rebar_count": 6, "stirrup_size": "D10", "stirrup_spacing_m": 0.2, "has_hook": false, "is_top_bar": false, "zone": "기준층(2~15층)", "floor_repeat_count": 14, "section": "101동", "bbox": {"page": 12, "box_2d": [280, 300, 320, 600]}, "rebar_layers": [{"role": "주근", "position": "상부", "size": "D22", "count": 4, "zone": "단부"}, {"role": "주근", "position": "상부", "size": "D22", "count": 2, "zone": "중앙부"}, {"role": "주근", "position": "하부", "size": "D22", "count": 4}, {"role": "스터럽", "size": "D10", "spacing_m": 0.1, "zone": "단부"}, {"role": "스터럽", "size": "D10", "spacing_m": 0.2, "zone": "중앙부"}]}],
  "slabs": [{"mark": "SL1", "area_m2": 120.0, "thickness_m": 0.15, "count": 1, "rebar_size": "D13", "rebar_spacing_m": 0.2, "has_hook": false, "is_top_bar": false, "is_deck_slab": false, "openings": [{"label": "계단실 개구부", "width_m": 2.4, "height_m": 4.0, "count": 1}], "zone": "기준층(2~15층)", "floor_repeat_count": 14, "section": "101동", "bbox": {"page": 12, "box_2d": [100, 100, 700, 700]}, "rebar_layers": [{"role": "주근", "position": "상부", "direction": "X", "strip": "주열대", "size": "D13", "spacing_m": 0.15}, {"role": "주근", "position": "상부", "direction": "X", "strip": "중간대", "size": "D13", "spacing_m": 0.25}, {"role": "주근", "position": "하부", "direction": "X", "size": "D13", "spacing_m": 0.2}, {"role": "주근", "position": "상부", "direction": "Y", "size": "D13", "spacing_m": 0.2}, {"role": "주근", "position": "하부", "direction": "Y", "size": "D13", "spacing_m": 0.2}]}],
  "walls": [{"mark": "W1", "length_m": 5.0, "height_m": 3.2, "thickness_m": 0.2, "count": 2, "rebar_size": "D13", "rebar_spacing_m": 0.2, "has_hook": false, "end_condition": "모서리", "is_single_face": null, "openings": [{"label": "출입구", "width_m": 0.9, "height_m": 2.1, "count": 1}], "zone": "지하1층", "floor_repeat_count": 1, "section": "지하주차장", "bbox": {"page": 8, "box_2d": [400, 100, 900, 250]}, "rebar_layers": [{"role": "수직근", "size": "D13", "spacing_m": 0.2}, {"role": "수평근", "size": "D13", "spacing_m": 0.2}, {"role": "모서리보강근", "size": "D16", "count": 4}, {"role": "개구부보강근", "size": "D13", "count": 2, "note": "출입구 상하좌우"}]}],
  "stairs": [{"mark": "ST1", "width_m": 1.2, "length_m": 4.5, "thickness_m": 0.15, "count": 2, "rebar_size": "D13", "rebar_spacing_m": 0.2, "distribution_rebar_size": "D10", "distribution_rebar_spacing_m": 0.3, "is_top_bar": false, "has_hook": false, "zone": "1층", "floor_repeat_count": 1, "section": "101동", "bbox": {"page": 3, "box_2d": [500, 600, 650, 750]}, "rebar_layers": [{"role": "주근", "position": "하부", "size": "D13", "spacing_m": 0.2}, {"role": "주근", "position": "상부", "size": "D13", "spacing_m": 0.2, "zone": "계단참"}, {"role": "배력근", "size": "D10", "spacing_m": 0.3}]}],
  "notes": ["확인이 필요하거나 근사치인 항목에 대한 메모"],
  "general_spec": {"concrete_fck_mpa": 30, "rebar_grade": "SD500", "lap_splice_class": "B", "cover_thickness_mm": 40, "chair_bar_size": "D10", "chair_bar_height_m": 0.1, "concrete_fck_table": [{"category": "기초", "fck_mpa": 24}, {"category": "기둥", "fck_mpa": 30}, {"category": "보", "fck_mpa": 30}, {"category": "슬래브", "fck_mpa": 30}, {"category": "전단벽", "fck_mpa": 30}, {"category": "계단", "fck_mpa": 30}], "rebar_grade_table": [{"bar_size_max": 10, "grade": "SD500", "fy_mpa": 500}, {"bar_size_max": 999, "grade": "SD500S", "fy_mpa": 500}], "lap_splice_table": [{"bar_size": "D25", "length_m": 1.0, "position": "하부", "splice_class": "A"}, {"bar_size": "D25", "length_m": 1.3, "position": "하부", "splice_class": "B"}], "anchorage_table": [{"bar_size": "D25", "length_m": 1.0, "hook": false}]}
}"""

_EMPTY_MEMBERS = {
    "foundations": [], "columns": [], "beams": [], "slabs": [], "walls": [], "stairs": [],
    "notes": [], "general_spec": {},
}


# ─────────────────────────────────────────────
#  개요/구조일반사항 사전 확인 단계
#  — 전체 도면을 배치로 나눠 다 읽는 본 추출(extract_structural_members)은 비용/시간이
#  꽤 든다. 그 전에 "이 프로젝트가 뭔지, 구조 설계 기준이 뭔지"부터 사용자에게 맞는지
#  확인받아서, 애초에 도면을 잘못 이해한 채로 비싼 본 추출을 돌리는 걸 막기 위한
#  가벼운 사전 단계다. 표지(개요) 몇 페이지 + 구조일반사항 페이지만 뽑아서 1번만
#  Gemini에 보낸다 — 본 추출(배치 여러 번)에 비하면 토큰이 훨씬 적게 든다.
# ─────────────────────────────────────────────
OVERVIEW_SPEC_SYSTEM_PROMPT = """당신은 건축/구조도면의 표지(개요)와 구조일반사항을 판독하는 전문가입니다.
아래 두 가지만 "있는 그대로" 읽어서 구조화된 JSON으로 추출하세요. 절대 물량을 계산하지 마세요.
확인할 수 없는 값은 반드시 null(또는 빈 배열)로 두고, 절대 추정해서 채우지 마세요 — 확인 안 되는
항목은 각각 overview.unconfirmed_items / general_spec.unconfirmed_items에 짧은 설명으로 남기세요
(예: "구조형식", "정확한 동별 연면적", "D25 상부근 이음길이").

표/근거가 있는 값은 반드시 그 값을 어느 도면·페이지·표에서 봤는지 source 필드에 남기세요.
overview의 source는 반드시
{"pdf_type":"건축","page":3,"table":"동별자료","quote":"원문"}
객체(근거가 여러 개면 객체 배열)여야 합니다. 문자열 source는 허용되지 않습니다.
source를 모르면 null로 두세요.

1) 프로젝트 개요(overview) — 표지, 개요표, 건축계획개요 등에서:

⚠ 매우 중요한 주의사항: 도면목록/시트 인덱스(예: "A-001 도면목록", "A-101 지하1층 평면도" ~
"A-126 옥탑층 평면도"처럼 도면 번호가 나열된 표)는 그 프로젝트에 도면이 몇 장 있는지를 보여줄
뿐, 실제 층수·개요 정보가 아닙니다. 도면 번호의 마지막 숫자나 목록에 나열된 항목 개수를 보고
"지하 26개층"처럼 층수를 유추하지 마세요 — 이는 절대 금지된 추론입니다. basement_floor_count와
  buildings[].floor_count는 반드시 "사업개요", "건축계획개요", "동별 면적표"처럼 명시적으로 층수
자체를 나타내는 표/문구에 실제로 적힌 숫자만 사용하세요. 그런 표를 찾지 못했으면
도면목록으로 대신 유추하지 말고 반드시
null로 두고 unconfirmed_items에 남기세요.

- project_name: 사업개요의 "사업명칭" 셀에 명확한 값이 있으면 그 값을 채우세요.
  셀의 값은 글자 하나도 고치거나 생략하지 말고 원문 그대로 옮기세요. 특히 "번지" 같은
  단어를 임의로 삭제하지 마세요.
  PROJECT TITLE이 함께 보이면 공백·줄바꿈·'번지' 표기 차이를 감안한 보조 교차검증에만
  사용하고, 완전히 동일하지 않다는 이유만으로 사업명칭 셀의 값을 버리지 마세요.
- site_location, usage, structure_type, household_count, site_area_m2, building_area_m2,
  aboveground_floor_area_m2, basement_floor_area_m2, total_floor_area_m2: 해당 셀에 명시된
  값만 채우고 각각 sources에 근거를 남기세요.
- structure_type: "주요구조/규모"처럼 구조와 층수가 결합된 셀에 "철근콘크리트조"가
  보이면 structure_type에는 구조 부분인 "철근콘크리트조"를 추출하세요. 이때
  sources.structure_type의 quote에는 구조 부분만 잘라 쓰지 말고 "주요구조/규모"의
  라벨과 철근콘크리트조 및 지하·지상 층수까지 보이는 전체 셀 원문을 넣으세요.
- basement_floor_count: 지하층 수(정수). 위 주의사항대로 도면목록이 아닌 실제 개요표/면적표
  근거로만 채우세요.
- aboveground_max_floor: 지상 최고층(정수). "주요구조/규모", 동별면적표 또는 층별면적표에
  명시된 값으로만 채우세요.
- buildings: 동별자료/동별면적표에서 "숫자+동"인 모든 행을 각각 별도
  항목으로 추출하세요. "아파트", "공동주택" 같은 공통 용도명으로 합치지 마세요.
  각 항목에는 label, floor_range, floor_count, building_area_m2,
  total_floor_area_m2, household_count, source를 넣으세요.
  각 항목의 source는 문자열이 아니라 반드시
  {"pdf_type":"건축","page":8,"table":"동별자료","quote":"해당 행 전체 원문"}
  객체로 넣으세요. quote에는 그 행에서 값을 사용한 동명·층수·건축면적·연면적·세대수를
  실제로 보이는 범위에서 모두 포함하세요. 보이지 않는 값은 만들지 말고 해당 필드를 null로 두세요.
- underground_parking_note: 지하주차장 층 범위를 설명하는 문장(예: "지하 1~2층"). 확인 안 되면 null.
- amenity_facilities: 부대복리시설 표의 모든 행을 빠짐없이 추출하세요.
  대표 항목만 고르지 마세요. 면적이 없거나 작아 보이는 시설도
  면적이 없거나 작아 보여도 표에 행이 있으면 반드시 포함하고, 없는 숫자는 null로 두세요.
  각 행은 buildings와 같은 필드 형식으로 만들고 source는 문자열이 아니라 반드시
  {pdf_type, page, table, quote} 객체로 넣으세요. quote에는 그 행에서 값을 사용한
  시설명·층수·건축면적·연면적·세대수를 실제로 보이는 범위에서 모두 포함하세요.
- utility_facilities: 별도 설비 표가 있으면 동일한 행 객체 형식으로 모든 행을 추출하세요.
  각 행의 source 역시 문자열이 아닌 {pdf_type, page, table, quote} 객체여야 하며,
  quote에는 그 행에서 값을 사용한 시설명·층수·건축면적·연면적·세대수를 실제로
  보이는 범위에서 모두 포함하세요.
- commercial_note: 근린생활시설 관련 설명(예: "근린생활시설 1개 동"). 확인 안 되면 null.
- unconfirmed_items: 개요 중 도면에서 확인하지 못한 항목 목록(문자열 배열). 예: "구조형식", "정확한
  동별 연면적". 확인 못한 게 없으면 빈 배열.
- apartment_total_floor_area_m2: 동별자료에 공동주택/아파트 연면적 소계가 명시돼 있으면 그 값.
- conflicts: OCR 숫자가 표의 구성행 합계와 다르다고 보이면 추측으로 고치지 말고
  field, reported, calculated, formula, message를 가진 객체 배열로 남기세요.
  특히 동별·시설별 건축면적 합, 동별 연면적 합, 동별 세대수 합, 지상+지하 연면적을
  각각 표의 소계/합계와 반드시 검산하세요. 숫자가 다르게 보이면 원문 OCR값과
  계산값을 모두 conflict에 남기고 임의 확정하지 마세요.
- sources: project_name/basement_floor_count/aboveground_max_floor/buildings/underground_parking_note/
  amenity_facilities/utility_facilities/commercial_note 중 실제로 값을 채운 키에 대해서만,
  {"pdf_type":"건축","page":3,"table":"사업개요","quote":"사업명칭: OO 신축공사"}
  형식으로 실제 PDF 종류, 실제 1부터 시작하는 페이지 번호, 표 이름, 근거 원문을 담으세요.
  프로젝트명은 사업명칭 셀 근거 하나가 명확하면 인정하세요. PROJECT TITLE은 보이면
  보조 교차검증 근거로 추가할 수 있습니다. buildings처럼 여러 근거가 필요하면 객체 배열을 사용하세요.
  overview.sources.buildings, overview.sources.amenity_facilities,
  overview.sources.utility_facilities에는 각 행의 source와 동일한 {pdf_type, page, table, quote}
  객체를 행별로 빠짐없이 객체 배열로 다시 넣으세요. 문자열이나 문자열 배열은 허용되지 않습니다.
  각 quote에는 해당 행에서 사용한 동명/시설명·층수·건축면적·연면적·세대수를 실제로
  보이는 범위에서 모두 포함해야 합니다.
  값을 못 채운 키(즉 null이거나 unconfirmed_items에 있는 항목)는 sources에 넣지 마세요.
  근거를 알 수 없는 값을 확정적으로 채우지 마세요 — source를 댈 수 없으면 그 값 자체를
  null로 두고 unconfirmed_items에 넣는 것이 원칙입니다.

2) 구조일반사항(general_spec) — 구조일반사항/구조설계개요 페이지에서:
- concrete_fck_mpa: 가장 일반적인 대표 콘크리트강도(MPa)
- concrete_fck_table: 부위별 Fck가 다르면 각 행을 만드세요. category는 반드시 "기초", "기둥",
  "보", "슬래브", "전단벽", "계단" 6개 중 정확히 하나여야 합니다(계산 엔진이 이 6개 이름만
  그대로 매칭합니다 — "기초 콘크리트"나 "벽·기둥"처럼 다르게 쓰면 계산에 반영되지 않습니다).
  지하층/지상층처럼 위치에 따라 같은 부재 종류라도 Fck가 다르면, category는 그대로 두고
  zone_scope에 "지하" 또는 "지상"을 채워 같은 category로 행을 2개 만드세요(구분이 없으면
  zone_scope는 null). 예: 지하 기둥 30MPa/지상 기둥 27MPa면
  [{"category":"기둥","zone_scope":"지하","fck_mpa":30,"source":"..."},
   {"category":"기둥","zone_scope":"지상","fck_mpa":27,"source":"..."}] 형식으로. 표에 "전부재"
  처럼 여러 부재가 묶여 표기돼 있으면 해당하는 개별 category 각각에 같은 값으로 행을 만드세요.
- rebar_grade: 가장 일반적인 대표 철근강종(SD400/SD500 등)
- rebar_grade_table: 철근 지름별 강종이 다르면
  [{"bar_size_min": null, "bar_size_max": 10, "grade": "SD400", "fy_mpa": 400, "source": "S-002 철근재료표"},
   {"bar_size_min": 13, "bar_size_max": 25, "grade": "SD500", "fy_mpa": 500, "source": "S-002 철근재료표"}] 형식으로
  (bar_size_min/bar_size_max는 mm 숫자만, 하한이 없으면 null).
- lap_splice_class: 도면에 명시된 이음등급(A급/B급) — 참고용으로만 채우세요(실제 계산은 정책상 항상
  B급 기준입니다).
- lap_splice_table: 이음길이표가 있으면 철근 지름별·상부/하부별로 전부
  [{"bar_size": "D25", "position": "상부", "length_m": 1.5, "splice_class": "B", "source": "S-001 이음기준"}, ...]
  형식으로 옮기세요. A급/B급이 각각 다른 열에 있으면 각각 행으로 나눠 splice_class를 채우세요.
- anchorage_table: 정착길이표가 있으면 철근 지름별·상부/하부별·직선/갈고리별로 전부
  [{"bar_size": "D25", "position": "상부", "hook": false, "length_m": 1.1, "source": "..."}, ...] 형식으로.
- cover_thickness_mm: 대표(가장 얇은) 피복두께(mm)
- cover_table: 부위별 피복두께가 다르면 각 행을 만드세요. category는 concrete_fck_table과
  동일하게 "기초"/"기둥"/"보"/"슬래브"/"전단벽"/"계단" 6개 중 정확히 하나로 쓰세요(예:
  "벽·기둥"처럼 묶어 쓰지 말고 "기둥"과 "전단벽" 각각 행으로 나누세요).
  [{"category": "기초", "thickness_mm": 80, "source": "S-001 피복표"}, {"category": "전단벽", "thickness_mm": 40, "source": "..."}] 형식으로.
- seismic_rebar_rules: 내진철근이 적용되는 부위/규격을 구조화해서
  [{"location": "기둥·보 접합부", "grade": "SD500S", "source": "S-002 내진상세"}] 형식으로. 확인 안 되면 빈 배열.
- summary_notes: 위 내용을 근거로 한 짧은 요약 문장들(문자열 배열). 예: "본 프로젝트는 B급 겹침이음을
  적용합니다.", "상부근과 하부근의 이음길이가 구분되어 있습니다.", "벽체는 양면 배근이며, 보 단부는
  중앙부보다 스터럽 간격이 조밀합니다." — 도면 근거가 있는 내용만 쓰세요, 지어내지 마세요.
- unconfirmed_items: 구조일반사항 중 도면에서 확인하지 못한 항목 목록(문자열 배열). 특히 특정 철근
  지름·위치 조합의 이음/정착길이가 표에 없으면 "D25 상부근 이음길이"처럼 구체적으로 남기세요 —
  이 목록에 있는 항목은 사용자가 직접 값을 채우기 전까지 철근 실수량 계산에서 제외됩니다.

사용자가 이전 확인 단계에서 정정한 내용(있다면 아래 [사용자 확인/정정 내용]으로 주어집니다)이 있으면
그 내용을 최우선 근거로 삼아 반영하세요 — 도면에서 다르게 보이더라도 사용자가 이미 확인한 내용을
우선하세요.

반드시 아래 키를 가진 JSON 객체 하나만 반환하세요. 다른 텍스트는 절대 포함하지 마세요.
예시(형식 참고용, 실제 값 아님):
{
  "overview": {
    "project_name": "OO 공동주택 신축공사",
    "structure_type": "철근콘크리트조",
    "basement_floor_count": 2,
    "buildings": [
      {"label": "101동", "floor_range": "지상1층~지상20층", "floor_count": 20,
       "building_area_m2": 600.1, "total_floor_area_m2": 5000.2, "household_count": 60,
       "source": {"pdf_type":"건축","page":8,"table":"동별자료","quote":"101동 | 지상1층~지상20층 | 건축면적 600.1㎡ | 연면적 5,000.2㎡ | 60세대"}},
      {"label": "102동", "floor_range": "지상1층~지상18층", "floor_count": 18,
       "building_area_m2": 580.3, "total_floor_area_m2": 4700.4, "household_count": 55,
       "source": {"pdf_type":"건축","page":8,"table":"동별자료","quote":"102동 | 지상1층~지상18층 | 건축면적 580.3㎡ | 연면적 4,700.4㎡ | 55세대"}}
    ],
    "underground_parking_note": "지하 1~2층",
    "amenity_facilities": [
      {"label":"관리사무소","floor_range":"지상1층","floor_count":1,
       "building_area_m2":30.0,"total_floor_area_m2":30.0,"household_count":null,
       "source":{"pdf_type":"건축","page":7,"table":"부대복리시설 설치계획","quote":"관리사무소 | 지상1층 | 건축면적 30.0㎡ | 연면적 30.0㎡"}}
    ],
    "utility_facilities": [
      {"label":"전기실","floor_range":"지하1층","floor_count":1,
       "building_area_m2":null,"total_floor_area_m2":45.0,"household_count":null,
       "source":{"pdf_type":"건축","page":7,"table":"설비시설","quote":"전기실 | 지하1층 | 연면적 45.0㎡"}}
    ],
    "commercial_note": "근린생활시설 1개 동",
    "unconfirmed_items": ["구조형식", "정확한 동별 연면적"],
    "sources": {
      "project_name": {"pdf_type":"건축","page":7,"table":"사업개요","quote":"사업명칭: OO 공동주택 신축공사"},
      "structure_type": {"pdf_type":"건축","page":7,"table":"사업개요","quote":"주요구조/규모: 철근콘크리트조, 지하2층 / 지상20층"},
      "basement_floor_count": {"pdf_type":"건축","page":7,"table":"사업개요","quote":"주요구조/규모: 철근콘크리트조, 지하2층 / 지상20층"},
      "buildings": [
        {"pdf_type":"건축","page":8,"table":"동별자료","quote":"101동 | 지상1층~지상20층 | 건축면적 600.1㎡ | 연면적 5,000.2㎡ | 60세대"},
        {"pdf_type":"건축","page":8,"table":"동별자료","quote":"102동 | 지상1층~지상18층 | 건축면적 580.3㎡ | 연면적 4,700.4㎡ | 55세대"}
      ],
      "amenity_facilities": [
        {"pdf_type":"건축","page":7,"table":"부대복리시설 설치계획","quote":"관리사무소 | 지상1층 | 건축면적 30.0㎡ | 연면적 30.0㎡"}
      ],
      "utility_facilities": [
        {"pdf_type":"건축","page":7,"table":"설비시설","quote":"전기실 | 지하1층 | 연면적 45.0㎡"}
      ]
    }
  },
  "general_spec": {
    "concrete_fck_mpa": 27,
    "concrete_fck_table": [
      {"category": "기초", "zone_scope": null, "fck_mpa": 24, "source": "S-001 3페이지"},
      {"category": "기둥", "zone_scope": "지하", "fck_mpa": 30, "source": "S-001 3페이지"},
      {"category": "기둥", "zone_scope": "지상", "fck_mpa": 27, "source": "S-001 3페이지"},
      {"category": "전단벽", "zone_scope": "지하", "fck_mpa": 30, "source": "S-001 3페이지"},
      {"category": "전단벽", "zone_scope": "지상", "fck_mpa": 27, "source": "S-001 3페이지"},
      {"category": "보", "zone_scope": null, "fck_mpa": 27, "source": "S-001 4페이지"},
      {"category": "슬래브", "zone_scope": null, "fck_mpa": 27, "source": "S-001 4페이지"}
    ],
    "rebar_grade": "SD500",
    "rebar_grade_table": [
      {"bar_size_min": null, "bar_size_max": 10, "grade": "SD400", "fy_mpa": 400, "source": "S-002 철근재료표"},
      {"bar_size_min": 13, "bar_size_max": 25, "grade": "SD500", "fy_mpa": 500, "source": "S-002 철근재료표"}
    ],
    "lap_splice_class": "B",
    "lap_splice_table": [{"bar_size": "D25", "position": "상부", "length_m": 1.5, "splice_class": "B", "source": "S-001 이음기준"}],
    "anchorage_table": [{"bar_size": "D25", "position": "상부", "hook": false, "length_m": 1.1, "source": "S-001 이음기준"}],
    "cover_thickness_mm": 40,
    "cover_table": [{"category": "기초", "thickness_mm": 80, "source": "S-001 피복표"}, {"category": "기둥", "thickness_mm": 40, "source": "S-001 피복표"}, {"category": "전단벽", "thickness_mm": 40, "source": "S-001 피복표"}],
    "seismic_rebar_rules": [{"location": "기둥·보 접합부", "grade": "SD500S", "source": "S-002 내진상세"}],
    "summary_notes": [
      "본 프로젝트는 B급 겹침이음을 적용합니다.",
      "상부근과 하부근의 이음길이가 구분되어 있습니다.",
      "벽체는 양면 배근이며, 보 단부는 중앙부보다 스터럽 간격이 조밀합니다."
    ],
    "unconfirmed_items": ["D25 상부근 이음길이"]
  },
  "notes": ["확인이 필요하거나 근사치인 항목에 대한 메모"]
}"""

_EMPTY_OVERVIEW_SPEC = {
    "overview": {
    "project_name": None, "basement_floor_count": None, "buildings": [],
        "aboveground_max_floor": None,
        "underground_parking_note": None, "amenity_facilities": [], "utility_facilities": [],
        "commercial_note": None, "unconfirmed_items": [],
    },
    "general_spec": {},
    "notes": [],
}

OVERVIEW_CHECK_MAX_PAGES = 12  # 사업개요 관련 최대 6장 + 구조일반사항 최대 6장 정도로 토큰을 억제한다

# 실제 프로젝트에서 확인된 문제: 무조건 앞 2페이지만 "표지"로 간주해서 보냈더니, 표지/목차/
# 조감도가 앞쪽에 오고 정작 사업명·대지위치·층별면적·부대복리시설이 담긴 "사업개요표"는
# 3페이지 이후에 있는 문서에서 프로젝트명/지하층수/부대시설 목록이 통째로 "확인 안 됨"으로
# 나왔다(그 페이지 자체를 Gemini에 아예 보내지 않았으므로 당연한 결과). _detect_schedule_
# pages와 같은 방식으로 "사업개요" 관련 키워드가 있는 페이지를 먼저 찾아서 우선 포함한다.
OVERVIEW_CLASSIFY_DPI = 120
OVERVIEW_CLASSIFY_MIN_LONG_EDGE = 2000
OVERVIEW_CLASSIFY_MAX_LONG_EDGE = 2800
# 실제 프로젝트들에서 사업개요/동별면적표는 거의 항상 앞쪽 30쪽 안에 있었다(2026-07-27
# 사용자 확인: "보통 30쪽 이상은 안 간다"). 이전에는 이 값이 정의만 되고 실제 호출부
# (extract_overview_and_spec -> _find_incremental_overview_pages)에 전달되지 않아서
# 사실상 페이지 수 제한 없이 문서 끝까지(120초 시간제한에만 걸려서) 스캔하고 있었다.
# 아래에서 실제로 max_pages로 전달해 30쪽 이후는 스캔하지 않도록 한다.
OVERVIEW_CLASSIFY_MAX_PAGES = int(os.environ.get("QTY_OVERVIEW_CLASSIFY_MAX_PAGES", "30"))
OVERVIEW_LOCATOR_VERSION = "vision-evidence-validated-v12"
OVERVIEW_VISION_TIMEOUT_SEC = 60
OVERVIEW_LOCATOR_TIMEOUT_SEC = 120


class OverviewLocatorTimeout(RuntimeError):
    pass
OVERVIEW_PAGE_TYPES = {
    "overview", "building_area_table", "drawing_list", "other",
}
OVERVIEW_PAGE_CLASSIFIER_PROMPT = """당신은 건축 PDF 페이지 종류 분류기입니다.
값을 추출하거나 층수·동·면적을 해석하지 말고 각 이미지의 페이지 종류만 분류하세요.
page_type은 overview / building_area_table / drawing_list / other 중 하나입니다.
각 페이지는 오직 그 이미지 자체의 제목, 표 머리글, 셀 내용만 보고 판정하세요.
페이지 번호, 앞뒤 페이지, 문서 내 위치, 다른 페이지의 판정 결과로 종류를 추정하지 마세요.
overview는 사업개요 또는 건축개요 제목과 함께 사업명칭, 대지위치, 주용도,
구조·규모, 면적 항목이 실제 표에 나타나는 페이지입니다.
building_area_table은 동별면적표·층별면적표 등의 제목과 동 구분 및
건축면적·연면적 열이 실제 표에 나타나는 페이지입니다.
도면번호 A-001 등 시트 번호와 도면명이 나열된 목록 페이지는 drawing_list입니다.
'건축개요'라는 도면명이 목록 일부에 있어도 실제 개요표가 아니면 overview가 아닙니다.
각 결과에 page_number, page_type, confidence, title_text, evidence_terms를 반환하세요.
evidence_terms는 실제 이미지에서 읽힌 판정 근거 표제·셀명·열 이름만 문자열 배열로 넣으세요.
각 페이지의 evidence_terms는 최대 6개의 짧은 용어 또는 셀 근거만 넣고,
설명문·판단 과정·장문 문장은 반환하지 마세요.
building_area_table로 판정한 경우 실제 표에서 보이는 101동·102동 같은 독립 동명 셀도
evidence_terms에 포함하세요. 제목이 불명확하면 건축면적·연면적 머리글과 독립 동명 셀을
반드시 함께 반환하고, 실제로 보이지 않는 동명이나 숫자는 만들지 마세요.
반드시 {"pages":[{"page_number":1,"page_type":"overview","confidence":0.98,
"title_text":"사업개요","evidence_terms":["사업명칭","대지위치","주용도","규모"]}]} JSON만 반환하세요.
그 외 키와 값 추출 결과는 반환하지 마세요."""

_OVERVIEW_CLASSIFICATION_CACHE = {}
_OVERVIEW_CLASSIFICATION_LOCK = threading.Lock()


class _OverviewClassificationResult(list):
    """기존 list 호출부와 호환하면서 locator 응답 완전성 메타를 전달한다."""

    def __init__(self, pages=(), *, requested_pages=(), finish_reason=None,
                 missing_page_numbers=(), partial_repair_used=False):
        super().__init__(pages)
        self.requested_pages = list(requested_pages)
        self.finish_reason = finish_reason
        self.missing_page_numbers = list(missing_page_numbers)
        self.partial_repair_used = bool(partial_repair_used)


def _incremental_overview_ranges(total_pages, max_pages=None):
    end_limit = total_pages if max_pages is None else min(total_pages, max_pages)
    starts_and_sizes = ((1, 10), (11, 5), (16, 5))
    for start, size in starts_and_sizes:
        if start <= end_limit:
            yield start, min(start + size - 1, end_limit)
    start = 21
    while start <= end_limit:
        yield start, min(start + 9, end_limit)
        start += 10


def _prepare_overview_classifier_image(image):
    image = image.convert("RGB")
    long_edge = max(image.size)
    if long_edge < OVERVIEW_CLASSIFY_MIN_LONG_EDGE:
        scale = OVERVIEW_CLASSIFY_MIN_LONG_EDGE / float(long_edge)
    elif long_edge > OVERVIEW_CLASSIFY_MAX_LONG_EDGE:
        scale = OVERVIEW_CLASSIFY_MAX_LONG_EDGE / float(long_edge)
    else:
        return image
    size = (
        max(1, int(round(image.width * scale))),
        max(1, int(round(image.height * scale))),
    )
    return image.resize(size, Image.Resampling.LANCZOS)


def _classify_overview_page_batch(pdf_bytes, page_numbers, needed_types=None):
    """페이지를 독립 이미지로 렌더링하되 단 한 번의 Vision 요청으로 분류한다."""
    client = get_gemini_client()
    if client is None or not page_numbers:
        return []
    contents = [
        "각 이미지의 페이지 종류를 그 페이지 자체 내용만으로 분류하세요. "
        "값을 추출하지 말고 모든 이미지에 대해 page_number, page_type, confidence, "
        "title_text, evidence_terms만 반환하세요."
    ]
    for page_num in page_numbers:
        try:
            images = _render_pdf_page_range(
                pdf_bytes, page_num, page_num, dpi=OVERVIEW_CLASSIFY_DPI,
                timeout=OVERVIEW_VISION_TIMEOUT_SEC,
            )
        except Exception as exc:
            logger.warning(
                "quantity_overview_locator_render_failed page=%s error=%s",
                page_num, str(exc)[:160],
            )
            continue
        if not images:
            continue
        image = _prepare_overview_classifier_image(images[0])
        logger.info(
            "[OVERVIEW_LOCATOR] page=%s image_pixels=%sx%s",
            page_num, image.width, image.height,
        )
        contents.append(f"[PDF_PAGE={page_num}]")
        contents.append(types.Part.from_bytes(
            data=image_to_jpeg_bytes(
                image,
                max_size=(OVERVIEW_CLASSIFY_MAX_LONG_EDGE, OVERVIEW_CLASSIFY_MAX_LONG_EDGE),
            ),
            mime_type="image/jpeg",
        ))
    if len(contents) == 1:
        logger.warning("[OVERVIEW_LOCATOR] no_page_images_rendered pages=%s", page_numbers)
        return []
    response = client.models.generate_content(
        model=GEMINI_QUANTITY_MODEL,
        contents=contents,
        config=types.GenerateContentConfig(
            system_instruction=OVERVIEW_PAGE_CLASSIFIER_PROMPT,
            # 실제 사용자 문서에서 재현된 버그: 10페이지 배치(첫 배치 1-10쪽, 이후 21쪽부터의
            # 10쪽 단위 배치)는 페이지당 title_text/evidence_terms까지 포함하면 2048 토큰을
            # 쉽게 넘겨 응답이 중간에 잘렸다. json.loads가 실패하면 무조건 빈 리스트를
            # 반환했으므로, 그 배치의 모든 페이지가 "다시 시도 없이 통째로 other/미발견"
            # 처리됐다(5쪽 배치인 11-15/16-20은 토큰이 부족하지 않아 정상 동작했음 —
            # 로그에서 10쪽 배치만 raw_response가 중간에 끊긴 것으로 확인됨). 5쪽 배치
            # 기준 실측 크기에 여유를 크게 둬서 10쪽 배치도 안전하게 담기도록 올린다.
            response_mime_type="application/json", temperature=0.0, max_output_tokens=8192,
            thinking_config=types.ThinkingConfig(thinking_budget=512),
            http_options=types.HttpOptions(timeout=OVERVIEW_VISION_TIMEOUT_SEC * 1000),
        ),
    )
    raw = _extract_text_from_gemini_response(response).replace("```json", "").replace("```", "").strip()
    logger.info("[OVERVIEW_LOCATOR] raw_response=%s", raw[:12000])
    partial_repair_used = False
    try:
        payload = json.loads(raw)
    except (TypeError, ValueError):
        # 토큰 상향 조정 후에도 혹시 잘리는 경우를 대비한 안전망 — 끝까지 못 받았어도
        # 앞부분에 완전히 끝난 페이지 항목까지는 있는 그대로 살려서 쓴다(전체를 버리고
        # "전부 못 찾음" 처리하는 것보다 일부라도 건지는 편이 낫다).
        payload = _try_repair_truncated_json(raw)
        if payload is None:
            logger.warning(
                "[OVERVIEW_LOCATOR] json_parse_failed_even_after_repair pages=%s raw_len=%s",
                page_numbers, len(raw),
            )
            payload = {"pages": []}
        else:
            partial_repair_used = True
            logger.warning(
                "[OVERVIEW_LOCATOR] json_truncated_partial_repair_used pages=%s raw_len=%s",
                page_numbers, len(raw),
            )
    pages = payload.get("pages", []) if isinstance(payload, dict) else payload
    if not isinstance(pages, list):
        pages = []
    requested_page_numbers = [int(page) for page in page_numbers]
    requested_set = set(requested_page_numbers)
    parsed_page_numbers = sorted({
        int(item.get("page_number"))
        for item in pages
        if isinstance(item, dict)
        and str(item.get("page_number") or "").strip().isdigit()
        and int(item.get("page_number")) in requested_set
    })
    missing_page_numbers = [
        page for page in requested_page_numbers if page not in set(parsed_page_numbers)
    ]
    response_diagnostics = _gemini_response_diagnostics(response)
    candidate_diagnostics = response_diagnostics.get("candidates") or []
    finish_reason = (
        candidate_diagnostics[0].get("finish_reason")
        if candidate_diagnostics else None
    )
    usage = response_diagnostics.get("usage") or {}
    _log_overview_diagnostic(
        "locator_vision_response",
        requested_page_numbers=requested_page_numbers,
        finish_reason=finish_reason,
        prompt_token_count=usage.get("prompt_token_count"),
        thoughts_token_count=usage.get("thoughts_token_count"),
        candidates_token_count=usage.get("candidates_token_count"),
        total_token_count=usage.get("total_token_count"),
        raw_length=len(raw),
        parsed_page_numbers=parsed_page_numbers,
        missing_page_numbers=missing_page_numbers,
        partial_json_repair_used=partial_repair_used,
    )
    if missing_page_numbers:
        _log_overview_diagnostic(
            "locator_vision_incomplete_response",
            level=logging.WARNING,
            requested_page_numbers=requested_page_numbers,
            finish_reason=finish_reason,
            missing_page_numbers=missing_page_numbers,
            partial_json_repair_used=partial_repair_used,
        )
    logger.info("[OVERVIEW_LOCATOR] decisions=%s", json.dumps(pages, ensure_ascii=False))
    return _OverviewClassificationResult(
        pages,
        requested_pages=requested_page_numbers,
        finish_reason=finish_reason,
        missing_page_numbers=missing_page_numbers,
        partial_repair_used=partial_repair_used,
    )


def _normalize_page_classification(item, allowed_pages):
    if not isinstance(item, dict):
        return None
    try:
        page_number = int(item.get("page_number"))
        confidence = float(item.get("confidence", 0))
    except (TypeError, ValueError):
        return None
    page_type = str(item.get("page_type") or "other")
    if page_number not in allowed_pages or page_type not in OVERVIEW_PAGE_TYPES:
        return None
    return {
        "page_number": page_number,
        "page_type": page_type,
        "confidence": max(0.0, min(1.0, confidence)),
        "title_text": str(item.get("title_text") or "")[:200],
        "evidence_terms": [
            str(term).strip()[:100]
            for term in (
                item.get("evidence_terms")
                if isinstance(item.get("evidence_terms"), list)
                else [item.get("evidence_terms")]
            )
            if str(term or "").strip()
        ][:6],
    }


def _classification_has_evidence(item):
    """높은 confidence만으로 표 종류를 확정하지 않고 실제 표제와 판정 근거를 확인한다."""
    page_type = item.get("page_type")
    title = re.sub(r"\s+", "", str(item.get("title_text") or ""))
    terms = {
        re.sub(r"\s+", "", str(term))
        for term in item.get("evidence_terms") or []
        if str(term).strip()
    }
    if page_type == "overview":
        title_ok = any(value in title for value in ("사업개요", "건축개요", "건축계획개요"))
        evidence_keys = ("사업명칭", "사업명", "대지위치", "주용도", "용도", "구조", "규모", "면적")
        return title_ok and sum(any(key in term for term in terms) for key in evidence_keys) >= 2
    if page_type == "building_area_table":
        title_ok = any(value in title for value in ("동별면적", "층별면적", "동별자료"))
        evidence_keys = ("동", "층", "건축면적", "연면적")
        standard_evidence = sum(
            any(key in term for term in terms) for key in evidence_keys
        ) >= 2
        if title_ok:
            return standard_evidence

        combined = " ".join([title, *terms])
        if "도면목록" in combined.replace(" ", "") or "DRAWINGLIST" in combined.replace(" ", "").upper():
            return False
        building_labels = {
            match.group(0).replace(" ", "")
            for match in re.finditer(r"(?<!\d)\d{1,4}\s*동(?!\w)", combined)
        }
        has_area_headers = (
            any("건축면적" in term for term in terms)
            and any("연면적" in term for term in terms)
        )
        if not has_area_headers:
            return False
        if len(building_labels) >= 2:
            return True
        if len(building_labels) == 1:
            has_row_context = any(
                key in combined for key in ("세대수", "세대", "층수", "지상", "지하")
            )
            has_area_number = bool(re.search(r"\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+\.\d+", combined))
            return has_row_context and has_area_number
        return False
    return True


def _cache_validated_overview_pages(pdf_bytes, page_detection, overview):
    """실제 핵심 개요값까지 검증된 경우에만 locator 성공 페이지를 캐시한다."""
    if not page_detection.get("complete"):
        return False
    if (
        not overview.get("project_name")
        or overview.get("basement_floor_count") is None
        or overview.get("aboveground_max_floor") is None
        or not any(
            isinstance(building, dict) and building.get("label")
            for building in overview.get("buildings") or []
        )
    ):
        return False
    overview_page = page_detection.get("overview", {}).get("page_number")
    area_table_page = page_detection.get("area_table", {}).get("page_number")
    if not isinstance(overview_page, int) or not isinstance(area_table_page, int):
        return False
    cache_key = f"{OVERVIEW_LOCATOR_VERSION}:{_sha256_bytes(pdf_bytes)}"
    with _OVERVIEW_CLASSIFICATION_LOCK:
        _OVERVIEW_CLASSIFICATION_CACHE[cache_key] = {
            "overview_page": overview_page,
            "area_table_page": area_table_page,
        }
    return True


_OVERVIEW_TEXT_KEYWORDS = {
    "overview": (
        ("사업개요", 8), ("건축개요", 6), ("사업명칭", 7),
        ("주요구조", 4), ("규모", 2),
    ),
    "building_area_table": (
        ("동별면적표", 10), ("층별면적표", 8), ("동별 면적", 8),
        ("동별자료", 8), ("101동", 3), ("102동", 3),
    ),
}


def _extract_pdf_page_texts(pdf_bytes, first_page=1, last_page=None, timeout_sec=20):
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as handle:
            handle.write(pdf_bytes)
            tmp_path = handle.name
        command = ["pdftotext", "-layout"]
        if first_page is not None:
            command += ["-f", str(first_page)]
        if last_page is not None:
            command += ["-l", str(last_page)]
        command += [tmp_path, "-"]
        result = subprocess.run(
            command,
            capture_output=True, timeout=timeout_sec,
        )
        if result.returncode != 0:
            return {}
        pages = result.stdout.decode("utf-8", errors="ignore").split("\f")
        return {
            index: text
            for index, text in enumerate(pages, start=first_page or 1)
            if text.strip()
        }
    except Exception:
        return {}
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def _text_page_score(text, page_type):
    compact = re.sub(r"\s+", "", str(text or ""))
    score = 0
    hits = []
    for keyword, weight in _OVERVIEW_TEXT_KEYWORDS[page_type]:
        if keyword in text or keyword.replace(" ", "") in compact:
            score += weight
            hits.append(keyword)
    if "도면목록" in compact or "DRAWINGLIST" in compact.upper():
        score -= 8
    return score, hits


def _log_locator_page(page_number, text_score, predicted_type, confidence, selection_reason,
                      scan_range="text-all"):
    logger.info(
        "quantity_overview_locator %s",
        json.dumps({
            "locator_version": OVERVIEW_LOCATOR_VERSION,
            "range": scan_range,
            "page_number": page_number,
            "text_score": text_score,
            "predicted_type": predicted_type,
            "confidence": confidence,
            "selection_reason": selection_reason,
        }, ensure_ascii=False, sort_keys=True),
    )


def _log_overview_diagnostic(event, level=logging.INFO, **details):
    """사업개요 탐색의 실행 흐름만 JSON 한 줄로 남기는 진단 전용 로그."""
    payload = {
        "event": event,
        "locator_version": OVERVIEW_LOCATOR_VERSION,
    }
    payload.update(details)
    logger.log(
        level,
        "quantity_overview_diagnostic %s",
        json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str),
    )


def _overview_diagnostic_missing_fields(overview):
    """판독 결과에서 비어 있는 주요 개요 필드를 진단용으로만 계산한다."""
    overview = overview if isinstance(overview, dict) else {}
    missing = []
    for field in (
        "project_name",
        "site_location",
        "usage",
        "structure_type",
        "basement_floor_count",
        "aboveground_max_floor",
        "household_count",
    ):
        value = overview.get(field)
        if value is None or value == "":
            missing.append(field)
    if not overview.get("buildings"):
        missing.append("buildings")
    if not overview.get("amenity_facilities"):
        missing.append("amenity_facilities")
    return missing


def _select_text_locator_candidates(page_texts, page_numbers, selections, scan_range):
    diagnostics = []
    for page_number in page_numbers:
        text = page_texts.get(page_number, "")
        overview_score, overview_hits = _text_page_score(text, "overview")
        area_score, area_hits = _text_page_score(text, "building_area_table")
        predicted_type = "other"
        text_score = max(overview_score, area_score)
        hits = []
        if overview_score >= 7 and overview_score >= area_score:
            predicted_type, hits = "overview", overview_hits
        elif area_score >= 6:
            predicted_type, hits = "building_area_table", area_hits
        reason = "text_keywords:" + ",".join(hits) if hits else "text_no_match"
        diagnostic = {
            "page_number": page_number, "page_type": predicted_type,
            "confidence": 1.0 if predicted_type != "other" else 0.0,
            "title_text": "", "text_score": text_score, "selection_reason": reason,
        }
        diagnostics.append(diagnostic)
        _log_locator_page(
            page_number, text_score, predicted_type, diagnostic["confidence"], reason, scan_range,
        )
        if predicted_type == "overview":
            current = selections["overview"]
            if current is None or text_score > current["text_score"]:
                selections["overview"] = diagnostic
        elif predicted_type == "building_area_table":
            current = selections["area_table"]
            if current is None or text_score > current["text_score"]:
                selections["area_table"] = diagnostic
    return selections, diagnostics


def _find_incremental_overview_pages(pdf_bytes, total_pages, classify_batch=None,
                                     max_pages=None,
                                     page_texts=None, progress_callback=None):
    """범위별 텍스트 점수 후 부족한 페이지만 Vision으로 보완하고 즉시 중단한다."""
    classify_batch = classify_batch or _classify_overview_page_batch
    cache_key = f"{OVERVIEW_LOCATOR_VERSION}:{_sha256_bytes(pdf_bytes)}"
    with _OVERVIEW_CLASSIFICATION_LOCK:
        cached_result = _OVERVIEW_CLASSIFICATION_CACHE.get(cache_key)
    if cached_result:
        _log_overview_diagnostic(
            "locator_cache_hit",
            selected_pages={
                "overview": cached_result["overview_page"],
                "area_table": cached_result["area_table_page"],
            },
        )
        return {
            "overview": {
                "page_number": cached_result["overview_page"], "page_type": "overview",
                "confidence": 1.0, "title_text": "", "text_score": 0,
                "selection_reason": "versioned_success_cache",
            },
            "area_table": {
                "page_number": cached_result["area_table_page"],
                "page_type": "building_area_table", "confidence": 1.0,
                "title_text": "", "text_score": 0,
                "selection_reason": "versioned_success_cache",
            },
            "classifications": [], "top_candidates": [], "scanned_pages": [],
            "complete": True, "from_cache": True,
            "vision_call_count": 0, "vision_pages": [],
        }

    supplied_page_texts = page_texts
    locator_started = time.monotonic()
    deadline = locator_started + OVERVIEW_LOCATOR_TIMEOUT_SEC

    def remaining_seconds():
        return deadline - time.monotonic()

    def ensure_time(phase):
        if remaining_seconds() <= 0:
            _log_overview_diagnostic(
                "locator_timeout",
                level=logging.WARNING,
                phase=phase,
                elapsed_seconds=round(time.monotonic() - locator_started, 3),
                selected_pages={
                    key: (value or {}).get("page_number")
                    for key, value in selections.items()
                },
                scanned_pages=list(scanned_pages),
                vision_call_count=vision_call_count,
            )
            raise OverviewLocatorTimeout(
                f"사업개요 페이지 자동 탐색이 {OVERVIEW_LOCATOR_TIMEOUT_SEC}초를 초과했습니다 "
                f"(단계: {phase})."
            )

    selections = {"overview": None, "area_table": None}
    diagnostics = []
    scanned_pages = []
    vision_call_count = 0
    vision_pages = []
    _log_overview_diagnostic(
        "locator_start",
        total_pages=total_pages,
        max_pages=max_pages,
        timeout_seconds=OVERVIEW_LOCATOR_TIMEOUT_SEC,
    )
    for first_page, last_page in _incremental_overview_ranges(total_pages, max_pages):
        page_numbers = list(range(first_page, last_page + 1))
        scan_range = f"{first_page}-{last_page}"
        ensure_time(f"{scan_range} 텍스트 검색")
        if progress_callback:
            progress_callback(f"개요 페이지 탐색 · {scan_range}쪽 · 텍스트 검색")
        current_page_texts = (
            _extract_pdf_page_texts(
                pdf_bytes, first_page, last_page,
                timeout_sec=max(1, min(20, int(remaining_seconds()))),
            )
            if supplied_page_texts is None else supplied_page_texts
        )
        selections, text_diagnostics = _select_text_locator_candidates(
            current_page_texts, page_numbers, selections, scan_range,
        )
        diagnostics.extend(text_diagnostics)
        scanned_pages.extend(page_numbers)
        vision_types = {"overview", "building_area_table"}
        # 텍스트 점수는 후보 생성과 로그에만 사용한다. 두 후보가 모두 있어도 같은 배치의
        # 독립 이미지 Vision 분류를 생략하지 않는다. 최종 종류는 페이지 자체 콘텐츠를
        # 읽은 Vision 결과로 검증한다.
        # 이 배치에서 문서 종류가 하나라도 부족하면 같은 배치의 모든 페이지를 Vision
        # 요청에 넣는다. 텍스트 판정 여부나 페이지 위치 때문에 이미지를 제외하지 않는다.
        # 각 페이지는 독립 이미지와 정확한 PDF_PAGE 라벨로 자체 내용만 분류한다.
        diagnostic_by_page = {
            item["page_number"]: item for item in text_diagnostics
        }
        pages_to_classify = list(page_numbers)
        # 같은 범위의 Vision 단계가 실제로 시작됐음을 응답 전에 기록한다. 이 호출이 끝나기
        # 전에는 아래 for-loop가 다음 범위로 진행될 수 없다.
        for page_number in pages_to_classify:
            _log_locator_page(
                page_number, diagnostic_by_page[page_number]["text_score"],
                "other", 0.0, "image_fallback_queued", scan_range,
            )
        vision_call_count += 1
        vision_pages.append(list(pages_to_classify))
        ensure_time(f"{scan_range} Vision 시작")
        if progress_callback:
            progress_callback(f"개요 페이지 탐색 · {scan_range}쪽 · Vision 분류")
        vision_started = time.monotonic()
        logger.info("[OVERVIEW_LOCATOR] pages=%s phase=vision start", scan_range)
        logger.info(
            "quantity_overview_locator_vision_call call=%s pages=%s range=%s",
            vision_call_count, pages_to_classify, scan_range,
        )
        try:
            returned = classify_batch(pdf_bytes, pages_to_classify, vision_types) or []
        except TypeError:
            # 기존 테스트/내부 주입 함수와의 호환.
            try:
                returned = classify_batch(pdf_bytes, pages_to_classify) or []
            except Exception as exc:
                if isinstance(exc, (TimeoutError, OverviewLocatorTimeout)) or "timeout" in str(exc).lower():
                    logger.info(
                        "[OVERVIEW_LOCATOR] pages=%s phase=vision end elapsed=%.3f found=timeout",
                        scan_range, time.monotonic() - vision_started,
                    )
                    _log_overview_diagnostic(
                        "locator_timeout",
                        level=logging.WARNING,
                        phase=f"{scan_range} Vision 분류",
                        elapsed_seconds=round(time.monotonic() - locator_started, 3),
                        selected_pages={
                            key: (value or {}).get("page_number")
                            for key, value in selections.items()
                        },
                        scanned_pages=list(scanned_pages),
                        vision_call_count=vision_call_count,
                    )
                    raise OverviewLocatorTimeout(
                        f"{scan_range}쪽 Vision 분류가 {OVERVIEW_VISION_TIMEOUT_SEC}초 제한을 초과했습니다."
                    ) from exc
                logger.exception(
                    "quantity_overview_locator_fallback_failed range=%s error=%s",
                    scan_range, str(exc)[:160],
                )
                returned = []
        except Exception as exc:
            if isinstance(exc, (TimeoutError, OverviewLocatorTimeout)) or "timeout" in str(exc).lower():
                logger.info(
                    "[OVERVIEW_LOCATOR] pages=%s phase=vision end elapsed=%.3f found=timeout",
                    scan_range, time.monotonic() - vision_started,
                )
                _log_overview_diagnostic(
                    "locator_timeout",
                    level=logging.WARNING,
                    phase=f"{scan_range} Vision 분류",
                    elapsed_seconds=round(time.monotonic() - locator_started, 3),
                    selected_pages={
                        key: (value or {}).get("page_number")
                        for key, value in selections.items()
                    },
                    scanned_pages=list(scanned_pages),
                    vision_call_count=vision_call_count,
                )
                raise OverviewLocatorTimeout(
                    f"{scan_range}쪽 Vision 분류가 {OVERVIEW_VISION_TIMEOUT_SEC}초 제한을 초과했습니다."
                ) from exc
            logger.exception(
                "quantity_overview_locator_fallback_failed range=%s error=%s",
                scan_range, str(exc)[:160],
            )
            returned = []
        if remaining_seconds() <= 0:
            logger.info(
                "[OVERVIEW_LOCATOR] pages=%s phase=vision end elapsed=%.3f found=timeout",
                scan_range, time.monotonic() - vision_started,
            )
            _log_overview_diagnostic(
                "locator_timeout",
                level=logging.WARNING,
                phase=f"{scan_range} Vision 완료",
                elapsed_seconds=round(time.monotonic() - locator_started, 3),
                selected_pages={
                    key: (value or {}).get("page_number")
                    for key, value in selections.items()
                },
                scanned_pages=list(scanned_pages),
                vision_call_count=vision_call_count,
            )
            raise OverviewLocatorTimeout(
                f"사업개요 페이지 자동 탐색이 {OVERVIEW_LOCATOR_TIMEOUT_SEC}초를 초과했습니다 "
                f"(단계: {scan_range} Vision 완료)."
            )
        # 실제 Gemini 분류기는 요청/응답 페이지 집합을 함께 반환한다. 부분 JSON이
        # 복구됐더라도 누락 페이지는 성공으로 확정하지 않고, 다음 범위로 넘어가기 전에
        # 최대 3페이지씩 한 번만 재분류한다. 일반 list를 반환하는 기존 주입 테스트와
        # 내부 호출은 기존 동작을 유지한다.
        missing_page_numbers = list(
            getattr(returned, "missing_page_numbers", ()) or ()
        )
        finish_reason = str(getattr(returned, "finish_reason", "") or "").upper()
        if missing_page_numbers or finish_reason == "MAX_TOKENS":
            merged_by_page = {
                int(item.get("page_number")): item
                for item in returned
                if isinstance(item, dict)
                and str(item.get("page_number") or "").strip().isdigit()
            }
            for offset in range(0, len(missing_page_numbers), 3):
                retry_pages = missing_page_numbers[offset:offset + 3]
                if not retry_pages:
                    continue
                ensure_time(f"{scan_range} 누락 페이지 Vision 재분류 시작")
                vision_call_count += 1
                vision_pages.append(list(retry_pages))
                logger.info(
                    "quantity_overview_locator_vision_call call=%s pages=%s "
                    "range=%s retry=missing_pages",
                    vision_call_count, retry_pages, scan_range,
                )
                try:
                    retry_returned = classify_batch(
                        pdf_bytes, retry_pages, vision_types,
                    ) or []
                except TypeError:
                    retry_returned = classify_batch(pdf_bytes, retry_pages) or []
                except Exception as exc:
                    if (
                        isinstance(exc, (TimeoutError, OverviewLocatorTimeout))
                        or "timeout" in str(exc).lower()
                    ):
                        raise OverviewLocatorTimeout(
                            f"{scan_range}쪽 누락 페이지 Vision 재분류가 "
                            f"{OVERVIEW_VISION_TIMEOUT_SEC}초 제한을 초과했습니다."
                        ) from exc
                    logger.exception(
                        "quantity_overview_locator_missing_retry_failed "
                        "range=%s pages=%s error=%s",
                        scan_range, retry_pages, str(exc)[:160],
                    )
                    retry_returned = []
                ensure_time(f"{scan_range} 누락 페이지 Vision 재분류 완료")
                for item in retry_returned:
                    if not isinstance(item, dict):
                        continue
                    try:
                        retry_page_number = int(item.get("page_number"))
                    except (TypeError, ValueError):
                        continue
                    if retry_page_number in retry_pages:
                        merged_by_page[retry_page_number] = item
            returned = [
                merged_by_page[page]
                for page in pages_to_classify
                if page in merged_by_page
            ]
            remaining_missing = [
                page for page in pages_to_classify
                if page not in merged_by_page
            ]
            _log_overview_diagnostic(
                "locator_missing_page_retry_complete",
                scan_range=scan_range,
                requested_page_numbers=pages_to_classify,
                retried_page_numbers=missing_page_numbers,
                remaining_missing_page_numbers=remaining_missing,
            )
        allowed = set(pages_to_classify)
        returned_page_numbers = set()
        for raw_item in returned:
            normalized_item = _normalize_page_classification(raw_item, allowed)
            raw_page_number = raw_item.get("page_number") if isinstance(raw_item, dict) else None
            if normalized_item is None:
                _log_overview_diagnostic(
                    "candidate_rejected",
                    scan_range=scan_range,
                    page_number=raw_page_number,
                    reason="invalid_classifier_result",
                )
                continue
            returned_page_numbers.add(normalized_item["page_number"])
            if normalized_item["page_type"] not in ("overview", "building_area_table"):
                _log_overview_diagnostic(
                    "candidate_rejected",
                    scan_range=scan_range,
                    page_number=normalized_item["page_number"],
                    predicted_type=normalized_item["page_type"],
                    confidence=normalized_item["confidence"],
                    title_text=normalized_item["title_text"],
                    evidence_terms=normalized_item["evidence_terms"],
                    reason="non_target_page_type",
                )
            elif not _classification_has_evidence(normalized_item):
                _log_overview_diagnostic(
                    "candidate_rejected",
                    scan_range=scan_range,
                    page_number=normalized_item["page_number"],
                    predicted_type=normalized_item["page_type"],
                    confidence=normalized_item["confidence"],
                    title_text=normalized_item["title_text"],
                    evidence_terms=normalized_item["evidence_terms"],
                    reason="evidence_validation_failed",
                )
            else:
                _log_overview_diagnostic(
                    "candidate_accepted",
                    scan_range=scan_range,
                    page_number=normalized_item["page_number"],
                    predicted_type=normalized_item["page_type"],
                    confidence=normalized_item["confidence"],
                    title_text=normalized_item["title_text"],
                    evidence_terms=normalized_item["evidence_terms"],
                )
        for page_number in pages_to_classify:
            if page_number not in returned_page_numbers:
                _log_overview_diagnostic(
                    "candidate_rejected",
                    scan_range=scan_range,
                    page_number=page_number,
                    reason="classifier_no_result",
                )
        normalized = [
            item for raw in returned
            if (item := _normalize_page_classification(raw, allowed)) is not None
            and _classification_has_evidence(item)
        ]

        # 면적표 근거는 찾았지만 overview 근거가 없으면 다음 범위로 넘어가기 전에 동일한
        # 이미지들을 overview 전용으로 딱 한 번 재분류한다. 페이지 위치는 사용하지 않는다.
        # (예전에는 first_page == 1인 배치에서만 이 재시도를 했는데, 사업개요표가 첫 배치
        # 이후 범위에 있는 문서에서는 재시도 자체가 걸리지 않았다 — 모든 배치에 동일하게
        # 적용한다.)
        found_types = {item["page_type"] for item in normalized}
        if (
            "building_area_table" in found_types
            and "overview" not in found_types
            and selections["overview"] is None
        ):
            ensure_time(f"{scan_range} overview 집중 Vision 시작")
            if progress_callback:
                progress_callback(f"개요 페이지 탐색 · {scan_range}쪽 · overview 집중 재분류")
            vision_call_count += 1
            vision_pages.append(list(pages_to_classify))
            logger.info(
                "quantity_overview_locator_vision_call call=%s pages=%s range=%s focus=overview",
                vision_call_count, pages_to_classify, scan_range,
            )
            try:
                focused_returned = classify_batch(
                    pdf_bytes, pages_to_classify, {"overview"},
                ) or []
            except TypeError:
                focused_returned = classify_batch(pdf_bytes, pages_to_classify) or []
            except Exception as exc:
                if isinstance(exc, (TimeoutError, OverviewLocatorTimeout)) or "timeout" in str(exc).lower():
                    raise OverviewLocatorTimeout(
                        f"{scan_range}쪽 overview 집중 Vision 분류가 "
                        f"{OVERVIEW_VISION_TIMEOUT_SEC}초 제한을 초과했습니다."
                    ) from exc
                logger.exception(
                    "quantity_overview_locator_focus_failed range=%s error=%s",
                    scan_range, str(exc)[:160],
                )
                focused_returned = []
            focused_normalized = [
                item for raw in focused_returned
                if (item := _normalize_page_classification(raw, allowed)) is not None
                and item["page_type"] == "overview"
                and _classification_has_evidence(item)
            ]
            normalized.extend(focused_normalized)

        by_page = {item["page_number"]: item for item in normalized}
        for page_number in pages_to_classify:
            item = by_page.get(page_number)
            predicted = item["page_type"] if item else "other"
            confidence = item["confidence"] if item else 0.0
            reason = "image_fallback" if item else "image_no_result"
            _log_locator_page(page_number, 0, predicted, confidence, reason, scan_range)
        for item in normalized:
            item["text_score"] = 0
            item["selection_reason"] = "image_fallback"
            diagnostics.append(item)

        # 같은 배치의 Vision 응답은 텍스트 후보보다 우선한다. 예를 들어 도면목록에서
        # '건축개요'라는 도면명이 검출돼도 Vision이 drawing_list로 판정하면 overview
        # 후보를 폐기하고, 실제 개요표로 판정된 페이지를 선택한다.
        selection_specs = (
            ("overview", "overview"),
            ("area_table", "building_area_table"),
        )
        for selection_key, expected_type in selection_specs:
            current = selections[selection_key]
            candidates = [
                item for item in normalized if item["page_type"] == expected_type
            ]
            if candidates:
                selections[selection_key] = max(
                    candidates, key=lambda item: item["confidence"],
                )
            elif current and current["page_number"] in allowed:
                selections[selection_key] = None
        found_labels = []
        if selections["overview"]:
            found_labels.append(f"overview:{selections['overview']['page_number']}")
        if selections["area_table"]:
            found_labels.append(f"building_area_table:{selections['area_table']['page_number']}")
        logger.info(
            "[OVERVIEW_LOCATOR] pages=%s phase=vision end elapsed=%.3f found=%s",
            scan_range, time.monotonic() - vision_started,
            ",".join(found_labels) if found_labels else "none",
        )
        _log_overview_diagnostic(
            "range_complete",
            scan_range=scan_range,
            elapsed_seconds=round(time.monotonic() - locator_started, 3),
            selected_pages={
                key: (value or {}).get("page_number")
                for key, value in selections.items()
            },
            vision_call_count=vision_call_count,
        )
        if selections["overview"] and selections["area_table"]:
            break

    top_candidates = sorted(
        (
            item for item in diagnostics
            if item.get("text_score", 0) > 0 or item.get("confidence", 0) > 0
        ),
        key=lambda item: (item.get("text_score", 0), item.get("confidence", 0)),
        reverse=True,
    )[:10]
    result = {
        "overview": selections["overview"],
        "area_table": selections["area_table"],
        "classifications": diagnostics,
        "top_candidates": top_candidates,
        "scanned_pages": scanned_pages,
        "vision_call_count": vision_call_count,
        "vision_pages": vision_pages,
        "complete": bool(selections["overview"] and selections["area_table"]),
    }
    _log_overview_diagnostic(
        "locator_complete",
        complete=result["complete"],
        elapsed_seconds=round(time.monotonic() - locator_started, 3),
        selected_pages={
            key: (value or {}).get("page_number")
            for key, value in selections.items()
        },
        scanned_pages=scanned_pages,
        vision_call_count=vision_call_count,
        vision_pages=vision_pages,
    )
    # 페이지 발견만으로는 캐시하지 않는다. 고해상도 실제 값 추출 후 프로젝트명·층수·동
    # 검증까지 통과한 경우에만 extract_overview_and_spec에서 성공 페이지를 캐시한다.
    return result


def _tile_page_image(image, columns=2, rows=2, overlap_ratio=0.04):
    """A1 전체 축소 대신 원본 렌더의 표 영역을 겹치는 타일로 잘라 반환한다."""
    width, height = image.size
    tiles = []
    for row in range(rows):
        for col in range(columns):
            x0 = int(col * width / columns)
            y0 = int(row * height / rows)
            x1 = int((col + 1) * width / columns)
            y1 = int((row + 1) * height / rows)
            pad_x, pad_y = int(width * overlap_ratio), int(height * overlap_ratio)
            box = (max(0, x0 - pad_x), max(0, y0 - pad_y),
                   min(width, x1 + pad_x), min(height, y1 + pad_y))
            tiles.append(image.crop(box))
    return tiles


def extract_overview_and_spec(structural_pdf_bytes=None, architectural_pdf_bytes=None,
                              correction_context=None, progress_callback=None,
                              architectural_page_hints=None, structural_page_hints=None):
    """표지(개요) 몇 페이지 + 구조일반사항 페이지만 뽑아서 가볍게 1번만 Gemini에 보내
    프로젝트 개요와 구조일반사항을 추출한다. 본 추출(extract_structural_members)보다
    훨씬 적은 페이지/토큰만 쓰는 사전 확인용 함수다.

    correction_context: 사용자가 이전 확인 단계에서 "아니요"를 누르고 입력한 정정 텍스트.
    주어지면 프롬프트에 최우선 근거로 반영하라고 명시해서 재요청한다(재검토).

    architectural_page_hints/structural_page_hints: 사용자가 낱장 파일(예: "A-015,016
    사업개요,동별개요.dwg")을 여러 개 선택해서 올린 경우 _merge_uploaded_pdfs가 만든
    {페이지번호: {"filename":..., "hints": {"overview", "area_table", "general_spec"}}}
    매핑. 파일명 자체에 내용이 명시돼 있으면 Vision 추측보다 이 힌트를 우선 신뢰한다
    (2026-07-27 사용자 지적: "캐드에 별도로 있잖아")."""
    client = get_gemini_client()
    if client is None:
        _log_overview_diagnostic(
            "extractor_stopped",
            level=logging.WARNING,
            stage="client_initialization",
            reason="gemini_client_unavailable",
        )
        return dict(_EMPTY_OVERVIEW_SPEC, notes=["GEMINI_API_KEY가 설정되지 않았습니다. .env 확인 후 서버를 재시작해 주세요."])

    if not structural_pdf_bytes and not architectural_pdf_bytes:
        _log_overview_diagnostic(
            "extractor_stopped",
            level=logging.WARNING,
            stage="input_validation",
            reason="no_pdf_input",
        )
        return dict(_EMPTY_OVERVIEW_SPEC, notes=["개요/구조일반사항을 확인할 PDF가 없습니다."])

    _log_overview_diagnostic(
        "extractor_start",
        has_architectural_pdf=bool(architectural_pdf_bytes),
        has_structural_pdf=bool(structural_pdf_bytes),
        architectural_hint_pages=len(architectural_page_hints or {}),
        structural_hint_pages=len(structural_page_hints or {}),
    )

    # 먼저 저해상도 이미지로 페이지 "종류"만 증분 분류한다. 값 추출은 이 단계에서 금지하고,
    # overview + 동별/층별면적표로 확정된 실제 페이지만 아래에서 고해상도 타일로 다시 렌더한다.
    overview_pairs = []  # (라벨, PIL Image)
    page_detection = {}
    try:
        cover_bytes = architectural_pdf_bytes or structural_pdf_bytes
        cover_source = "건축" if architectural_pdf_bytes else "구조"
        cover_page_hints = architectural_page_hints if architectural_pdf_bytes else structural_page_hints
        cover_info = pdfinfo_from_bytes(cover_bytes)
        cover_total = int(cover_info.get("Pages", 0) or 0)

        # 파일명 힌트로 이미 확정된 페이지가 있으면 그 자체를 최우선 신뢰한다 — 파일명이
        # "사업개요"/"동별개요"라고 명시하면 그게 정답이므로, Vision한테 다시 추측시킬
        # 필요가 없다(틀릴 여지 자체가 없어짐). 둘 다 힌트로 나오면 Vision 로케이터
        # 자체를 건너뛰어 비용/시간도 아낀다.
        hint_selected = {"overview": None, "area_table": None}
        if cover_page_hints:
            for page_num in sorted(cover_page_hints):
                info = cover_page_hints[page_num]
                hints = info.get("hints") or set()
                if "overview" in hints and hint_selected["overview"] is None:
                    hint_selected["overview"] = {
                        "page_number": page_num, "page_type": "overview",
                        "confidence": 1.0, "title_text": info.get("filename") or "",
                        "text_score": 0, "selection_reason": "filename_hint",
                    }
                if "area_table" in hints and hint_selected["area_table"] is None:
                    hint_selected["area_table"] = {
                        "page_number": page_num, "page_type": "building_area_table",
                        "confidence": 1.0, "title_text": info.get("filename") or "",
                        "text_score": 0, "selection_reason": "filename_hint",
                    }

        if hint_selected["overview"] and hint_selected["area_table"]:
            logger.info(
                "[OVERVIEW_LOCATOR] filename_hint_short_circuit overview=%s area_table=%s",
                hint_selected["overview"]["page_number"], hint_selected["area_table"]["page_number"],
            )
            selected = dict(
                hint_selected, complete=True, classifications=[], top_candidates=[],
                scanned_pages=[], vision_call_count=0, vision_pages=[],
                from_filename_hint=True,
            )
        else:
            selected = _find_incremental_overview_pages(
                cover_bytes, cover_total, max_pages=OVERVIEW_CLASSIFY_MAX_PAGES,
                progress_callback=progress_callback,
            )
            # 힌트로 확정된 쪽이 있으면 Vision 결과보다 그 힌트를 우선한다(파일명이
            # 명시적으로 알려주는 정답이 Vision의 추측보다 신뢰도가 높음).
            if hint_selected["overview"]:
                selected["overview"] = hint_selected["overview"]
            if hint_selected["area_table"]:
                selected["area_table"] = hint_selected["area_table"]
        page_detection = selected
        _log_overview_diagnostic(
            "pages_selected",
            cover_source=cover_source,
            total_pages=cover_total,
            complete=bool(selected.get("overview") and selected.get("area_table")),
            selected_pages={
                "overview": (selected.get("overview") or {}).get("page_number"),
                "area_table": (selected.get("area_table") or {}).get("page_number"),
            },
            selection_reasons={
                "overview": (selected.get("overview") or {}).get("selection_reason"),
                "area_table": (selected.get("area_table") or {}).get("selection_reason"),
            },
            scanned_pages=selected.get("scanned_pages") or [],
            vision_call_count=selected.get("vision_call_count", 0),
        )

        # 이전에는 overview/area_table 둘 다 찾았을 때만(selected["complete"]) 페이지를
        # 보냈다 — 그래서 동별면적표는 찾았는데 사업개요 페이지만 못 찾은 흔한 경우에도
        # "둘 다 못 찾음" 취급으로 이미 찾은 동별면적표 페이지까지 통째로 버리고 완전히
        # 빈 결과를 반환했다(실제 사용자 프로젝트에서 재현됨 — 101동/102동/부대시설처럼
        # 동별면적표만으로 읽을 수 있는 값까지 전부 "확인 안 됨"으로 나온 원인).
        # 이제는 둘 중 하나라도 찾았으면 그 페이지만이라도 보내고, 나머지 값은 기존처럼
        # unconfirmed_items/근거없음으로 정직하게 남긴다 — 이미 만들어둔 "핵심 항목
        # 미확인시 확인버튼 차단" 게이트가 그 나머지를 안전하게 막아준다.
        if selected.get("overview") or selected.get("area_table"):
            page_types = (
                ("overview", selected.get("overview")),
                ("area_table", selected.get("area_table")),
            )
            seen_pages = set()
            for role, info in page_types:
                if not info:
                    continue
                page_num = int(info["page_number"])
                if page_num in seen_pages or not (1 <= page_num <= cover_total):
                    continue
                seen_pages.add(page_num)
                images = _render_pdf_page_range(
                    cover_bytes, page_num, page_num, dpi=OVERVIEW_TABLE_RENDER_DPI,
                )
                if not images:
                    continue
                for tile_index, tile in enumerate(_tile_page_image(images[0]), start=1):
                    overview_pairs.append((
                        f"{cover_source} PDF 실제 {page_num}페이지 "
                        f"{info['page_type']} 표 타일 {tile_index}/4",
                        tile,
                    ))
    except OverviewLocatorTimeout:
        raise
    except Exception as exc:
        logger.exception(
            "quantity_overview_page_preparation_failed error=%s",
            str(exc)[:200],
        )
        _log_overview_diagnostic(
            "extractor_error",
            level=logging.ERROR,
            stage="page_detection_or_render",
            error=str(exc)[:200],
        )
        pass

    # 구조일반사항 페이지: 파일명 힌트(예: "S-002 구조일반사항.dwg")가 있으면 그 페이지를
    # 최우선으로 쓰고, 없으면 기존 일람표 자동감지를 재사용한다(구조 PDF에서만 찾는다 —
    # 구조일반사항은 관례상 구조도면집에 있음). max_results는 이전에 4로 하드코딩돼 있어
    # 실제 설계 의도(모듈 상수 SCHEDULE_PAGE_MAX_RESULTS=6, OVERVIEW_CHECK_MAX_PAGES 주석
    # 참고: "구조일반사항 최대 6장")보다 낮게 잡혀 있었다 — 구조일반사항이 4페이지를
    # 넘는 프로젝트에서 뒷부분 표(정착길이/피복두께 등)가 잘려나가는 원인이었다.
    spec_pairs = []
    if structural_pdf_bytes:
        try:
            spec_info = pdfinfo_from_bytes(structural_pdf_bytes)
            spec_total = int(spec_info.get("Pages", 0) or 0)
            spec_pages_from_hints = sorted(
                p for p, info in (structural_page_hints or {}).items()
                if "general_spec" in (info.get("hints") or set()) and 1 <= p <= spec_total
            )
            if spec_pages_from_hints:
                logger.info(
                    "[OVERVIEW_LOCATOR] general_spec_filename_hint pages=%s",
                    spec_pages_from_hints,
                )
                spec_pages = spec_pages_from_hints[:SCHEDULE_PAGE_MAX_RESULTS]
            else:
                spec_pages = _detect_schedule_pages(
                    structural_pdf_bytes, spec_total, max_results=SCHEDULE_PAGE_MAX_RESULTS,
                )
            for p in spec_pages:
                if len(overview_pairs) + len(spec_pairs) >= OVERVIEW_CHECK_MAX_PAGES:
                    break
                imgs = _render_pdf_page_range(
                    structural_pdf_bytes, p, p, dpi=PDF_RENDER_DPI,
                )
                if imgs:
                    spec_pairs.append((f"구조도면 {p}페이지(구조일반사항 추정)", imgs[0]))
        except Exception:
            pass

    all_pairs = overview_pairs + spec_pairs
    if not overview_pairs:
        result = _fill_overview_spec_defaults(dict(
            _EMPTY_OVERVIEW_SPEC,
            notes=["사업개요 또는 동별면적표를 자동으로 찾지 못했습니다"],
        ))
        result["page_detection"] = page_detection
        result["locator_failed"] = True
        _log_overview_diagnostic(
            "extractor_complete",
            locator_failed=True,
            selected_pages={
                "overview": (page_detection.get("overview") or {}).get("page_number"),
                "area_table": (page_detection.get("area_table") or {}).get("page_number"),
            },
            missing_fields=_overview_diagnostic_missing_fields(result.get("overview")),
            unconfirmed_items=(result.get("overview") or {}).get("unconfirmed_items") or [],
            notes=result.get("notes") or [],
        )
        return result

    contents = ["아래 도면 이미지를 보고 개요와 구조일반사항을 읽어서 추출해주세요."]
    if correction_context:
        contents.append(f"\n[사용자 확인/정정 내용]\n{correction_context}\n")
    for label, img in all_pairs:
        contents.append(f"[{label}]")
        contents.append(types.Part.from_bytes(
            data=image_to_jpeg_bytes(img, max_size=OVERVIEW_TABLE_IMAGE_MAX_SIZE),
            mime_type="image/jpeg",
        ))

    try:
        _log_overview_diagnostic(
            "value_extraction_start",
            image_count=len(all_pairs),
            image_labels=[label for label, _image in all_pairs],
        )
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=OVERVIEW_SPEC_SYSTEM_PROMPT,
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=8192,
                # Gemini 2.5 Pro의 동적 생각이 실제 재현에서 7,862토큰을 사용해
                # JSON 출력이 316토큰만 남고 MAX_TOKENS로 잘렸다. 이 요청은
                # 복잡한 추론이 아닌 표 원문 전사이므로 생각을 1,024토큰으로 제한해
                # 값과 sources 근거 JSON이 끝까지 출력될 공간을 확보한다.
                thinking_config=types.ThinkingConfig(thinking_budget=1024),
            ),
        )
    except Exception as e:
        _log_overview_diagnostic(
            "extractor_error",
            level=logging.ERROR,
            stage="value_extraction_request",
            error=str(e)[:200],
        )
        return dict(_EMPTY_OVERVIEW_SPEC, notes=[f"Gemini 개요/구조일반사항 판독 요청 중 오류가 발생했습니다: {str(e)[:200]}"])

    raw = _extract_text_from_gemini_response(response)
    _log_overview_diagnostic(
        "value_extraction_raw_response",
        raw_length=len(raw),
        raw_preview=raw[:12000],
        response_diagnostics=_gemini_response_diagnostics(response),
    )
    try:
        raw_clean = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(raw_clean)
    except json.JSONDecodeError:
        data = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        if data is None:
            _log_overview_diagnostic(
                "extractor_error",
                level=logging.ERROR,
                stage="value_extraction_json_parse",
                raw_length=len(raw),
            )
            return dict(_EMPTY_OVERVIEW_SPEC, notes=[f"JSON 파싱 오류 - 원본 응답 확인 필요: {raw[:300]}"])

    parsed_overview = data.get("overview") if isinstance(data, dict) else {}
    parsed_overview = parsed_overview if isinstance(parsed_overview, dict) else {}
    parsed_sources = parsed_overview.get("sources")
    parsed_sources = parsed_sources if isinstance(parsed_sources, dict) else {}
    _log_overview_diagnostic(
        "value_extraction_before_validation",
        extracted_fields={
            field: parsed_overview.get(field)
            for field in (
                "project_name",
                "site_location",
                "usage",
                "structure_type",
                "basement_floor_count",
                "aboveground_max_floor",
                "household_count",
            )
        },
        building_count=len(parsed_overview.get("buildings") or []),
        amenity_facility_count=len(parsed_overview.get("amenity_facilities") or []),
        source_keys=sorted(parsed_sources),
        unconfirmed_items=parsed_overview.get("unconfirmed_items") or [],
    )

    data = _fill_overview_spec_defaults(data)
    data["page_detection"] = page_detection
    data["locator_failed"] = False
    _cache_validated_overview_pages(
        cover_bytes, page_detection, data.get("overview") or {},
    )
    _log_overview_diagnostic(
        "extractor_complete",
        locator_failed=False,
        selected_pages={
            "overview": (page_detection.get("overview") or {}).get("page_number"),
            "area_table": (page_detection.get("area_table") or {}).get("page_number"),
        },
        missing_fields=_overview_diagnostic_missing_fields(data.get("overview")),
        unconfirmed_items=(data.get("overview") or {}).get("unconfirmed_items") or [],
        notes=data.get("notes") or [],
    )
    return data


_MAX_REASONABLE_BASEMENT_FLOORS = 10  # 국내 실제 프로젝트에서 지하 10개층을 넘는 사례는 극히 이례적
_MAX_REASONABLE_ABOVEGROUND_FLOORS = 100  # 초고층이라도 국내 주거동이 100층을 넘는 사례는 없음
_MAX_REASONABLE_ROOFTOP_FLOORS = 10


def _source_items(source):
    if isinstance(source, dict):
        return [source]
    if isinstance(source, list):
        return [item for item in source if isinstance(item, dict)]
    return []


def _source_quote(source):
    return " ".join(str(item.get("quote") or "") for item in _source_items(source)).strip()


def _source_is_complete(source):
    items = _source_items(source)
    return bool(items) and all(
        item.get("pdf_type") and isinstance(item.get("page"), int) and item.get("quote")
        for item in items
    )


def _source_uses_drawing_list(source):
    return any(
        "도면목록" in (str(item.get("table") or "") + str(item.get("quote") or "")).replace(" ", "")
        or "DRAWINGLIST" in str(item.get("quote") or "").replace(" ", "").upper()
        for item in _source_items(source)
    )


def _parse_explicit_floor_count(text, level):
    """도면번호가 아니라 '지하 2층/지상13층/옥탑 1층' 명시 문구만 파싱한다."""
    text = str(text or "")
    label = {"basement": "지하", "aboveground": "지상", "rooftop": "옥탑"}[level]
    matches = re.findall(rf"{label}\s*([0-9]{{1,3}})\s*층", text)
    return max((int(value) for value in matches), default=None)


def _normalize_project_title(value):
    value = re.sub(r"\s+", "", str(value or "")).lower()
    return value.replace("번지", "")


def _business_name_cell_value(source):
    """근거 quote에서 셀 라벨만 제거하고 값의 원문 표기는 그대로 보존한다."""
    for item in _source_items(source):
        quote = str(item.get("quote") or "")
        match = re.search(r"사업명(?:칭)?\s*(?:[:：]\s*)?(.+)", quote, flags=re.S)
        if match:
            value = re.sub(r"^[\s|:：\-–—]+", "", match.group(1))
            return value.strip()
    return None


def _as_number(value):
    if isinstance(value, bool) or value is None:
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _sum_numbers(items, key):
    values = [_as_number(item.get(key)) for item in items if isinstance(item, dict)]
    values = [value for value in values if value is not None]
    return (round(sum(values), 2), values) if values else (None, [])


def _format_sum(values):
    return " + ".join(f"{value:.2f}".rstrip("0").rstrip(".") for value in values)


def _source_supports_number(source, value):
    """쉼표·공백·단위·정수/소수 표기만 무시하고 quote의 독립 숫자와 정확히 비교한다."""
    expected = _as_number(value)
    if expected is None:
        return True
    quote = _source_quote(source)
    tokens = re.findall(
        r"(?<![\d.])[-+]?(?:\d{1,3}(?:\s*,\s*\d{3})+|\d+)(?:\.\d+)?(?![\d.])",
        quote,
    )
    return any(
        math.isclose(expected, parsed, rel_tol=0, abs_tol=1e-9)
        for parsed in (_as_number(token.replace(" ", "")) for token in tokens)
        if parsed is not None
    )


def _validate_overview_arithmetic(overview):
    """표 구성행 합계를 독립 계산하고 OCR 표시값과 다르면 conflict로 남긴다."""
    validations = []
    conflicts = []

    def compare(field, reported, calculated, values, label, replace_with_calculated=False):
        if calculated is None or reported is None:
            return
        matched = math.isclose(reported, calculated, rel_tol=0, abs_tol=0.01)
        entry = {
            "field": field, "reported": round(reported, 2), "calculated": calculated,
            "formula": _format_sum(values), "status": "matched" if matched else "conflict",
            "message": f"{label} 일치" if matched else f"{label} 표시값과 구성행 합계 불일치",
        }
        validations.append(entry)
        if not matched:
            conflicts.append(dict(entry))
            if replace_with_calculated:
                overview[field] = calculated

    buildings = overview.get("buildings") or []

    building_floor_sum, building_floor_values = _sum_numbers(buildings, "total_floor_area_m2")
    compare(
        "apartment_total_floor_area_m2",
        _as_number(overview.get("apartment_total_floor_area_m2")),
        building_floor_sum, building_floor_values, "공동주택 연면적 소계",
        replace_with_calculated=False,
    )

    household_sum, household_values = _sum_numbers(buildings, "household_count")
    compare(
        "household_count", _as_number(overview.get("household_count")),
        household_sum, household_values, "세대수", replace_with_calculated=False,
    )

    above = _as_number(overview.get("aboveground_floor_area_m2"))
    basement = _as_number(overview.get("basement_floor_area_m2"))
    total = _as_number(overview.get("total_floor_area_m2"))
    if above is not None and basement is not None:
        compare(
            "total_floor_area_m2", total, round(above + basement, 2),
            [above, basement], "전체 연면적", replace_with_calculated=False,
        )

    overview["validations"] = validations
    overview["conflicts"] = conflicts
    return overview


def _sanity_check_overview(data):
    """Gemini가 사업개요를 읽으면서 도면목록/시트 인덱스의 도면 번호(예: "A-101"~"A-126")를
    실제 층수로 착각해 "지하 26개층"처럼 명백히 비현실적인 값을 만들어내는 사례가 실제로
    확인됐다(2026-07-26, 사용자 프로젝트에서 재현). 프롬프트 지시만으로 완전히 막힌다는
    보장이 없어서, 코드에서 한 번 더 상식적인 범위를 벗어난 값을 걸러 null로 되돌리고
    unconfirmed_items에 사유를 남긴다. 미확인 값은 이후 도면 판독과 최종 검산에서
    참고할 수 있도록 null/미확인 상태 그대로 유지한다."""
    overview = data.get("overview")
    if not isinstance(overview, dict):
        return data
    unconfirmed = overview.get("unconfirmed_items")
    if not isinstance(unconfirmed, list):
        unconfirmed = []
        overview["unconfirmed_items"] = unconfirmed
    sources = overview.get("sources")
    if not isinstance(sources, dict):
        sources = {}
        overview["sources"] = sources

    bfc_source = sources.get("basement_floor_count")
    bfc = _parse_explicit_floor_count(_source_quote(bfc_source), "basement")
    if not _source_is_complete(bfc_source) or _source_uses_drawing_list(bfc_source) or bfc is None:
        overview["basement_floor_count"] = None
        sources.pop("basement_floor_count", None)
        unconfirmed.append("지하층 수(명시 문구와 페이지 근거가 없어 제외됨)")
    else:
        overview["basement_floor_count"] = bfc

    above_source = sources.get("aboveground_max_floor")
    above = _parse_explicit_floor_count(_source_quote(above_source), "aboveground")
    if not _source_is_complete(above_source) or _source_uses_drawing_list(above_source) or above is None:
        overview["aboveground_max_floor"] = None
        sources.pop("aboveground_max_floor", None)
        unconfirmed.append("지상 최고층(명시 문구와 페이지 근거가 없어 제외됨)")
    else:
        overview["aboveground_max_floor"] = above

    project_sources = _source_items(sources.get("project_name"))
    project_quotes = [str(item.get("quote") or "") for item in project_sources]
    has_business_name = any("사업명" in quote for quote in project_quotes)
    project_name = str(overview.get("project_name") or "").strip()
    raw_business_name = _business_name_cell_value(project_sources)
    if raw_business_name:
        # 비교용 정규화와 표시용 원문을 분리한다. 표시값에서는 '번지' 등을 절대 삭제하지 않는다.
        overview["project_name"] = raw_business_name
        project_name = raw_business_name
    normalized_name = _normalize_project_title(project_name)
    business_name_matches = any(
        "사업명" in quote and normalized_name
        and normalized_name in _normalize_project_title(quote)
        for quote in project_quotes
    )
    if not (
        _source_is_complete(project_sources) and not _source_uses_drawing_list(project_sources)
        and has_business_name and business_name_matches
    ):
        overview["project_name"] = None
        sources.pop("project_name", None)
        unconfirmed.append("프로젝트명(사업개요의 사업명칭 셀 근거가 없음)")
    project_title_quotes = [
        quote for quote in project_quotes if "PROJECT TITLE" in quote.upper()
    ]
    if project_title_quotes and normalized_name:
        overview["project_title_crosscheck"] = (
            "matched" if any(
                normalized_name in _normalize_project_title(quote)
                for quote in project_title_quotes
            ) else "mismatch"
        )
    else:
        overview["project_title_crosscheck"] = "not_available"

    for key, label in (
        ("site_location", "대지 위치"), ("usage", "용도"), ("structure_type", "구조"),
        ("household_count", "세대수"), ("site_area_m2", "대지면적"),
        ("building_area_m2", "건축면적"), ("aboveground_floor_area_m2", "지상 연면적"),
        ("basement_floor_area_m2", "지하 연면적"), ("total_floor_area_m2", "전체 연면적"),
    ):
        if overview.get(key) is not None and (
            not _source_is_complete(sources.get(key))
            or _source_uses_drawing_list(sources.get(key))
        ):
            overview[key] = None
            sources.pop(key, None)
            unconfirmed.append(f"{label}(페이지 원문 근거가 없음)")

    for key, label, empty_value in (
        ("underground_parking_note", "지하주차장", None),
        ("commercial_note", "근린생활시설", None),
    ):
        value = overview.get(key)
        if value and (
            not _source_is_complete(sources.get(key))
            or _source_uses_drawing_list(sources.get(key))
        ):
            overview[key] = empty_value
            sources.pop(key, None)
            unconfirmed.append(f"{label}(페이지 원문 근거가 없음)")

    normalized_buildings = []
    for b in overview.get("buildings") or []:
        if not isinstance(b, dict):
            continue
        label = str(b.get("label") or "").strip()
        source = b.get("source")
        if not re.fullmatch(r"\d+\s*동", label):
            continue
        label = re.sub(r"\s+", "", label)
        if (
            not _source_is_complete(source) or _source_uses_drawing_list(source)
            or label not in _source_quote(source).replace(" ", "")
        ):
            unconfirmed.append(f"{label}(동별자료 행 원문 근거가 없음)")
            continue
        row = {
            "label": label,
            "floor_range": b.get("floor_range"),
            "floor_count": _as_number(b.get("floor_count")),
            "building_area_m2": _as_number(b.get("building_area_m2")),
            "total_floor_area_m2": _as_number(b.get("total_floor_area_m2")),
            "household_count": _as_number(b.get("household_count")),
            "source": source,
        }
        for field, field_label in (
            ("floor_count", "층수"),
            ("building_area_m2", "건축면적"),
            ("total_floor_area_m2", "연면적"),
            ("household_count", "세대수"),
        ):
            if row[field] is not None and not _source_supports_number(source, row[field]):
                row[field] = None
                unconfirmed.append(
                    f"{label} {field_label}(동별자료 행 source.quote에 해당 숫자가 없어 제외됨)"
                )
        if row["floor_count"] is not None:
            row["floor_count"] = int(row["floor_count"])
        if row["household_count"] is not None:
            row["household_count"] = int(row["household_count"])
        normalized_buildings.append(row)
    overview["buildings"] = normalized_buildings
    if not normalized_buildings:
        sources.pop("buildings", None)
        unconfirmed.append("동 목록(숫자+동 형식의 독립 행 근거가 없음)")

    def normalize_facility_rows(rows, category_label):
        normalized = []
        for facility in rows or []:
            if not isinstance(facility, dict) or not str(facility.get("label") or "").strip():
                continue
            source = facility.get("source")
            if not _source_is_complete(source) or _source_uses_drawing_list(source):
                continue
            facility_label = str(facility["label"]).strip()
            row = {
                "label": facility_label,
                "source": source,
            }
            for field, field_label in (
                ("building_area_m2", "건축면적"),
                ("total_floor_area_m2", "연면적"),
                ("household_count", "세대수"),
            ):
                value = _as_number(facility.get(field))
                if value is not None and not _source_supports_number(source, value):
                    value = None
                    unconfirmed.append(
                        f"{category_label} {facility_label} {field_label}"
                        "(행 source.quote에 해당 숫자가 없어 제외됨)"
                    )
                if field == "household_count" and value is not None:
                    value = int(value)
                row[field] = value
            normalized.append(row)
        return normalized

    overview["amenity_facilities"] = normalize_facility_rows(
        overview.get("amenity_facilities"), "부대시설",
    )
    overview["utility_facilities"] = normalize_facility_rows(
        overview.get("utility_facilities"), "설비시설",
    )
    _validate_overview_arithmetic(overview)
    overview["unconfirmed_items"] = list(dict.fromkeys(unconfirmed))

    return data


def _fill_overview_spec_defaults(data):
    """extract_overview_and_spec / revise_overview_and_spec_with_text가 공통으로 쓰는
    기본값 채우기. 응답에 일부 키가 빠져도 프론트가 죽지 않도록 방어한다."""
    data = dict(data or {})
    data.setdefault("overview", {})
    data["overview"] = dict(data["overview"] or {})
    data["overview"].setdefault("project_name", None)
    data["overview"].setdefault("basement_floor_count", None)
    data["overview"].setdefault("aboveground_max_floor", None)
    data["overview"].setdefault("buildings", [])
    data["overview"].setdefault("apartment_total_floor_area_m2", None)
    data["overview"].setdefault("validations", [])
    data["overview"].setdefault("conflicts", [])
    for key in (
        "site_location", "usage", "structure_type", "household_count", "site_area_m2",
        "building_area_m2", "aboveground_floor_area_m2", "basement_floor_area_m2",
        "total_floor_area_m2",
    ):
        data["overview"].setdefault(key, None)
    data["overview"].setdefault("underground_parking_note", None)
    data["overview"].setdefault("amenity_facilities", [])
    data["overview"].setdefault("utility_facilities", [])
    data["overview"].setdefault("commercial_note", None)
    data["overview"].setdefault("unconfirmed_items", [])
    data["overview"].setdefault("sources", {})
    if not isinstance(data["overview"]["sources"], dict):
        data["overview"]["sources"] = {}
    data.setdefault("general_spec", {})
    data["general_spec"] = dict(data["general_spec"] or {})
    data["general_spec"].setdefault("concrete_fck_table", [])
    data["general_spec"].setdefault("rebar_grade_table", [])
    data["general_spec"].setdefault("lap_splice_table", [])
    data["general_spec"].setdefault("anchorage_table", [])
    data["general_spec"].setdefault("cover_table", [])
    data["general_spec"].setdefault("seismic_rebar_rules", [])
    data["general_spec"].setdefault("summary_notes", [])
    data["general_spec"].setdefault("unconfirmed_items", [])
    data.setdefault("notes", [])
    data = _sanity_check_overview(data)
    return data


def revise_overview_and_spec_with_text(prior_result, correction_text):
    """이전 개요/구조일반사항 추출 결과(prior_result)에 사용자의 정정 텍스트만 반영해서
    수정된 JSON을 만든다. 도면 이미지를 다시 보내지 않고 텍스트만 주고받으므로
    extract_overview_and_spec()(이미지 재전송)보다 훨씬 저렴하다 — 사용자가 "아니요"를
    누르고 정정 내용을 입력했을 때, 전체 도면을 다시 읽는 대신 이 함수로 가볍게 재검토한다."""
    client = get_gemini_client()
    if client is None:
        return dict(prior_result or _EMPTY_OVERVIEW_SPEC, notes=["GEMINI_API_KEY가 설정되지 않았습니다. .env 확인 후 서버를 재시작해 주세요."])
    if not correction_text:
        return _fill_overview_spec_defaults(prior_result or _EMPTY_OVERVIEW_SPEC)

    prompt = (
        "아래는 이전에 도면에서 읽은 프로젝트 개요/구조일반사항 JSON입니다. 사용자가 이 중 일부가 "
        "틀렸다고 정정 내용을 알려줬습니다. 정정 내용을 최우선 근거로 반영해서 전체 JSON을 다시 "
        "만들어 반환하세요 — 정정과 무관한 값은 그대로 유지하세요. 정정 내용에 없는 값을 새로 "
        "추측해서 바꾸지 마세요. 정정으로 새롭게 확인된 항목은 unconfirmed_items에서 제거하고, "
        "그 항목의 overview.sources 값을 \"사용자 확인\"으로 갱신하세요(도면 근거가 아니라 "
        "사용자가 직접 알려준 값임을 구분할 수 있어야 합니다).\n\n"
        f"[이전 추출 결과]\n{json.dumps(prior_result or {}, ensure_ascii=False)[:8000]}\n\n"
        f"[사용자 정정 내용]\n{correction_text}\n"
    )

    try:
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=[prompt],
            config=types.GenerateContentConfig(
                system_instruction=OVERVIEW_SPEC_SYSTEM_PROMPT,
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=8192,
            ),
        )
    except Exception as e:
        return dict(prior_result or _EMPTY_OVERVIEW_SPEC, notes=[f"Gemini 재검토 요청 중 오류가 발생했습니다: {str(e)[:200]}"])

    raw = _extract_text_from_gemini_response(response)
    try:
        raw_clean = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(raw_clean)
    except json.JSONDecodeError:
        data = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        if data is None:
            return dict(prior_result or _EMPTY_OVERVIEW_SPEC, notes=[f"JSON 파싱 오류 - 원본 응답 확인 필요: {raw[:300]}"])

    return _fill_overview_spec_defaults(data)


# 대형 프로젝트(수십~백 페이지)를 한 번의 Gemini 호출에 몰아넣으면 부재가 많을수록
# 응답 JSON이 커져서 max_output_tokens(65536)를 넘어 잘리거나, 요청 자체가 타임아웃/오류로
# 실패해서 "전체 결과 0건"이 되는 문제가 실제로 발생했다. 그래서 페이지를 이 크기 단위로
# 나눠 순차 호출하고 결과를 합친다 — 배치 하나가 실패해도 나머지 배치의 데이터는 살아남는다.
# 15페이지였을 때 실제 대형 아파트 프로젝트(경성빌라 참고 사례)에서 배치 6개 중 4개가
# max_output_tokens에 걸려 응답이 중간에 잘리는 게 확인돼서 8로 낮췄다 — 배치당 예상
# 출력 토큰을 줄여 잘림 위험을 낮추는 대신, Gemini 호출 횟수(=비용)는 그만큼 늘어난다.
EXTRACTION_BATCH_PAGE_SIZE = 8


# (mark, zone, section)이 같은 항목끼리 중복 여부를 판단할 때 비교할 "핵심 치수" 필드들.
# 이 필드들이 서로 크게 다르면 배치 중복 추출이 아니라 같은 mark를 쓰는 진짜 다른
# 부재(예: 같은 W1이라도 위치마다 실제 길이가 다른 경우)일 가능성이 커서 자동으로
# 지우지 않는다 — ChatGPT 2차 검토에서 지적받은 "같은 mark/zone/section이 항상 같은
# 부재라고 단정할 수 없다"는 위험에 대한 대응.
_DEDUP_DIMENSION_FIELDS = {
    "foundations": ["length_m", "width_m", "thickness_m"],
    "columns": ["width_m", "depth_m", "height_m"],
    "beams": ["width_m", "depth_m", "length_m"],
    "slabs": ["area_m2", "thickness_m"],
    "walls": ["length_m", "height_m", "thickness_m"],
    "stairs": ["width_m", "length_m", "thickness_m"],
}


def _dims_roughly_match(a, b, fields, rel_tol=0.03, abs_tol=0.02):
    """두 항목의 핵심 치수 필드가 (둘 다 값이 있는 경우에 한해) 오차범위 내로 같은지 본다.
    한쪽만 값이 있으면(다른 배치가 그 필드를 못 읽었을 뿐일 수 있어서) 그 필드는 판단에서
    제외한다 — 값이 있는 필드끼리만 비교해서, 둘 다 있는데 실제로 다르면 다른 부재로 본다."""
    for f in fields:
        av, bv = a.get(f), b.get(f)
        if av is None or bv is None:
            continue
        if not isinstance(av, (int, float)) or not isinstance(bv, (int, float)):
            continue
        if abs(av - bv) > max(abs_tol, rel_tol * max(abs(av), abs(bv))):
            return False
    return True


def _dedupe_category_items(items, cat_key=None):
    """같은 부재가 두 번 이상 추출된 항목을 걸러낸다.

    일람표/구조일반사항 페이지를 모든 배치에 함께 끼워 보내다 보니(마크의 실제 치수를
    다른 배치에서도 찾을 수 있게 하려고), 그 일람표 자체에 나열된 마크가 "평면도에서
    본 부재"와 별개로 한 번 더 추출되거나, 같은 일람표 페이지가 여러 배치에 반복
    전송되면서 같은 부재가 배치마다 중복 추출될 수 있다 — 실제로 이런 위험이 있다는
    지적을 받고 코드 차원의 안전망으로 추가했다.

    mark가 있는 항목만 (mark, zone, section) 조합으로 중복을 판단한다. mark가 없는
    ("무명") 항목은 서로 구분할 근거가 없어 잘못 합쳤다가 실제로 다른 부재를 지울
    위험이 더 크므로 그대로 둔다.

    (mark, zone, section)이 같아도 핵심 치수가 서로 크게 다르면(_dims_roughly_match가
    False) 자동으로 지우지 않는다 — 같은 mark를 쓰는 서로 다른 실제 부재를 배치 중복으로
    오인해서 물량을 잃어버릴 위험이, 진짜 배치 중복을 한 번 더 보여주는 것보다 크다고
    판단했다. 이 경우 둘 다 남기고 _dedup_flag를 붙여서 검토 화면에서 사람이 확인하게 한다.

    치수가 같은 범위 내(진짜 중복으로 판단됨)면, 필드가 더 많이 채워진(정보가 더 풍부한)
    쪽을 남긴다 — 같은 부재를 서로 다른 배치가 서로 다르게(한쪽은 자세히, 한쪽은
    부분적으로) 읽었을 수 있어서다."""
    dim_fields = _DEDUP_DIMENSION_FIELDS.get(cat_key, [])
    seen_at = {}
    result = []
    removed = 0
    flagged = 0
    for it in items:
        if not isinstance(it, dict) or not it.get("mark"):
            result.append(it)
            continue
        key = (it.get("mark"), it.get("zone") or "", it.get("section") or "")
        if key not in seen_at:
            seen_at[key] = len(result)
            result.append(it)
            continue
        idx = seen_at[key]
        existing = result[idx]
        if dim_fields and not _dims_roughly_match(existing, it, dim_fields):
            it["_dedup_flag"] = "같은 mark/구역인데 치수가 달라 자동 병합하지 않았습니다 — 실제로 다른 부재인지 확인해 주세요."
            existing["_dedup_flag"] = existing.get("_dedup_flag") or it["_dedup_flag"]
            result.append(it)
            flagged += 1
            continue
        removed += 1
        existing_filled = sum(1 for v in existing.values() if v not in (None, "", []))
        new_filled = sum(1 for v in it.values() if v not in (None, "", []))
        if new_filled > existing_filled:
            result[idx] = it
    return result, removed, flagged


def _merge_extracted_members(batch_results: list) -> dict:
    """extract_structural_members()가 배치별로 얻은 결과 리스트를 하나로 합친다.
    - foundations/columns/beams/slabs/walls/stairs: 배열을 이어붙인 뒤 중복 추출을 걸러낸다
      (_dedupe_category_items — 일람표 페이지 반복 전송으로 같은 부재가 두 번 뽑힐 수 있어서).
    - notes: 배치별 메모를 순서대로 이어붙인다(각 메모에 이미 배치 태그가 붙어 있음).
    - general_spec: 구조일반사항은 보통 도면 앞쪽 한 곳에만 있으므로, 스칼라 필드는
      처음 나온 non-null 값을, 표(테이블) 필드는 처음 나온 비어있지 않은 표를 채택한다
      (여러 배치에 같은 정보가 중복 추출돼도 뒤에 나온 값으로 덮어쓰지 않는다)."""
    merged = {
        "foundations": [], "columns": [], "beams": [], "slabs": [], "walls": [], "stairs": [],
        "notes": [], "general_spec": {},
    }
    scalar_fields = [
        "concrete_fck_mpa", "rebar_grade", "lap_splice_class", "cover_thickness_mm",
        "chair_bar_size", "chair_bar_height_m",
    ]
    table_fields = ["concrete_fck_table", "rebar_grade_table", "lap_splice_table", "anchorage_table"]

    for data in batch_results:
        for key in ("foundations", "columns", "beams", "slabs", "walls", "stairs"):
            merged[key].extend(data.get(key) or [])
        merged["notes"].extend(data.get("notes") or [])

        gs = data.get("general_spec") or {}
        for field in scalar_fields:
            if merged["general_spec"].get(field) in (None, "") and gs.get(field) not in (None, ""):
                merged["general_spec"][field] = gs[field]
        for field in table_fields:
            if not merged["general_spec"].get(field) and gs.get(field):
                merged["general_spec"][field] = gs[field]

    dedup_total = 0
    for key, label in _CATEGORY_KEY_TO_LABEL.items():
        merged[key], removed, flagged = _dedupe_category_items(merged[key], key)
        if removed:
            dedup_total += removed
            merged["notes"].append(f"{label}: 배치 간 중복 추출된 항목 {removed}건을 병합했습니다(같은 mark/zone/section 기준).")
        if flagged:
            merged["notes"].append(
                f"{label}: 같은 mark/zone/section인데 치수가 달라 자동 병합하지 않고 둘 다 남긴 항목이 "
                f"{flagged}건 있습니다 — 검토 화면에서 실제로 다른 부재인지 확인해 주세요."
            )

    return merged


def _apply_confirmed_general_spec_override(result, confirmed_general_spec):
    """사용자가 개요/구조일반사항 사전 확인 단계에서 이미 확인한 general_spec을, 본 추출
    결과의 general_spec 위에 강제로 덮어쓴다. 배치 프롬프트에 이미 확인된 값을 우선하라고
    안내는 해두었지만(_extract_structural_members_one_batch의 confirmed_note), Gemini가
    그래도 다르게 읽어올 가능성을 코드 차원에서 한 번 더 막기 위한 마지막 안전장치다 —
    사용자가 직접 확인한 값이 AI가 다시 읽은 값보다 항상 우선해야 한다."""
    if not confirmed_general_spec:
        return result
    result = dict(result)
    gs = dict(result.get("general_spec") or {})
    overridden_fields = []
    # AI판독값 vs 사용자확정값 병행 보관 — 사용자가 "나중에 산출 근거도 추적할 수 있게"
    # 요청한 감사(audit) 기록. 덮어쓰기 전에 이 본 추출이 자체적으로 읽은 값(ai_value)을
    # 먼저 남겨두고, 그 다음에 사용자 확정값(confirmed_value)으로 덮어쓴다 — 나중에 결과를
    # 다시 열어봤을 때 "AI는 뭐라고 읽었고 사람은 뭐라고 확정했는지"를 함께 볼 수 있다.
    audit = []
    for k, v in confirmed_general_spec.items():
        if v in (None, "", []):
            continue
        ai_value = gs.get(k)
        gs[k] = v
        overridden_fields.append(k)
        audit.append({"field": k, "ai_value": ai_value, "confirmed_value": v})
    # 사용자가 개요/구조일반사항 사전 확인 단계를 실제로 거쳤다는 표시.
    # quantity_calc.py의 get_splice_length/get_anchorage_length가 이 플래그를 보고,
    # 표에서 못 찾은 이음/정착길이 항목을 공식으로 추정하지 않고 완전히 제외한다
    # (사용자가 명시적으로 요청한 정책: "확정 전까지 계산제외").
    gs["_confirmed"] = True
    result["general_spec"] = gs
    if audit:
        result["general_spec_confirmation_audit"] = audit
    if overridden_fields:
        result["notes"] = list(result.get("notes") or []) + [
            f"사용자가 사전 확인 단계에서 확인한 구조일반사항({', '.join(overridden_fields)})을 "
            "그대로 반영했습니다."
        ]
    return result


def _extract_structural_members_one_batch(client, dwg_data, image_batch, batch_idx, total_batches, page_numbers=None, confirmed_general_spec=None):
    """extract_structural_members()의 배치 1개를 처리한다. 실패해도 예외를 던지지 않고
    _EMPTY_MEMBERS(+오류 메모)를 반환해서, 이 배치가 실패해도 다른 배치는 계속 처리되게 한다.

    page_numbers: image_batch와 같은 길이의 리스트로, 각 이미지가 실제 PDF의 몇 페이지인지
    알려준다. 반드시 넘겨야 한다 — 안 넘기면(None) 이미지 순번(1,2,3...)을 페이지 번호처럼
    라벨링하게 되는데, 이 배치가 예를 들어 16~30페이지 범위라면 실제로는 16페이지인 이미지가
    "1페이지"로 잘못 표시되는 문제가 있었다(과거 버그). 부재 위치(bbox.page)를 실제 페이지
    번호와 연결해서 도면에 색칠 미리보기를 그리려면 이 번호가 정확해야 한다.

    confirmed_general_spec: 사용자가 개요/구조일반사항 사전 확인 단계에서 이미 확인/정정한
    general_spec(dict). 주어지면 이 배치가 구조일반사항 페이지를 못 보거나 다르게 읽더라도
    이미 확인된 값을 최우선으로 쓰라고 명시해서, 사전 확인 단계에서 잡은 실수가 본 추출에서
    다시 반복되지 않게 한다."""
    is_multi_batch = total_batches > 1
    batch_tag = f"[배치 {batch_idx}/{total_batches}] " if is_multi_batch else ""

    def _tag_notes(note_list):
        return [f"{batch_tag}{n}" for n in note_list] if is_multi_batch else list(note_list)

    if page_numbers is None:
        # 하위 호환: 실제 페이지 번호를 모르면 순번으로라도 라벨링한다(이전 동작과 동일).
        page_numbers = list(range(1, len(image_batch) + 1))

    confirmed_note = ""
    if confirmed_general_spec:
        confirmed_note = (
            "\n\n[사용자가 이미 확인한 구조일반사항 — 이 값을 general_spec에 그대로 반영하세요. "
            "도면에서 다르게 보이더라도 아래 값을 우선하세요]\n"
            f"{json.dumps(confirmed_general_spec, ensure_ascii=False)[:4000]}\n"
        )

    contents = [
        f"[구조도면 DWG 파싱 데이터]\n"
        f"{json.dumps(dwg_data, ensure_ascii=False, indent=2)[:60000]}\n\n"
        "위 데이터와 아래 도면 이미지를 보고, 부재일람표/배근도에 적힌 치수와 철근 정보를 "
        "그대로 읽어서 스키마에 맞게 추출해주세요. 계산하지 말고 읽기만 하세요."
        + (
            f"\n\n※ 이 요청은 전체 도면 중 {batch_idx}/{total_batches}번째 배치입니다. "
            "이 배치에 포함된 이미지에서 실제로 보이는 부재만 추출하세요 — 다른 배치에 있을 "
            "부재를 추측해서 채우거나 이 배치에 없는 부재를 생략하지 마세요."
            if is_multi_batch else ""
        )
        + confirmed_note
    ]
    for i, img in enumerate(image_batch):
        page_label = page_numbers[i] if i < len(page_numbers) else (i + 1)
        contents.append(f"[도면 {page_label}페이지]")
        contents.append(types.Part.from_bytes(data=image_to_jpeg_bytes(img), mime_type="image/jpeg"))

    try:
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=MEMBER_EXTRACTION_SYSTEM_PROMPT,
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=65536,
            ),
        )
    except Exception as e:
        return dict(_EMPTY_MEMBERS, notes=_tag_notes([f"Gemini 부재 추출 요청 중 오류가 발생했습니다: {str(e)[:200]}"]))

    raw = _extract_text_from_gemini_response(response)

    try:
        raw_clean = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(raw_clean)
    except json.JSONDecodeError:
        data = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        if data is None:
            return dict(_EMPTY_MEMBERS, notes=_tag_notes([f"JSON 파싱 오류 - 원본 응답 확인 필요: {raw[:300]}"]))
        data.setdefault("notes", [])
        data["notes"].append(
            "Gemini 응답이 도중에 잘려(max_output_tokens 한도) 마지막 일부 부재는 누락됐을 수 있습니다 "
            "— 부분 복구된 데이터입니다. 개수가 적어 보이면 도면을 더 잘게 나눠서 재실행해 주세요."
        )

    # 응답에 일부 키가 빠져도 계산 엔진이 죽지 않도록 방어
    for key in ("foundations", "columns", "beams", "slabs", "walls", "stairs", "notes"):
        data.setdefault(key, [])
    data.setdefault("general_spec", {})

    data["notes"] = _tag_notes(data.get("notes") or [])
    return data


def extract_structural_members(dwg_data: dict, pdf_bytes, progress_cb=None, cancel_cb=None, confirmed_general_spec=None) -> dict:
    """
    DWG 파싱 데이터 + 구조도면 PDF(원본 바이트) 를 Gemini Vision에 보내
    '부재 리스트만' 구조화된 JSON으로 추출한다 (계산은 하지 않음).
    실제 물량 계산은 quantity_calc.compute_structural_quantities가 결정론적으로 수행한다.

    confirmed_general_spec: 개요/구조일반사항 사전 확인 단계에서 사용자가 이미 확인/정정한
    general_spec. 주어지면 모든 배치 요청에 함께 실어보내서, 본 추출이 사전 확인 단계에서
    잡은 실수(Fck/철근강종/이음등급 등)를 다시 반복하지 않게 한다.

    두 가지 방식으로 대형 프로젝트(수십~백 페이지)에서의 실패를 막는다:
    1) Gemini 호출을 EXTRACTION_BATCH_PAGE_SIZE 페이지 단위로 나눠 순차 호출 후 병합
       — 한 번에 몰아넣으면 응답이 max_output_tokens를 넘어 잘리거나 요청이 타임아웃/
       오류로 실패해서 전체 결과가 0건이 되는 문제를 막는다.
    2) PDF 페이지도 한꺼번에 다 렌더링하지 않고 배치 단위로 그때그때만 렌더링한다
       — 79페이지 같은 대형 도면집을 한 번에 전부 이미지로 뽑으면(pdf2image가 전체를
       메모리에 올림) 수 GB 메모리가 필요해져 서버가 죽거나(MemoryError) 매우 느려질
       수 있기 때문이다.
    배치 1개(렌더링이든 Gemini 호출이든)가 실패해도 다른 배치의 데이터는 그대로 살아남는다.

    progress_cb(idx, total_batches)가 주어지면, 각 배치를 처리하기 직전에 호출해서
    "지금 몇 배치째"인지 외부(예: 진행률 폴링 저장소)에 알릴 수 있게 한다. 필수 아님 —
    None이면 그냥 진행률 보고를 생략한다(기존 호출부와 호환).

    cancel_cb()가 주어지고 True를 반환하면, 다음 배치를 아예 Gemini에 보내지 않고
    루프를 즉시 멈춘다(이미 시작된 배치 1개는 어차피 끝까지 처리— Gemini 호출 자체를
    중간에 끊을 방법은 없다). 사용자가 "취소"를 누른 뒤 아직 시작 안 한 배치들의
    비용을 아끼기 위한 용도다. None이면 취소 체크를 생략한다.
    """
    cache_key = _dev_cache_key(pdf_bytes, dwg_data) if _DEV_CACHE_ENABLED else None
    if cache_key:
        cached = _dev_cache_load(cache_key)
        if cached is not None:
            if progress_cb:
                progress_cb(1, 1)
            return _apply_confirmed_general_spec_override(cached, confirmed_general_spec)

    client = get_gemini_client()
    if client is None:
        return dict(_EMPTY_MEMBERS, notes=["GEMINI_API_KEY가 설정되지 않았습니다. .env 확인 후 서버를 재시작해 주세요."])

    if not pdf_bytes:
        # PDF가 없고 DWG 데이터만 있는 경우(구조 ZIP만 업로드) — 배치 없이 한 번만 호출
        if progress_cb:
            progress_cb(1, 1)
        result = _merge_extracted_members(
            [_extract_structural_members_one_batch(client, dwg_data, [], 1, 1, confirmed_general_spec=confirmed_general_spec)]
        )
        if cache_key:
            _dev_cache_save(cache_key, result)
        return _apply_confirmed_general_spec_override(result, confirmed_general_spec)

    try:
        info = pdfinfo_from_bytes(pdf_bytes)
        actual_pages = int(info.get("Pages", 0) or 0)
        total_pages = min(actual_pages, MAX_PDF_PAGES_TO_GEMINI)
    except Exception as e:
        return dict(_EMPTY_MEMBERS, notes=[f"구조도면 PDF 페이지 수 확인 중 오류가 발생했습니다: {str(e)[:200]}"])

    if total_pages <= 0:
        return dict(_EMPTY_MEMBERS, notes=["구조도면 PDF에서 읽을 수 있는 페이지가 없습니다 — 파일이 비어있거나 손상됐을 수 있습니다."])

    # 페이지가 MAX_PDF_PAGES_TO_GEMINI(80장)를 넘으면 그 이후 페이지는 아예 처리하지 않고
    # 조용히 잘려나갔다 — 대형 프로젝트(80장 초과 도면집)에서 결과가 "완전"해 보이지만
    # 실제로는 일부 부재가 통째로 누락될 수 있는 위험한 상황이라, 반드시 눈에 띄는 note를
    # 남긴다. 이 문구는 _INCOMPLETE_EXTRACTION_MARKERS에도 등록해서, 검토 팝업의 "불완전
    # 추출" 경고 배너(진행 전 확인 강제)가 이 경우도 자동으로 잡아내게 한다.
    page_cap_truncated = actual_pages > MAX_PDF_PAGES_TO_GEMINI
    page_cap_note = None
    if page_cap_truncated:
        page_cap_note = (
            f"구조도면 PDF가 총 {actual_pages}페이지인데, 한 번에 처리 가능한 최대 페이지 수"
            f"({MAX_PDF_PAGES_TO_GEMINI}장)를 넘어서 {MAX_PDF_PAGES_TO_GEMINI + 1}페이지부터 "
            f"{actual_pages}페이지까지({actual_pages - MAX_PDF_PAGES_TO_GEMINI}페이지)는 도중에 잘려 "
            "처리하지 못했습니다 — 해당 페이지의 부재는 결과에서 완전히 누락됩니다. 도면을 나눠서 "
            "여러 번 업로드하거나, 필요한 페이지만 추려서 다시 시도해 주세요."
        )

    page_ranges = [
        (start, min(start + EXTRACTION_BATCH_PAGE_SIZE - 1, total_pages))
        for start in range(1, total_pages + 1, EXTRACTION_BATCH_PAGE_SIZE)
    ]
    total_batches = len(page_ranges)

    # 일람표/구조일반사항으로 보이는 페이지를 미리 찾아서, 각 배치 이미지에 함께 끼워
    # 넣는다 — 여러 배치로 나눠 처리할 때만 의미가 있다(배치가 1개면 어차피 전체를 다
    # 보내므로 불필요). 페이지 텍스트가 안 뽑히는 PDF면 schedule_pages가 그냥 빈 리스트로
    # 나오고, 이 기능만 조용히 꺼진다(에러 아님).
    schedule_pages = []
    schedule_page_images = {}
    if total_batches > 1:
        try:
            schedule_pages = _detect_schedule_pages(pdf_bytes, total_pages)
        except Exception:
            schedule_pages = []
        for p in schedule_pages:
            try:
                imgs = _render_pdf_page_range(pdf_bytes, p, p)
                if imgs:
                    schedule_page_images[p] = imgs[0]
            except Exception:
                continue
        # 렌더링까지 실제로 성공한 페이지만 남긴다(감지는 됐는데 렌더링이 실패한 페이지 제외)
        schedule_pages = [p for p in schedule_pages if p in schedule_page_images]

    batch_results = []
    schedule_note_added = False
    for idx, (first_page, last_page) in enumerate(page_ranges, 1):
        if cancel_cb and cancel_cb():
            batch_results.append(dict(
                _EMPTY_MEMBERS,
                notes=[f"사용자가 취소를 요청하여 배치 {idx}/{total_batches}부터 남은 페이지({first_page}~{total_pages}페이지)는 처리하지 않았습니다."],
            ))
            break
        if progress_cb:
            progress_cb(idx, total_batches)
        try:
            image_batch = _render_pdf_page_range(pdf_bytes, first_page, last_page)
        except Exception as e:
            tag = f"[배치 {idx}/{total_batches}] " if total_batches > 1 else ""
            batch_results.append(dict(
                _EMPTY_MEMBERS,
                notes=[f"{tag}PDF {first_page}~{last_page}페이지 렌더링 중 오류가 발생했습니다: {str(e)[:200]}"],
            ))
            continue

        # 이 배치 자체 범위와 겹치지 않는 일람표 페이지만 앞에 붙인다(겹치면 이미 image_batch
        # 안에 들어있으니 중복 전송할 필요 없음). 이렇게 하면 평면도만 보이는 배치도 같은
        # 요청 안에서 부재 마크의 실제 치수(일람표)를 함께 보고 매칭할 수 있다.
        extra_pairs = [
            (p, img) for p, img in schedule_page_images.items()
            if not (first_page <= p <= last_page)
        ]
        extra_images = [img for _, img in extra_pairs]
        combined_batch = extra_images + image_batch
        # image_batch의 각 이미지는 first_page..last_page 순서 그대로이므로, 실제 페이지
        # 번호를 그 순서에 맞춰 만들어 붙인다 — 이래야 Gemini에게 보여주는 "[도면 N페이지]"
        # 라벨과 부재 위치(bbox.page)가 실제 PDF 페이지 번호와 정확히 일치한다.
        combined_page_numbers = [p for p, _ in extra_pairs] + list(range(first_page, last_page + 1))

        one_result = _extract_structural_members_one_batch(
            client, dwg_data, combined_batch, idx, total_batches, page_numbers=combined_page_numbers,
            confirmed_general_spec=confirmed_general_spec,
        )
        if extra_images and not schedule_note_added:
            schedule_note_added = True
            one_result = dict(one_result)
            one_result["notes"] = list(one_result.get("notes") or []) + [
                f"일람표/구조일반사항으로 추정되는 페이지({', '.join(str(p) for p in schedule_pages)})를 "
                "모든 배치 이미지에 함께 포함해서, 평면도만 보이는 배치에서도 부재 마크의 치수를 "
                "찾을 수 있도록 했습니다."
            ]
        batch_results.append(one_result)
        del image_batch  # 다음 배치로 넘어가기 전에 명시적으로 메모리 해제

    result = _merge_extracted_members(batch_results)
    if page_cap_note:
        result["notes"] = [page_cap_note] + list(result.get("notes") or [])
    if cache_key:
        _dev_cache_save(cache_key, result)
    return _apply_confirmed_general_spec_override(result, confirmed_general_spec)


# ─────────────────────────────────────────────
#  색칠된 도면 미리보기: Gemini가 채운 bbox(부재 위치)를 원본 도면 페이지 위에
#  부재종류별 색상 박스로 그려서, 의뢰인이 채팅창에서 "빠진 벽/보가 있는지" 눈으로
#  직접 확인할 수 있게 한다. bbox가 없는 항목/페이지는 조용히 건너뛴다(이 기능이
#  실패하거나 데이터가 없어도 나머지 수량산출 결과에는 영향이 없어야 한다).
# ─────────────────────────────────────────────
_CATEGORY_KEY_TO_LABEL = {
    "foundations": "기초", "columns": "기둥", "beams": "보",
    "slabs": "슬래브", "walls": "전단벽", "stairs": "계단",
}
# 이미지 위에 직접 그려 넣는 라벨은 일부러 영문 약어를 쓴다 — PIL의 기본 폰트는 한글
# 글리프가 없어서(운영 서버에 한글 지원 폰트가 없을 수도 있음), 한글로 그리면 네모(tofu)로
# 깨져 보인다. 한글 부재종류 이름은 categories 필드(HTML/채팅창 텍스트)로만 노출하고,
# 픽셀에 직접 굽는 라벨은 폰트 의존성 없는 영문 약어로 통일한다.
_CATEGORY_SHORT_LABEL = {
    "기초": "FTG", "기둥": "COL", "보": "BEAM",
    "슬래브": "SLAB", "전단벽": "WALL", "계단": "STAIR",
}
_CATEGORY_BBOX_COLOR = {
    "기초": (230, 126, 34),   # 주황
    "기둥": (155, 89, 182),   # 보라
    "보": (231, 76, 60),      # 빨강
    "슬래브": (52, 152, 219), # 파랑 (바닥)
    "전단벽": (241, 196, 15), # 노랑 (벽체)
    "계단": (46, 204, 113),   # 초록
}
ANNOTATED_PAGE_MAX_PAGES = 12  # 채팅창에 한 번에 띄울 색칠 페이지 수 상한(응답 크기 제한용)


def _draw_member_boxes_on_image(img: Image.Image, boxes):
    """boxes: [(category_label_kr, mark, box_2d[ymin,xmin,ymax,xmax](0~1000 정규화), color_rgb), ...]
    반투명 채우기 + 테두리 + 라벨 텍스트를 원본 이미지 위에 겹쳐 그린 새 이미지를 반환한다
    (원본은 건드리지 않음). RGBA 오버레이를 따로 그려서 알파 합성해야 실제로 반투명하게
    보인다 — RGB 이미지에 바로 그리면 PIL이 알파를 무시하고 불투명하게 그려버린다.
    라벨 텍스트는 category_label_kr을 영문 약어(_CATEGORY_SHORT_LABEL)로 바꿔서 그린다
    (한글 폰트 미설치 서버에서도 깨지지 않도록)."""
    base = img.convert("RGBA")
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    w, h = base.size
    for category_label, mark, box_2d, color in boxes:
        try:
            ymin, xmin, ymax, xmax = (float(v) for v in box_2d)
        except (TypeError, ValueError):
            continue
        x0, y0 = xmin / 1000.0 * w, ymin / 1000.0 * h
        x1, y1 = xmax / 1000.0 * w, ymax / 1000.0 * h
        if x1 <= x0 or y1 <= y0:
            continue
        draw.rectangle([x0, y0, x1, y1], outline=color + (255,), width=3, fill=color + (55,))
        short_label = _CATEGORY_SHORT_LABEL.get(category_label, category_label)
        label = f"{short_label} {mark}".strip() if mark else short_label
        text_y = max(0, y0 - 14)
        draw.rectangle([x0, text_y, x0 + 7 * len(label) + 4, text_y + 13], fill=color + (200,))
        # 라벨 배경색 밝기에 따라 글자색을 흰/검 중에서 고른다 — 노랑처럼 밝은 배경에
        # 흰 글자를 고정으로 쓰면 안 보여서(전단벽 색을 노랑으로 바꾸면서 실제로 발생), YIQ
        # 공식으로 인지 밝기를 계산해 128 기준으로 나눈다.
        brightness = color[0] * 0.299 + color[1] * 0.587 + color[2] * 0.114
        text_color = (0, 0, 0, 255) if brightness > 150 else (255, 255, 255, 255)
        draw.text((x0 + 2, text_y), label, fill=text_color)
    return Image.alpha_composite(base, overlay).convert("RGB")


def build_annotated_pages(pdf_bytes, members: dict, max_pages=ANNOTATED_PAGE_MAX_PAGES):
    """추출된 members(foundations/columns/beams/slabs/walls/stairs)의 bbox 필드를 모아
    페이지별로 묶고, 그 페이지의 원본 이미지를 다시 렌더링해서 부재종류별 색상 박스를
    그린 뒤 base64 JPEG로 인코딩한다.
    Returns: [{"page": int, "image_data_url": str, "categories": [...], "member_count": int}, ...]
    (페이지 번호 오름차순). pdf_bytes가 없거나 bbox 데이터가 하나도 없으면 빈 리스트를
    반환한다 — 이 기능은 있으면 좋은 보조 기능이라, 실패해도 예외를 던지지 않고 조용히
    빈 리스트로 넘어간다(호출부에서 굳이 try/except로 감싸지 않아도 되게)."""
    if not pdf_bytes or not isinstance(members, dict):
        return []

    page_boxes = {}
    for key, label in _CATEGORY_KEY_TO_LABEL.items():
        color = _CATEGORY_BBOX_COLOR[label]
        for it in (members.get(key) or []):
            if not isinstance(it, dict):
                continue
            bbox = it.get("bbox")
            if not isinstance(bbox, dict):
                continue
            page = bbox.get("page")
            box_2d = bbox.get("box_2d")
            if not isinstance(page, (int, float)) or not (isinstance(box_2d, list) and len(box_2d) == 4):
                continue
            try:
                page = int(page)
            except (TypeError, ValueError):
                continue
            mark = it.get("mark") or ""
            page_boxes.setdefault(page, []).append((label, mark, box_2d, color))

    return _render_page_boxes_to_annotated(pdf_bytes, page_boxes, max_pages)


def _render_page_boxes_to_annotated(pdf_bytes, page_boxes, max_pages):
    """page_boxes: {page_num: [(category_label_kr, mark, box_2d, color_rgb), ...]} 를 받아
    해당 페이지들만 다시 렌더링하고 부재종류별 색상 박스를 그린 뒤 base64 JPEG 미리보기
    리스트로 반환한다. build_annotated_pages(본 추출 후 전체 부재 색칠)와
    build_basement_plan_preview(본 추출 전 지하주차장 평면도만 가볍게 색칠) 둘 다 이
    렌더링 로직을 공유한다."""
    if not page_boxes:
        return []

    annotated = []
    for page in sorted(page_boxes.keys())[:max_pages]:
        try:
            imgs = _render_pdf_page_range(pdf_bytes, page, page)
            if not imgs:
                continue
            annotated_img = _draw_member_boxes_on_image(imgs[0], page_boxes[page])
            jpeg_bytes = image_to_jpeg_bytes(annotated_img, max_size=(1400, 1400))
            data_url = "data:image/jpeg;base64," + base64.b64encode(jpeg_bytes).decode("ascii")
            annotated.append({
                "page": page,
                "image_data_url": data_url,
                "categories": sorted({b[0] for b in page_boxes[page]}),
                "member_count": len(page_boxes[page]),
            })
        except Exception:
            continue

    return annotated


# ─────────────────────────────────────────────
#  지하주차장 각층평면도 사전 확인 — 본 추출(전체 도면 배치 읽기, 비용 큼) 전에
#  지하주차장 평면도로 감지된 몇 페이지만 가볍게 읽어서 벽/보/슬래브/계단 등 부재 위치만
#  뽑고, build_annotated_pages와 같은 색상 규칙으로 색칠 미리보기를 만든다. 치수/철근은
#  아직 안 읽는다 — "AI가 부재 종류를 제대로 구분하는지"만 사람이 눈으로 먼저 확인하고,
#  틀렸으면 본 추출을 시작하기 전에 바로잡을 수 있게 하기 위한 사전 점검용이다.
# ─────────────────────────────────────────────
BASEMENT_PLAN_CHECK_SYSTEM_PROMPT = """당신은 지하주차장 구조평면도를 판독하는 전문가입니다.
주어진 평면도 이미지에서 보이는 구조부재의 위치만 표시하세요. 치수나 철근 정보는 아직 읽지
마세요 — 이 단계는 부재 종류 인식이 맞는지 사람이 색깔로 먼저 확인하기 위한 사전 점검입니다
(치수/철근은 나중 단계에서 따로 정밀하게 읽습니다).

이미지에 보이는 구조부재마다 아래 정보를 남기세요:
- category: "기초", "기둥", "보", "슬래브", "전단벽", "계단" 중 하나 (반드시 이 6개 중 하나만 쓰세요)
- mark: 도면에 표기된 부재기호(예: W1, G2, C3, SL1). 표기가 안 보이면 null.
- bbox: {"page": <그 부재가 보이는 이미지 위에 표시된 "[도면 N페이지]"의 N>, "box_2d": [ymin, xmin, ymax, xmax]}
  (0~1000 정규화 좌표, 좌상단이 (0,0), 우하단이 (1000,1000))

같은 종류/같은 모양이 반복돼도 평면도에 보이는 개별 위치마다 각각 항목을 만드세요(개수로 묶지
마세요) — 이 단계는 "AI가 벽/보/슬래브/계단을 제대로 구분하는지" 시각적으로 검증하는 용도입니다.

사용자가 이전 확인 단계에서 정정한 내용(있다면 [사용자 확인/정정 내용]으로 주어집니다)이 있으면
그 내용을 최우선 근거로 반영하세요.

반드시 아래 키를 가진 JSON 객체 하나만 반환하세요. 다른 텍스트는 절대 포함하지 마세요.
{
  "members": [
    {"category": "전단벽", "mark": "W1", "bbox": {"page": 1, "box_2d": [400, 100, 900, 250]}},
    {"category": "보", "mark": "G1", "bbox": {"page": 1, "box_2d": [280, 300, 320, 600]}}
  ],
  "notes": ["확인이 필요한 사항"]
}"""

_EMPTY_BASEMENT_PLAN_CHECK = {"members": [], "notes": []}


def extract_basement_plan_members(pdf_bytes, page_numbers, correction_context=None):
    """지하주차장 각층평면도로 감지된 페이지(page_numbers)만 가볍게 읽어서 부재 종류별
    위치(bbox)만 뽑는다 — 본 추출(extract_structural_members)처럼 치수/철근까지 전부 읽지
    않고 "어디에 뭐가 있는지"만 확인해서, 본 추출 전에 색칠 미리보기로 사람이 눈으로
    검증할 수 있게 한다. 페이지 수가 적어(보통 지하층 수만큼, 몇 장) 본 추출보다 훨씬 싸다."""
    client = get_gemini_client()
    if client is None:
        return dict(_EMPTY_BASEMENT_PLAN_CHECK, notes=["GEMINI_API_KEY가 설정되지 않았습니다. .env 확인 후 서버를 재시작해 주세요."])
    if not pdf_bytes or not page_numbers:
        return dict(_EMPTY_BASEMENT_PLAN_CHECK)

    contents = []
    if correction_context:
        contents.append(
            "[사용자 확인/정정 내용 — 최우선 반영]\n" + str(correction_context)[:2000] + "\n"
        )

    page_count = 0
    for page_num in page_numbers:
        try:
            imgs = _render_pdf_page_range(pdf_bytes, page_num, page_num)
        except Exception:
            continue
        if not imgs:
            continue
        contents.append(f"[도면 {page_num}페이지]")
        contents.append(types.Part.from_bytes(data=image_to_jpeg_bytes(imgs[0]), mime_type="image/jpeg"))
        page_count += 1

    if page_count == 0:
        return dict(_EMPTY_BASEMENT_PLAN_CHECK, notes=["지하주차장 평면도로 보이는 페이지를 렌더링하지 못했습니다."])

    try:
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=BASEMENT_PLAN_CHECK_SYSTEM_PROMPT,
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=16384,
            ),
        )
    except Exception as e:
        return dict(_EMPTY_BASEMENT_PLAN_CHECK, notes=[f"지하주차장 평면도 사전 확인 요청 중 오류가 발생했습니다: {str(e)[:200]}"])

    raw = _extract_text_from_gemini_response(response)
    try:
        raw_clean = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(raw_clean)
    except json.JSONDecodeError:
        data = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        if data is None:
            return dict(_EMPTY_BASEMENT_PLAN_CHECK, notes=[f"JSON 파싱 오류 - 원본 응답 확인 필요: {raw[:300]}"])

    if not isinstance(data, dict):
        return dict(_EMPTY_BASEMENT_PLAN_CHECK, notes=["예상치 못한 응답 형식입니다."])
    data.setdefault("members", [])
    data.setdefault("notes", [])
    return data


def build_basement_plan_preview(pdf_bytes, members, max_pages=BASEMENT_PLAN_PAGE_MAX_RESULTS):
    """extract_basement_plan_members()가 뽑은 부재 목록을 build_annotated_pages와 동일한
    색상 규칙(_CATEGORY_BBOX_COLOR)으로 색칠해서 미리보기 이미지 리스트를 만든다."""
    if not pdf_bytes or not members:
        return []
    page_boxes = {}
    for it in members:
        if not isinstance(it, dict):
            continue
        label = it.get("category")
        color = _CATEGORY_BBOX_COLOR.get(label)
        if not color:
            continue
        bbox = it.get("bbox")
        if not isinstance(bbox, dict):
            continue
        page = bbox.get("page")
        box_2d = bbox.get("box_2d")
        if not isinstance(page, (int, float)) or not (isinstance(box_2d, list) and len(box_2d) == 4):
            continue
        try:
            page = int(page)
        except (TypeError, ValueError):
            continue
        mark = it.get("mark") or ""
        page_boxes.setdefault(page, []).append((label, mark, box_2d, color))
    return _render_page_boxes_to_annotated(pdf_bytes, page_boxes, max_pages)


# ─────────────────────────────────────────────
#  건축 입면도/단면도 검토: 층고 + 개구부(창호) 목록 읽기
#  (계산에 쓰이는 게 아니라, 구조 부재 데이터의 층고 누락을 메꾸고
#   개구부 정보의 정합성을 대조하기 위한 참고 데이터)
# ─────────────────────────────────────────────
ELEVATION_SECTION_SYSTEM_PROMPT = """당신은 건축 입면도/단면도를 판독하는 전문가입니다.
입면도(정면/배면/측면)와 단면도에서 아래 두 가지만 "있는 그대로" 읽어서 추출하세요.
계산하지 말고, 도면에 적힌 값을 옮겨 적기만 하세요.

1) 층고(floor_heights): 각 층의 바닥~바닥(FL~FL) 또는 바닥~천장 높이.
   단면도에 GL(지반), 1FL, 2FL... 식으로 레벨이 표기돼 있으면 그 차이로 계산해서 읽어주세요.
   같은 층고가 반복되면 "기준층" 등으로 대표해서 하나만 적고, repeat_count에 몇 개 층이
   그 층고로 반복되는지 적어주세요 (예: 2~15층이 전부 2.9m면 level="기준층(2~15층)",
   height_m=2.9, repeat_count=14). 반복 여부를 모르면 repeat_count=1로 두세요.

2) 개구부(openings): 입면도에 보이는 창호/출입구 등 벽면 개구부 목록.
   위치(어느 입면, 몇 층인지)와 폭/높이, 같은 크기가 반복되면 개수(count)로 묶어서 적으세요.
   정확한 치수를 알 수 없으면 note에 "치수 확인 필요"라고 남기고 width_m/height_m은 null로 두세요.

단위는 전부 미터(m) 기준으로 환산해서 넣으세요 (mm로 적혀 있으면 나눠서 변환).
도면에서 확인할 수 없는 값은 빈 배열로 두고 억지로 만들어내지 마세요.

반드시 아래와 같은 키를 가진 JSON 객체 하나만 반환하세요. 다른 텍스트는 절대 포함하지 마세요.
예시(형식 참고용, 실제 값 아님):
{
  "floor_heights": [{"level": "1층", "height_m": 3.6, "repeat_count": 1, "note": "GL~1FL"}, {"level": "기준층(2~15층)", "height_m": 2.9, "repeat_count": 14, "note": ""}],
  "openings": [{"location": "정면 2~15층 반복", "width_m": 1.5, "height_m": 1.8, "count": 42, "note": "거실 창호"}],
  "notes": ["확인이 필요하거나 근사치인 항목에 대한 메모"]
}"""

_EMPTY_ELEVATION_SECTION = {"floor_heights": [], "openings": [], "notes": []}


def extract_elevation_section_data(dwg_data: dict, pdf_images: list) -> dict:
    """
    건축 입면도/단면도(DWG 파싱 데이터 + PDF 이미지)에서 층고와 개구부(창호) 목록만
    구조화된 JSON으로 추출한다. 물량 계산에는 관여하지 않고, quantity_calc.py가
    구조 부재 데이터의 층고 누락 보완 / 개구부 정합성 대조용 참고자료로만 사용한다.
    """
    client = get_gemini_client()
    if client is None:
        return dict(_EMPTY_ELEVATION_SECTION, notes=["GEMINI_API_KEY가 설정되지 않았습니다. .env 확인 후 서버를 재시작해 주세요."])

    if not pdf_images:
        return dict(_EMPTY_ELEVATION_SECTION)

    contents = [
        f"[건축 입면/단면 DWG 파싱 데이터]\n"
        f"{json.dumps(dwg_data, ensure_ascii=False, indent=2)[:60000]}\n\n"
        "위 데이터와 아래 입면도/단면도 이미지를 보고 층고와 개구부 목록을 읽어서 추출해주세요."
    ]

    for i, img in enumerate(pdf_images[:MAX_PDF_PAGES_TO_GEMINI]):
        contents.append(f"[도면 {i + 1}페이지]")
        contents.append(types.Part.from_bytes(data=image_to_jpeg_bytes(img), mime_type="image/jpeg"))

    try:
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=ELEVATION_SECTION_SYSTEM_PROMPT,
                response_mime_type="application/json",
                temperature=0.0,
                max_output_tokens=32768,
            ),
        )
    except Exception as e:
        return dict(_EMPTY_ELEVATION_SECTION, notes=[f"Gemini 입면/단면 판독 요청 중 오류가 발생했습니다: {str(e)[:200]}"])

    raw = _extract_text_from_gemini_response(response)

    try:
        raw_clean = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(raw_clean)
    except json.JSONDecodeError:
        data = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        if data is None:
            return dict(_EMPTY_ELEVATION_SECTION, notes=[f"JSON 파싱 오류 - 원본 응답 확인 필요: {raw[:300]}"])
        data.setdefault("notes", [])
        data["notes"].append(
            "Gemini 응답이 도중에 잘려(max_output_tokens 한도) 마지막 일부 항목은 누락됐을 수 있습니다 "
            "— 부분 복구된 데이터입니다."
        )

    for key in ("floor_heights", "openings", "notes"):
        data.setdefault(key, [])
    return data


ARCHITECTURAL_SYSTEM_PROMPT = """당신은 건설 건축 수량산출 전문가입니다.
건축도면(PDF 이미지)과 DWG 파싱 데이터를 분석하여 수량을 산출합니다.

반드시 아래 JSON 형식으로만 응답하세요:
{
  "items": [
    {
      "category": "공종 대분류 (예: 조적, 미장, 타일, 창호, 도장, 내장재)",
      "sub_category": "세부 항목",
      "quantity": 숫자,
      "unit": "단위",
      "confidence": "high/medium/low",
      "note": "산출 근거"
    }
  ],
  "summary": "전체 산출 요약",
  "warnings": ["주의 사항"],
  "missing_info": ["확인 불가 항목"]
}"""


def analyze_with_gemini(dwg_data: dict, pdf_images: list, system_prompt: str, drawing_type: str) -> dict:
    """
    DWG 파싱 데이터 + PDF 이미지를 Gemini Vision으로 분석
    Returns: 수량산출 결과 dict
    """
    client = get_gemini_client()
    if client is None:
        return {
            "items": [],
            "summary": "GEMINI_API_KEY가 설정되지 않았습니다.",
            "warnings": [".env 파일에 GEMINI_API_KEY를 추가한 뒤 서버를 재시작해 주세요."],
            "missing_info": [],
        }

    # contents 구성: 파싱 데이터 텍스트 + 도면 이미지(최대 MAX_PDF_PAGES_TO_GEMINI 페이지)
    contents = [
        f"[{drawing_type} DWG 파싱 데이터]\n"
        f"{json.dumps(dwg_data, ensure_ascii=False, indent=2)[:60000]}\n\n"
        "위 데이터의 layer_geometry(레이어별 도형 개수/길이/닫힌면적), block_counts(블록 삽입 개수), "
        "texts(도면 내 문자)를 실제 수량 산출의 근거로 적극 활용하고, 아래 도면 이미지와 교차 검증해 주세요."
    ]

    for i, img in enumerate(pdf_images[:MAX_PDF_PAGES_TO_GEMINI]):
        contents.append(f"[도면 {i + 1}페이지]")
        contents.append(types.Part.from_bytes(data=image_to_jpeg_bytes(img), mime_type="image/jpeg"))

    try:
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL,
            contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=system_prompt,
                response_mime_type="application/json",
                temperature=0.1,
                max_output_tokens=32768,
            ),
        )
    except Exception as e:
        return {
            "items": [],
            "summary": f"Gemini 분석 요청 중 오류가 발생했습니다: {str(e)[:200]}",
            "warnings": [],
            "missing_info": [],
        }

    raw = _extract_text_from_gemini_response(response)

    # JSON 파싱
    try:
        raw_clean = raw.replace("```json", "").replace("```", "").strip()
        return json.loads(raw_clean)
    except json.JSONDecodeError:
        repaired = _try_repair_truncated_json(raw.replace("```json", "").replace("```", "").strip())
        if repaired is not None:
            repaired.setdefault("items", [])
            repaired.setdefault("warnings", [])
            repaired["warnings"].append(
                "Gemini 응답이 도중에 잘려(max_output_tokens 한도) 마지막 일부 항목은 누락됐을 수 있습니다 "
                "— 부분 복구된 데이터입니다."
            )
            repaired.setdefault("summary", "일부 데이터가 잘려 복구된 결과입니다.")
            repaired.setdefault("missing_info", [])
            return repaired
        return {
            "items": [],
            "summary": "JSON 파싱 오류 - 원본 응답 확인 필요",
            "warnings": [raw[:500]],
            "missing_info": [],
        }


GENERAL_NOTES_SYSTEM_PROMPT = """구조도면의 구조일반사항만 판독하세요. 구조평면도, 건축
평·입·단면도, 개구부, 부재 수량, 3D 및 물량은 판독하거나 계산하지 마세요. 없는 값은
추정하지 말고 null/빈 배열로 두세요. 모든 값/행에는
{"file_type":"구조 PDF","pdf_page":null,"drawing_number":"실제 보이는 도면번호 또는 null",
"drawing_title":"실제 보이는 도면명","quote":"조건을 포함한 실제 원문",
"method":"pdf_image","confidence":0.95} 형식의 evidence를 넣으세요. 파일명은 값의
근거가 아닙니다. 범위별 값은 별도 행으로 유지하고 실제 충돌은 덮어쓰지 마세요.
응답 최상위 source_text에는 선택된 페이지에서 실제 판독한 원문만 페이지·구역 제목과
함께 전사하세요. evidence.quote는 source_text에 문자 그대로 존재하는 일부여야 합니다.
concrete_materials는 반드시 "구조재료 및 강도" 제목 아래의 프로젝트 적용 문장에서만
추출하고 evidence.quote에 그 제목/문맥과 적용 위치 문장을 함께 넣으세요. 이음길이·
정착길이·피복두께 구역의 행은 숫자와 무관하게 lookup 데이터이며 프로젝트 적용 재료가
아닙니다. 철근 재료표의 실제 직경 범위·강종·fy를 그대로 분리하고 원문에 없는 강종을
생성하지 마세요. JSON 키는 source_text,
basic_info(structure_system, foundation_type, design_codes,
soil_bearing_capacity, groundwater_buoyancy, seismic_design; 각 값은 value/unit/evidence),
concrete_materials(location,member_type,floor_scope,fck_mpa,slump_mm,aggregate_mm,
exposure_condition,evidence), rebar_materials(diameter_min_mm,diameter_max_mm,grade,
fy_mpa,member_scope,material_type,evidence), cover_requirements(member_type,
exposure_condition,location,thickness_mm,evidence), anchorage_splice_requirements
(requirement_type,bar_size,position,concrete_fck_mpa,value,unit,splice_class,
top_bar_factor,seismic_condition,conditions,evidence), quantity_notes(category,text,
affected_work,evidence), conflicts(field,scope,values,message,evidences),
unconfirmed_items입니다. JSON 객체 하나만 반환하세요."""


def _empty_general_notes_result(reason=None):
    result = {
        "basic_info": {}, "concrete_materials": [], "rebar_materials": [],
        "cover_requirements": [], "anchorage_splice_requirements": [],
        "quantity_notes": [], "conflicts": [], "validation_rejections": [],
        "lookup_data": {"splice_length_rows": [], "anchorage_rows": [], "cover_fck_rows": []},
        "source_text": "",
        "unconfirmed_items": ["구조방식", "기초형식", "적용 구조설계기준", "콘크리트 재료",
                              "철근 재료", "피복두께", "정착·이음 기준"],
        "selected_pages": [], "page_candidates": [], "notes": [reason] if reason else [],
    }
    return _general_notes_add_legacy_fields(result)


def _general_notes_evidence_valid(evidence, selected_pages):
    if not isinstance(evidence, dict):
        return False
    file_type = str(evidence.get("file_type") or "")
    if "PDF" in file_type:
        page_valid = (
            isinstance(evidence.get("pdf_page"), int)
            and evidence["pdf_page"] in set(selected_pages or [])
        )
    elif "CAD" in file_type:
        page_valid = evidence.get("pdf_page") is None and evidence.get("method") == "cad_text"
    else:
        page_valid = False
    return bool(
        page_valid and evidence.get("drawing_number") and evidence.get("drawing_title")
        and str(evidence.get("quote") or "").strip()
        and evidence.get("method")
        and isinstance(evidence.get("confidence"), (int, float))
    )


def _general_notes_quote_supports(evidence, value):
    if value in (None, ""):
        return True
    quote = re.sub(r"\s+", "", str((evidence or {}).get("quote") or "")).upper()
    expected = re.sub(r"\s+", "", str(value)).upper()
    if isinstance(value, (int, float)):
        return _source_supports_number({"quote": quote}, value)
    return expected in quote


def _general_notes_normalize_source(value):
    return re.sub(r"\s+", "", str(value or "")).upper()


def _general_notes_quote_in_source(evidence, source_text):
    quote = _general_notes_normalize_source((evidence or {}).get("quote"))
    source = _general_notes_normalize_source(source_text)
    return bool(quote and source and quote in source)


def _general_notes_row_fields_supported(group, row, evidence):
    fields = {
        "concrete_materials": (
            "fck_mpa", "location", "member_type", "floor_scope",
        ),
        "rebar_materials": (
            "grade", "fy_mpa", "diameter_min_mm", "diameter_max_mm", "member_scope",
        ),
    }.get(group, ())
    return all(
        value in (None, "") or _general_notes_quote_supports(evidence, value)
        for value in (row.get(field) for field in fields)
    )


def _general_notes_is_material_strength_statement(evidence):
    quote = re.sub(r"\s+", "", str((evidence or {}).get("quote") or "")).upper()
    return "구조재료및강도" in quote and not any(
        marker in quote for marker in ("이음길이표", "정착길이표", "피복두께표")
    )


def _general_notes_is_lookup_row(group, row):
    quote = re.sub(r"\s+", "", str(((row or {}).get("evidence") or {}).get("quote") or "")).upper()
    if group == "anchorage_splice_requirements":
        return "표" in quote and any(marker in quote for marker in ("이음", "정착"))
    if group == "cover_requirements":
        return "피복두께" in quote and "표" in quote
    return False


def _general_notes_add_legacy_fields(result):
    concrete = []
    for row in result.get("concrete_materials") or []:
        category = {"벽": "전단벽", "지하외벽": "전단벽", "내부벽": "전단벽"}.get(
            row.get("member_type"), row.get("member_type"),
        )
        if category in ("기초", "기둥", "보", "슬래브", "전단벽", "계단"):
            concrete.append({"category": category, "zone_scope": row.get("location") or row.get("floor_scope"),
                             "fck_mpa": row.get("fck_mpa"), "source": row.get("evidence")})
    rebar = [{"bar_size_min": r.get("diameter_min_mm"), "bar_size_max": r.get("diameter_max_mm"),
              "grade": r.get("grade"), "fy_mpa": r.get("fy_mpa"), "source": r.get("evidence")}
             for r in result.get("rebar_materials") or []]
    cover = []
    for row in result.get("cover_requirements") or []:
        category = {"벽": "전단벽", "지하외벽": "전단벽", "내부벽": "전단벽",
                    "토양에 접하는 부재": "전단벽"}.get(row.get("member_type"), row.get("member_type"))
        if category in ("기초", "기둥", "보", "슬래브", "전단벽", "계단"):
            cover.append({"category": category, "thickness_mm": row.get("thickness_mm"),
                          "condition": row.get("exposure_condition") or row.get("location"),
                          "source": row.get("evidence")})
    laps, anchors = [], []
    for row in result.get("anchorage_splice_requirements") or []:
        common = {"bar_size": row.get("bar_size"), "position": row.get("position"),
                  "length_m": row.get("value") if row.get("unit") == "m" else None,
                  "source": row.get("evidence")}
        if "이음" in str(row.get("requirement_type") or ""):
            laps.append({**common, "splice_class": row.get("splice_class")})
        elif row.get("requirement_type") in ("인장정착", "압축정착", "갈고리"):
            anchors.append({**common, "hook": row.get("requirement_type") == "갈고리"})
    fcks = [r.get("fck_mpa") for r in result.get("concrete_materials") or [] if r.get("fck_mpa") is not None]
    grades = [r.get("grade") for r in result.get("rebar_materials") or [] if r.get("grade")]
    covers = [r.get("thickness_mm") for r in result.get("cover_requirements") or [] if r.get("thickness_mm") is not None]
    result.update({
        "concrete_fck_mpa": fcks[0] if fcks else None, "concrete_fck_table": concrete,
        "rebar_grade": grades[0] if grades else None, "rebar_grade_table": rebar,
        "lap_splice_class": None, "lap_splice_table": laps, "anchorage_table": anchors,
        "cover_thickness_mm": min(covers) if covers else None, "cover_table": cover,
        "seismic_rebar_rules": [],
        "summary_notes": [r.get("text") for r in result.get("quantity_notes") or [] if r.get("text")],
    })
    return result


def _validate_general_notes_result(data, selected_pages, overview=None, source_text=None):
    data = data if isinstance(data, dict) else {}
    result = _empty_general_notes_result()
    source_text = str(source_text if source_text is not None else data.get("source_text") or "")
    result["source_text"] = source_text
    result["selected_pages"] = list(selected_pages or [])
    rejected = []
    basic = data.get("basic_info") if isinstance(data.get("basic_info"), dict) else {}
    for field, item in basic.items():
        evidence = item.get("evidence") if isinstance(item, dict) else None
        if not _general_notes_evidence_valid(evidence, selected_pages):
            rejected.append(f"{field}: 구조화 근거 누락")
        elif not _general_notes_quote_in_source(evidence, source_text):
            rejected.append(f"{field}: evidence.quote 원문 불일치")
        elif not _general_notes_quote_supports(evidence, item.get("value")):
            rejected.append(f"{field}: 원문에 값 없음")
        else:
            result["basic_info"][field] = item
    specs = {
        "concrete_materials": ("fck_mpa", ("location", "member_type", "floor_scope")),
        "rebar_materials": ("grade", ("diameter_min_mm", "diameter_max_mm", "member_scope")),
        "cover_requirements": ("thickness_mm", ("member_type", "exposure_condition", "location")),
        "anchorage_splice_requirements": ("value", ("requirement_type", "bar_size", "conditions")),
        "quantity_notes": ("text", ("category", "affected_work")),
    }
    for group, (value_field, scopes) in specs.items():
        for index, row in enumerate(data.get(group) or []):
            evidence = row.get("evidence") if isinstance(row, dict) else None
            reason = None
            if _general_notes_is_lookup_row(group, row):
                if (
                    not _general_notes_evidence_valid(evidence, selected_pages)
                    or not _general_notes_quote_in_source(evidence, source_text)
                ):
                    rejected.append(f"{group}[{index}]: lookup 원문 불일치")
                    continue
                lookup_key = (
                    "splice_length_rows"
                    if "이음" in str(row.get("requirement_type") or "")
                    else "anchorage_rows"
                ) if group == "anchorage_splice_requirements" else "cover_fck_rows"
                result["lookup_data"][lookup_key].append(row)
                continue
            if not _general_notes_evidence_valid(evidence, selected_pages):
                reason = "구조화 근거 누락"
            elif not _general_notes_quote_in_source(evidence, source_text):
                reason = "evidence.quote 원문 불일치"
            elif row.get(value_field) in (None, "") or not _general_notes_quote_supports(evidence, row.get(value_field)):
                reason = f"{value_field} 원문 근거 없음"
            elif not _general_notes_row_fields_supported(group, row, evidence):
                reason = "필드별 원문 근거 없음"
            elif group == "concrete_materials" and not _general_notes_is_material_strength_statement(evidence):
                reason = "구조재료 및 강도 적용 문장 아님"
            elif not any(row.get(field) not in (None, "") for field in scopes):
                reason = "적용 범위/조건 없음"
            elif group == "anchorage_splice_requirements" and not row.get("unit"):
                reason = "단위 없음"
            if reason:
                rejected.append(f"{group}[{index}]: {reason}")
            else:
                result[group].append(row)
    result["conflicts"] = [c for c in data.get("conflicts") or []
                           if isinstance(c, dict) and len(c.get("evidences") or []) >= 2]
    result["validation_rejections"] = rejected
    result["unconfirmed_items"] = list(dict.fromkeys(list(data.get("unconfirmed_items") or []) + rejected))
    for group, label in (("concrete_materials", "콘크리트 재료"), ("rebar_materials", "철근 재료"),
                         ("cover_requirements", "피복두께"),
                         ("anchorage_splice_requirements", "정착·이음 기준")):
        if not result[group] and label not in result["unconfirmed_items"]:
            result["unconfirmed_items"].append(label)
    structure_item = result["basic_info"].get("structure_system") or {}
    old, new = str((overview or {}).get("structure_type") or ""), str(structure_item.get("value") or "")
    if old and new and not (old == new or ("철근콘크리트" in old and "철근콘크리트" in new)):
        result["conflicts"].append({"field": "structure_system", "scope": "프로젝트 개요 교차검증",
                                    "values": [old, new], "message": "개요 구조형식과 충돌",
                                    "evidences": [structure_item.get("evidence")]})
    result["quantity_notes"].append({
        "category": "이음 정책",
        "text": "전 부재 B급 인장이음 적용",
        "affected_work": "전 부재",
        "source_type": "user_confirmed",
        "provenance": "사용자 확정조건",
        "evidence": None,
    })
    return _general_notes_add_legacy_fields(result)

def _run_general_notes_job(job_id, review_id, structural_pdf_bytes, structural_page_hints,
                           cad_candidate_records=None):
    started = time.monotonic()
    try:
        _progress_set(job_id, "general_spec", 0, 1, "구조일반사항을 확인하고 있어요...",
                      stage_index=1, total_stages=1)
        rec = _review_get(review_id) or {}
        cad_context = []
        if cad_candidate_records:
            parsed, attempted, capped = _parse_precheck_candidates(cad_candidate_records)
            for record in cad_candidate_records:
                info = parsed.get(record.get("content_sha256")) or {}
                if "error" in info:
                    _general_notes_log("cad_parse", job_id=job_id, path=record.get("path"),
                                       success=False, error=str(info.get("error"))[:160])
                    continue
                texts = [str(value) for value in info.get("texts") or [] if str(value).strip()]
                if texts:
                    cad_context.append({"path": record.get("path"),
                                        "filename": record.get("filename"),
                                        "text": "\n".join(texts)})
                _general_notes_log("cad_parse", job_id=job_id, path=record.get("path"),
                                   success=bool(texts), parse_capped=capped,
                                   attempted=record.get("content_sha256") in attempted)
        result = extract_general_notes(
            structural_pdf_bytes, structural_page_hints, rec.get("overview"),
            job_id=job_id, timeout_seconds=GENERAL_NOTES_TIMEOUT_SEC,
            cad_context=cad_context,
        )
        _review_update(review_id, general_spec=result)
        _review_reset_confirmations_from(review_id, "general_spec")
        _result_set(job_id, {"ok": True, "results": {"general_spec": result}})
    except TimeoutError as exc:
        _general_notes_log("job_timeout", level=logging.WARNING, job_id=job_id,
                           stage="candidate_or_extraction",
                           elapsed_seconds=round(time.monotonic() - started, 3),
                           error=str(exc)[:200])
        _result_set(job_id, {"ok": False, "error": f"구조일반사항 확인 시간 초과: {str(exc)[:200]}"})
    except Exception as exc:
        logger.exception("quantity_general_notes_job_failed job_id=%s error=%s", job_id, str(exc)[:200])
        _general_notes_log("job_error", level=logging.ERROR, job_id=job_id,
                           elapsed_seconds=round(time.monotonic() - started, 3),
                           error=str(exc)[:200])
        _result_set(job_id, {"ok": False, "error": f"구조일반사항 확인 중 오류: {str(exc)[:200]}"})
    finally:
        _progress_clear(job_id)


def _run_overview_check_job(job_id, review_id, structural_pdf_bytes, architectural_pdf_bytes,
                            correction_context, architectural_page_hints=None,
                            structural_page_hints=None):
    try:
        _log_overview_diagnostic(
            "job_start",
            job_id=job_id,
            review_id=review_id,
        )
        _progress_set(job_id, "overview", 0, 1, "프로젝트 개요를 확인하고 있어요...",
                      stage_index=1, total_stages=1)
        data = extract_overview_and_spec(
            structural_pdf_bytes, architectural_pdf_bytes, correction_context,
            progress_callback=lambda label: _progress_set(
                job_id, "overview_locator", 0, 1, label,
                stage_index=1, total_stages=1,
            ),
            architectural_page_hints=architectural_page_hints,
            structural_page_hints=structural_page_hints,
        )
        # review_id 상태머신에 이 단계에서 실제로 읽은 overview/general_spec을 기록해둔다 —
        # 사용자가 이후 "확정" 버튼을 누르면(api_quantity_review_confirm) 이 값이 확정된다.
        # 기존 결합 extractor가 호환성상 general_spec도 반환하지만, 3단계 전용 endpoint가
        # 실제 구조 PDF 근거로 다시 판독하기 전에는 review 상태에 저장하지 않는다.
        _review_update(
            review_id,
            overview=data.get("overview"),
            overview_page_detection=data.get("page_detection"),
            general_spec=None,
        )
        # 새 데이터로 덮어썼으니 overview 및 그 이후 단계(general_spec/basement_plan)의
        # 기존 확정은 전부 무효화한다 — 이 review_id로 이전에 이미 확정까지 갔었더라도
        # (예: 같은 세션으로 다른 도면을 다시 확인하는 경우) 새 데이터를 다시 확인받아야 한다.
        _review_reset_confirmations_from(review_id, "overview")
        _result_set(job_id, {"ok": True, "results": data})
        _log_overview_diagnostic(
            "job_complete",
            job_id=job_id,
            review_id=review_id,
            ok=True,
            locator_failed=bool(data.get("locator_failed")),
            missing_fields=_overview_diagnostic_missing_fields(data.get("overview")),
        )
    except OverviewLocatorTimeout as e:
        _log_overview_diagnostic(
            "job_timeout",
            level=logging.WARNING,
            job_id=job_id,
            review_id=review_id,
            stage="overview_locator",
            error=str(e)[:300],
        )
        _result_set(job_id, {
            "ok": False,
            "error": f"사업개요 페이지 자동 탐색 시간 초과: {str(e)[:300]}",
        })
    except Exception as e:
        logger.exception(
            "quantity_overview_job_failed job_id=%s review_id=%s error=%s",
            job_id, review_id, str(e)[:300],
        )
        _log_overview_diagnostic(
            "job_error",
            level=logging.ERROR,
            job_id=job_id,
            review_id=review_id,
            stage="overview_check",
            error=str(e)[:300],
        )
        _result_set(job_id, {"ok": False, "error": f"개요/구조일반사항 확인 중 오류가 발생했습니다: {str(e)[:300]}"})
    finally:
        _progress_clear(job_id)


@require_POST
@_admin_only_json
def api_quantity_overview_check(request):
    """개요/구조일반사항 사전 확인을 백그라운드로 시작한다. 본 추출(api_run_quantity)과
    완전히 동일한 kickoff+폴링 패턴이다 — 다만 표지/구조일반사항 몇 페이지만 쓰므로
    훨씬 빠르고 저렴하다.
    POST 필드: job_id(필수, 폴링용 일회성 토큰), review_id(필수, 이 확인 절차 세션 전체를
    식별하는 값 — 프론트가 흐름 시작 시 한 번만 만들어서 이후 모든 단계에 재사용해야 함),
    structural_pdf, architectural_pdf(둘 다 선택, 최소 1개는 있어야 함),
    correction_context(선택 — 사용자가 이전 확인에서 입력한 정정 텍스트, 재검토용)."""
    job_id = request.POST.get("job_id") or None
    if not job_id:
        return JsonResponse({"error": "job_id가 없습니다."}, status=400)

    # 낱장 파일(예: "A-015,016 사업개요,동별개요.dwg"처럼 파일명 자체에 내용이 적힌
    # 개별 PDF들)을 여러 개 선택해서 올릴 수도 있으므로 getlist로 전부 받는다 —
    # 예전에는 get()이라 프론트가 여러 파일을 보내도 그중 하나만 쓰였다.
    structural_pdf_files = request.FILES.getlist("structural_pdf")
    architectural_pdf_files = request.FILES.getlist("architectural_pdf")
    if not structural_pdf_files and not architectural_pdf_files:
        return JsonResponse({"error": "구조 또는 건축 PDF가 최소 1개는 필요합니다."}, status=400)

    structural_pdf_bytes, structural_page_hints = _merge_uploaded_pdfs(structural_pdf_files)
    architectural_pdf_bytes, architectural_page_hints = _merge_uploaded_pdfs(architectural_pdf_files)
    cad_uploads = _collect_request_cad_uploads(request)
    structural_zip_bytes, architectural_zip_bytes, _cad_merge_info = (
        _merge_uploaded_cad_sets(cad_uploads)
    )
    file_hashes = _review_file_hashes(
        structural_pdf_bytes, architectural_pdf_bytes,
        structural_zip_bytes, architectural_zip_bytes,
    )
    review_id = _canonical_review_id(request.user.pk, file_hashes)
    # 파일 내용이 달라지면 canonical id가 달라지고 새 레코드가 만들어지므로 이전 개요와
    # 모든 confirmed 상태가 새 파일에 승계되지 않는다.
    rec = _review_ensure(review_id, request.user.pk, file_hashes)
    correction_context = (request.POST.get("correction_context") or "").strip() or None
    force = str(request.POST.get("force") or "").lower() in {"1", "true", "yes"}
    if rec.get("overview") is not None and not force and not correction_context:
        _result_set(job_id, {
            "ok": True,
            "results": {
                "overview": rec["overview"],
                "general_spec": rec.get("general_spec"),
                "page_detection": rec.get("overview_page_detection"),
            },
            "cache_hit": True,
        })
        return JsonResponse({
            "accepted": True, "job_id": job_id, "review_id": review_id, "cache_hit": True,
        })
    _progress_set(job_id, "queued", 0, 1, "대기열에 등록됨", stage_index=1, total_stages=1)
    thread = threading.Thread(
        target=_run_overview_check_job,
        args=(job_id, review_id, structural_pdf_bytes, architectural_pdf_bytes,
              correction_context),
        kwargs={
            "architectural_page_hints": architectural_page_hints,
            "structural_page_hints": structural_page_hints,
        },
        daemon=True,
    )
    thread.start()
    return JsonResponse({"accepted": True, "job_id": job_id, "review_id": review_id})


@require_POST
@_admin_only_json
def api_quantity_general_notes_check(request):
    """개요 확정 뒤에만 실행되는 구조일반사항 전용 3단계."""
    job_id = request.POST.get("job_id") or None
    review_id = request.POST.get("review_id") or None
    if not job_id or not review_id:
        return JsonResponse({"error": "job_id와 review_id가 필요합니다."}, status=400)
    rec, err = _review_require_stage(review_id, "overview_confirmed", "프로젝트 개요")
    if err:
        return err
    if rec.get("_user_id") != str(request.user.pk):
        return JsonResponse({"error": "다른 사용자의 확인 절차입니다."}, status=403)
    structural_files = request.FILES.getlist("structural_pdf")
    cad_uploads = _collect_request_cad_uploads(request)
    if not structural_files and not cad_uploads:
        return JsonResponse({"error": "구조 PDF 또는 구조일반사항 CAD 후보가 필요합니다."}, status=400)
    structural_pdf_bytes, page_hints = _merge_uploaded_pdfs(structural_files)
    cad_names = [str(getattr(item, "name", "") or "") for item in cad_uploads]
    cad_inventory = _collect_cad_precheck_inventory(cad_uploads) if cad_uploads else {"records": []}
    general_item = CAD_PRECHECK_STRUCTURAL_ITEMS[0]
    cad_candidates = [
        record for record in cad_inventory.get("records") or []
        if _cad_item_matches(record, general_item)
    ]
    structural_zip_bytes, architectural_zip_bytes, _info = _merge_uploaded_cad_sets(cad_uploads)
    current_hashes = _review_file_hashes(
        structural_pdf_bytes=structural_pdf_bytes,
        structural_zip_bytes=structural_zip_bytes,
        architectural_zip_bytes=architectural_zip_bytes,
    )
    if not _matching_uploaded_files(rec, current_hashes):
        return JsonResponse({"error": "개요 확인 때와 업로드 파일이 다릅니다."}, status=409)
    force = str(request.POST.get("force") or "").lower() in {"1", "true", "yes"}
    if rec.get("general_spec") is not None and not force:
        _result_set(job_id, {
            "ok": True, "results": {"general_spec": rec["general_spec"]}, "cache_hit": True,
        })
        return JsonResponse({"accepted": True, "job_id": job_id, "cache_hit": True})
    _general_notes_log("job_received", job_id=job_id, review_id=review_id,
                       structural_pdf_names=[getattr(f, "name", "") for f in structural_files],
                       cad_candidate_names=cad_names,
                       cad_filename_only=True)
    _progress_set(job_id, "queued", 0, 1, "대기열에 등록됨", stage_index=1, total_stages=1)
    threading.Thread(
        target=_run_general_notes_job,
        args=(job_id, review_id, structural_pdf_bytes, page_hints, cad_candidates),
        daemon=True,
    ).start()
    return JsonResponse({"accepted": True, "job_id": job_id})


def _run_overview_revise_job(job_id, review_id, prior_result, correction_text, revision_stage):
    try:
        _progress_set(job_id, "overview_revise", 0, 1, "정정 내용 반영해서 재검토 중", stage_index=1, total_stages=1)
        data = revise_overview_and_spec_with_text(prior_result, correction_text)
        # 재검토 결과로 review 기록을 갱신한다. 정정된 새 데이터이므로, 혹시 이전에
        # 이미 확정돼 있던 게 있다면(예: 사용자가 확정 후 다시 이 단계로 돌아와 재검토한
        # 경우) 그 확정은 더 이상 유효하지 않다 — overview/general_spec/basement_plan
        # confirmed를 전부 무효화해서 사용자가 새 결과를 다시 확인하게 한다.
        _review_update(review_id, overview=data.get("overview"), general_spec=data.get("general_spec"))
        _review_reset_confirmations_from(review_id, revision_stage)
        _result_set(job_id, {"ok": True, "results": data})
    except Exception as e:
        _result_set(job_id, {"ok": False, "error": f"재검토 중 오류가 발생했습니다: {str(e)[:300]}"})
    finally:
        _progress_clear(job_id)


@require_POST
@_admin_only_json
def api_quantity_overview_revise(request):
    """개요/구조일반사항 "아니요" 정정을 텍스트 전용으로 재검토한다. api_quantity_overview_check와
    달리 도면 이미지를 다시 보내지 않는다 — 이전 추출 결과(JSON)와 사용자 정정 텍스트만 Gemini에
    보내서 토큰을 크게 아낀다.
    POST 필드: job_id(필수), review_id(필수), prior_result(필수, JSON 문자열), correction_text(필수)."""
    job_id = request.POST.get("job_id") or None
    if not job_id:
        return JsonResponse({"error": "job_id가 없습니다."}, status=400)

    review_id = request.POST.get("review_id") or None
    if not review_id:
        return JsonResponse({"error": "review_id가 없습니다."}, status=400)
    rec = _review_get(review_id)
    if rec is None or rec.get("_user_id") != str(request.user.pk):
        return JsonResponse({
            "error": "확인 절차 세션을 찾을 수 없습니다 — 시간이 너무 지났거나 잘못된 review_id입니다. "
                     "개요 확인부터 다시 진행해 주세요.",
        }, status=404)

    correction_text = (request.POST.get("correction_text") or "").strip()
    if not correction_text:
        return JsonResponse({"error": "정정 내용이 없습니다."}, status=400)

    # 브라우저 prior_result는 변조되거나 오래된 값일 수 있으므로 서버 저장값만 사용한다.
    prior_result = {
        "overview": rec.get("overview") or {},
        "general_spec": rec.get("general_spec") or {},
    }
    revision_stage = request.POST.get("stage") or "overview"
    if revision_stage not in ("overview", "general_spec"):
        return JsonResponse({"error": "재검토 단계가 올바르지 않습니다."}, status=400)

    _progress_set(job_id, "queued", 0, 1, "대기열에 등록됨", stage_index=1, total_stages=1)
    thread = threading.Thread(
        target=_run_overview_revise_job,
        args=(job_id, review_id, prior_result, correction_text, revision_stage),
        daemon=True,
    )
    thread.start()
    return JsonResponse({"accepted": True, "job_id": job_id})


# ─────────────────────────────────────────────
#  API: 지하주차장 각층평면도 사전 확인
#  — 구조일반사항 확인이 끝난 뒤, 본 추출(비용 큼) 전에 지하주차장 평면도로 보이는
#  페이지만 찾아서 부재 위치를 가볍게 읽고 색칠 미리보기를 만든다. 사용자가 색으로
#  부재 인식이 맞는지 확인한 뒤에만 비싼 본 추출을 시작한다.
# ─────────────────────────────────────────────
def _run_basement_plan_check_job(job_id, review_id, structural_pdf_bytes, correction_context):
    try:
        _progress_set(job_id, "basement_plan", 0, 1, "지하주차장 평면도 부재 인식 확인 중", stage_index=1, total_stages=1)

        if not structural_pdf_bytes:
            result = {
                "skipped": True, "reason": "구조 PDF가 없어 이 확인은 건너뜁니다.",
                "pages": [], "members": [], "annotated": [],
            }
            _review_update(review_id, basement_plan=result)
            _review_reset_confirmations_from(review_id, "basement_plan")
            _result_set(job_id, {"ok": True, "results": result})
            return

        total_pages = 0
        try:
            info = pdfinfo_from_bytes(structural_pdf_bytes)
            total_pages = int(info.get("Pages", 0) or 0)
        except Exception:
            total_pages = 0

        pages = _detect_basement_plan_pages(structural_pdf_bytes, total_pages)
        if not pages:
            result = {
                "skipped": True, "reason": "지하주차장 평면도로 보이는 페이지를 찾지 못했습니다.",
                "pages": [], "members": [], "annotated": [],
            }
            _review_update(review_id, basement_plan=result)
            _review_reset_confirmations_from(review_id, "basement_plan")
            _result_set(job_id, {"ok": True, "results": result})
            return

        data = extract_basement_plan_members(structural_pdf_bytes, pages, correction_context)
        annotated = build_basement_plan_preview(structural_pdf_bytes, data.get("members") or [])
        result = {
            "skipped": False,
            "pages": pages,
            "members": data.get("members") or [],
            "notes": data.get("notes") or [],
            "annotated": annotated,
        }
        _review_update(review_id, basement_plan=result)
        _review_reset_confirmations_from(review_id, "basement_plan")
        _result_set(job_id, {"ok": True, "results": result})
    except Exception as e:
        _result_set(job_id, {"ok": False, "error": f"지하주차장 평면도 확인 중 오류가 발생했습니다: {str(e)[:300]}"})
    finally:
        _progress_clear(job_id)


@require_POST
@_admin_only_json
def api_quantity_basement_plan_check(request):
    """본 추출 전, 지하주차장 각층평면도만 가볍게 읽어서 부재(벽/보/슬래브/계단 등) 색칠
    미리보기를 만든다. 지하주차장 평면도로 보이는 페이지를 못 찾으면 skipped=True로
    응답해서, 프론트가 이 확인 단계를 건너뛰고 바로 본 추출로 진행할 수 있게 한다.
    구조일반사항이 아직 확정되지 않은 review_id로는 진행할 수 없다(409) — 확인 순서를
    건너뛰고 API를 직접 두드리는 경로를 막기 위함.
    POST 필드: job_id(필수), review_id(필수, 구조일반사항까지 확정돼 있어야 함),
    structural_pdf(필수), correction_context(선택 — "아니요" 정정 재검토용, 해당 페이지
    이미지를 다시 보내되 이 몇 장뿐이라 여전히 저렴함)."""
    job_id = request.POST.get("job_id") or None
    if not job_id:
        return JsonResponse({"error": "job_id가 없습니다."}, status=400)

    review_id = request.POST.get("review_id") or None
    if not review_id:
        return JsonResponse({"error": "review_id가 없습니다."}, status=400)
    rec, err = _review_require_stage(review_id, "general_spec_confirmed", "구조일반사항")
    if err:
        return err
    if rec.get("_user_id") != str(request.user.pk):
        return JsonResponse({"error": "다른 사용자의 확인 세션입니다."}, status=403)

    # getlist: overview-check 단계와 똑같이 낱장 파일 여러 개를 받을 수 있어야 하고,
    # 병합 결과 바이트가 그 단계와 정확히 같아야 아래 해시 비교(_matching_uploaded_files)가
    # 통과한다 — 병합 로직(_merge_uploaded_pdfs)을 그대로 재사용해서 이를 보장한다.
    structural_pdf_files = request.FILES.getlist("structural_pdf")
    if not structural_pdf_files:
        return JsonResponse({"error": "구조 PDF가 필요합니다."}, status=400)

    structural_pdf_bytes, _structural_page_hints = _merge_uploaded_pdfs(structural_pdf_files)
    if not _matching_uploaded_files(
        rec, _review_file_hashes(structural_pdf_bytes=structural_pdf_bytes),
    ):
        _review_reset_confirmations_from(review_id, "overview")
        return JsonResponse({
            "error": "업로드 파일이 바뀌었습니다. 이전 개요와 확정 상태를 초기화했으니 개요부터 다시 확인해 주세요.",
        }, status=409)
    correction_context = (request.POST.get("correction_context") or "").strip() or None

    _progress_set(job_id, "queued", 0, 1, "대기열에 등록됨", stage_index=1, total_stages=1)
    thread = threading.Thread(
        target=_run_basement_plan_check_job,
        args=(job_id, review_id, structural_pdf_bytes, correction_context),
        daemon=True,
    )
    thread.start()
    return JsonResponse({"accepted": True, "job_id": job_id})


@require_POST
@_admin_only_json
def api_quantity_review_confirm(request):
    """개요/구조일반사항/지하주차장 평면도 각 단계를 "확정"한다 — 화면에서 사용자가
    "예, 맞습니다"를 눌렀을 때만 호출돼야 한다. 확정 후에는 그 값이 review_id 기록에
    영구히 남아 이후 단계(특히 api_run_quantity)가 클라이언트가 보내는 값이 아니라
    이 서버 기록을 신뢰하게 된다.
    POST 필드: review_id(필수), stage(필수, "overview" | "general_spec" | "basement_plan")."""
    review_id = request.POST.get("review_id") or None
    if not review_id:
        return JsonResponse({"error": "review_id가 없습니다."}, status=400)

    stage = request.POST.get("stage") or None
    if stage not in REVIEW_STAGE_ORDER:
        return JsonResponse({"error": f"알 수 없는 stage입니다: {stage!r}"}, status=400)

    rec = _review_get(review_id)
    if rec is None:
        return JsonResponse({
            "error": "확인 절차 세션을 찾을 수 없습니다 — 시간이 너무 지났거나 잘못된 review_id입니다. "
            "개요 확인부터 다시 진행해 주세요.",
        }, status=404)
    if rec.get("_user_id") != str(request.user.pk):
        return JsonResponse({"error": "다른 사용자의 확인 세션입니다."}, status=403)

    # 이 stage 앞 단계까지는 이미 확정돼 있어야 한다(순서를 건너뛰고 임의 stage를
    # 확정하는 걸 막기 위함). overview는 첫 단계라 앞 단계가 없다.
    stage_idx = REVIEW_STAGE_ORDER.index(stage)
    if stage_idx > 0:
        prior_stage = REVIEW_STAGE_ORDER[stage_idx - 1]
        if not rec.get(f"{prior_stage}_confirmed"):
            return JsonResponse({
                "error": f"이전 단계({prior_stage})가 아직 확정되지 않았습니다.",
            }, status=409)

    # 확정하려는 stage 자체의 데이터가 서버에 저장돼 있어야 한다(확인 호출을 아예
    # 건너뛰고 바로 확정만 두드리는 경로를 막기 위함). 예외: basement_plan은 애초에
    # "지하주차장 구조 PDF가 아예 없으면 확인 자체가 성립하지 않는" 선택적 단계다 —
    # 프론트가 구조 PDF가 없을 때 이 확인 호출 자체를 건너뛰므로 rec["basement_plan"]이
    # None으로 남을 수 있는데, 그 경우 api_run_quantity 쪽에서도 어차피 structural_pdf/
    # structural_zip이 없으면 구조 부재 추출 자체가 실행되지 않으므로 확정만 통과시켜도
    # "AI 부재인식을 확인 안 하고 넘어가는" 구멍이 생기지 않는다.
    if rec.get(stage) is None and stage != "basement_plan":
        return JsonResponse({
            "error": f"{stage} 확인 데이터가 없습니다 — 먼저 확인을 실행해 주세요.",
        }, status=409)

    ok = _review_update(review_id, **{f"{stage}_confirmed": True})
    if not ok:
        return JsonResponse({"error": "확정 처리 중 세션을 찾지 못했습니다."}, status=404)

    return JsonResponse({"ok": True, "review_id": review_id, "stage": stage, "confirmed": True})


# ─────────────────────────────────────────────
#  API: 수량산출 실행 (메인 엔드포인트)
# ─────────────────────────────────────────────
@require_POST
@_admin_only_json
def api_run_quantity(request):
    """
    파일 4개 받아서 수량산출을 백그라운드 스레드로 시작하고 즉시 응답한다.
    - structural_zip: 구조도면 ZIP (DWG)
    - structural_pdf: 구조도면 합본 PDF
    - architectural_zip: 건축도면 ZIP (DWG)
    - architectural_pdf: 건축도면 합본 PDF

    예전에는 이 요청 하나가 Gemini 배치 호출을 전부 끝낼 때까지(실제 도면 기준
    수십 분 걸릴 수 있음) 응답을 내보내지 않고 연결을 계속 붙잡고 있었다. 문제는
    응답 바이트가 1바이트도 안 나가는 채로 이렇게 오래 연결이 떠 있으면, 브라우저/
    OS/공유기 등 어딘가의 유휴(idle) 커넥션 타임아웃에 걸려 "네트워크 오류:
    Load failed"로 끊겨버릴 수 있다는 것 — 실제로 재현됐다. 이때 서버는 클라이언트가
    끊긴 걸 알 방법이 없어서 이미 시작한 Gemini 호출들을 계속 실행한다: 돈은
    나가는데 사용자는 결과를 영영 못 받는 최악의 조합이다.

    그래서 지금은 실제 작업(파일 파싱 + Gemini 호출)을 별도 스레드(_run_quantity_job)로
    던져놓고, 이 뷰는 파일을 바이트로 읽어 확보한 뒤 곧장 {"accepted": true, "job_id"}를
    돌려준다. 프론트엔드는 이미 갖고 있던 진행률 폴링(api_quantity_progress)을
    계속 돌리다가 done=true가 되면 같이 담겨오는 결과를 쓴다. 폴링은 몇 초짜리
    요청이 반복되는 구조라 중간에 한두 번 끊겨도 다음 폴링에서 다시 잡히고,
    서버 쪽 작업 자체는 브라우저 연결 상태와 무관하게 끝까지 실행된다.
    """
    cad_uploads = _collect_request_cad_uploads(request)
    # getlist: overview-check 단계와 동일하게 낱장 PDF 여러 개를 받을 수 있어야 하고,
    # 병합 결과가 그 단계와 바이트 단위로 똑같아야 아래 해시 비교가 통과한다 —
    # _merge_uploaded_pdfs를 그대로 재사용해서 이를 보장한다.
    structural_pdf_files = request.FILES.getlist("structural_pdf")
    architectural_pdf_files = request.FILES.getlist("architectural_pdf")

    if not any([cad_uploads, structural_pdf_files, architectural_pdf_files]):
        return JsonResponse({"error": "파일이 없습니다."}, status=400)

    # job_id는 프론트엔드가 업로드와 함께 보내는 임의 문자열이다. 백그라운드로
    # 처리한 결과를 나중에 어디로 돌려줘야 할지 이 값 하나로 찾으므로, 이제는
    # (진행률 표시가 선택사항이던 예전과 달리) 필수값이다.
    job_id = request.POST.get("job_id") or None
    if not job_id:
        return JsonResponse({"error": "job_id가 없습니다 — 새로고침 후 다시 시도해 주세요."}, status=400)

    # review_id 상태머신 강제: 개요→구조일반사항→지하주차장 평면도 확인을 전부
    # 확정하지 않고서는 본 추출(비용이 큰 실제 작업)을 시작할 수 없다. 클라이언트가
    # confirmed_general_spec을 뭘 보내든, 아래에서 서버가 review_id 기록으로 직접
    # 확정된 general_spec을 가져와 덮어쓴다 — API를 직접 두드려 임의의 구조일반사항을
    # 주입하거나 확인 단계를 생략하는 경로를 막기 위함.
    review_id = request.POST.get("review_id") or None
    if not review_id:
        return JsonResponse({"error": "review_id가 없습니다 — 개요/구조일반사항 확인부터 다시 진행해 주세요."}, status=400)
    review_rec, err = _review_require_stage(review_id, "basement_plan_confirmed", "지하주차장 평면도")
    if err:
        return err
    if review_rec.get("_user_id") != str(request.user.pk):
        return JsonResponse({"error": "다른 사용자의 확인 세션입니다."}, status=403)

    # 백그라운드 스레드에서는 Django의 request/FILES 객체를 더 이상 안전하게 쓸 수
    # 없다(요청-응답 주기가 끝나면 내부적으로 임시파일 등이 정리될 수 있음). 그래서
    # 응답을 만들기 전, 즉 요청이 아직 살아있는 지금 필요한 바이트를 전부 메모리로
    # 읽어두고 그 바이트만 스레드에 넘긴다.
    structural_zip_bytes, architectural_zip_bytes, cad_merge_info = (
        _merge_uploaded_cad_sets(cad_uploads)
    )
    structural_pdf_bytes, _structural_page_hints = _merge_uploaded_pdfs(structural_pdf_files)
    architectural_pdf_bytes, _architectural_page_hints = _merge_uploaded_pdfs(architectural_pdf_files)
    logger.info(
        "quantity_run_cad_uploads upload_count=%s upload_names=%s cad_count=%s "
        "structural_count=%s architectural_count=%s scan_complete=%s",
        cad_merge_info.get("upload_count"), cad_merge_info.get("upload_names"),
        cad_merge_info.get("cad_count"), cad_merge_info.get("structural_count"),
        cad_merge_info.get("architectural_count"), cad_merge_info.get("scan_complete"),
    )
    current_hashes = _review_file_hashes(
        structural_pdf_bytes, architectural_pdf_bytes,
        structural_zip_bytes, architectural_zip_bytes,
    )
    if not _matching_uploaded_files(review_rec, current_hashes):
        _review_reset_confirmations_from(review_id, "overview")
        return JsonResponse({
            "error": "업로드 파일이 바뀌었습니다. 이전 개요와 모든 확정 상태를 초기화했습니다.",
        }, status=409)

    # 개요/구조일반사항 사전 확인 단계에서 사용자가 확정한 general_spec — 클라이언트가
    # 보낸 값이 아니라 review_id 기록에 서버가 직접 저장해둔 값을 쓴다(신뢰의 기준점을
    # 서버로 옮김). general_spec_confirmed까지 확정돼야 이 함수까지 올 수 있으므로
    # (basement_plan_confirmed는 general_spec_confirmed보다 뒤 단계) 이 시점에는
    # review_rec["general_spec"]이 반드시 채워져 있다.
    confirmed_general_spec = review_rec.get("general_spec") or None

    progress_stages = []
    if architectural_zip_bytes or architectural_pdf_bytes:
        progress_stages += ["architectural", "elevation"]
    if structural_zip_bytes or structural_pdf_bytes:
        progress_stages.append("structural")
    progress_total_stages = len(progress_stages) or 1

    # 스레드가 실제로 시작해서 첫 _report_progress를 부르기 전까지는 진행률
    # 저장소가 비어있어, 그 짧은 틈에 프론트가 첫 폴링을 하면 found=false를 받고
    # "아직 시작 안 했나?"로 헷갈릴 수 있다. 미리 "대기열에 등록됨" 상태를 하나
    # 넣어둬서 그 틈을 없앤다.
    _progress_set(job_id, "queued", 0, 1, "대기열에 등록됨", stage_index=1, total_stages=progress_total_stages)

    thread = threading.Thread(
        target=_run_quantity_job,
        args=(job_id, structural_zip_bytes, structural_pdf_bytes,
              architectural_zip_bytes, architectural_pdf_bytes,
              progress_stages, progress_total_stages, confirmed_general_spec),
        daemon=True,
    )
    thread.start()

    return JsonResponse({"accepted": True, "job_id": job_id})


def _run_quantity_job(job_id, structural_zip_bytes, structural_pdf_bytes,
                       architectural_zip_bytes, architectural_pdf_bytes,
                       progress_stages, progress_total_stages, confirmed_general_spec=None):
    """api_run_quantity가 백그라운드 스레드로 실행시키는 실제 작업 본체.
    HTTP 요청/응답과 완전히 분리돼 있어서, 이 함수를 시작시킨 브라우저 연결이
    중간에 끊기든 말든 이 함수는 끝까지 실행되고, 최종 결과는 _RESULT_STORE에
    job_id로 저장된다 — 나중에 같은 job_id로 폴링하면(페이지를 새로고침해서
    다시 접속해도) 결과를 받을 수 있다."""

    def _report_progress(stage, current, total, label):
        try:
            stage_idx = progress_stages.index(stage) + 1
        except ValueError:
            stage_idx = 1
        _progress_set(job_id, stage, current, total, label,
                       stage_index=stage_idx, total_stages=progress_total_stages)

    def _cancelled():
        return _is_cancelled(job_id)

    results = {}
    elevation_data = None
    cancelled_note_added = False

    def _note_cancelled_once():
        nonlocal cancelled_note_added
        if not cancelled_note_added:
            cancelled_note_added = True
            return "사용자가 취소를 요청하여 이후 단계는 처리하지 않았습니다 — 여기까지 처리된 결과만 표시합니다."
        return None

    try:
        try:
            # ── 건축 수량산출 + 입면도/단면도 검토 (구조 계산보다 먼저 실행 —
            #    여기서 읽은 층고/개구부를 구조 계산의 층고 대체값/개구부 대조에 사용) ──
            if (architectural_zip_bytes or architectural_pdf_bytes) and _cancelled():
                results["architectural"] = {"items": [], "missing_info": [_note_cancelled_once() or "사용자가 취소를 요청했습니다."], "warnings": []}
            elif architectural_zip_bytes or architectural_pdf_bytes:
                arch_dwg_data = {}
                arch_pdf_images = []

                if architectural_zip_bytes:
                    # 2026-07-27: REQUIRED_ARCHITECTURAL의 임의 코드(A-001 등)로 필터링하면
                    # 실제 프로젝트 파일명과 안 맞아 매칭 0건이 될 수 있으므로, ZIP 안의
                    # 모든 dwg/dxf를 대상으로 한다(코드 목록에 의존하지 않음).
                    arch_dwg_data = parse_dwg_from_zip(architectural_zip_bytes)

                if architectural_pdf_bytes:
                    try:
                        arch_pdf_images = pdf_to_images(architectural_pdf_bytes, max_pages=MAX_PDF_PAGES_TO_GEMINI)
                    except Exception as e:
                        arch_pdf_images = []
                        results["architectural_pdf_error"] = str(e)

                if arch_dwg_data or arch_pdf_images:
                    _report_progress("architectural", 0, 1, "건축도면 분석 중")
                    try:
                        results["architectural"] = analyze_with_gemini(
                            arch_dwg_data, arch_pdf_images,
                            ARCHITECTURAL_SYSTEM_PROMPT, "건축도면"
                        )
                    except Exception as e:
                        results["architectural"] = {"error": str(e), "items": [], "missing_info": [f"건축 수량산출 처리 중 오류가 발생했습니다: {str(e)[:300]}"], "warnings": []}
                    _report_progress("architectural", 1, 1, "건축도면 분석 완료")

                    if _cancelled():
                        elevation_data = dict(_EMPTY_ELEVATION_SECTION, notes=[_note_cancelled_once() or "사용자가 취소를 요청했습니다."])
                    else:
                        # 입면도/단면도 검토: 층고 + 개구부(창호) 목록만 별도로 읽음
                        # (건축 수량산출과 같은 파싱 데이터/이미지를 재사용해 중복 변환을 피한다)
                        _report_progress("elevation", 0, 1, "입면/단면 검토 중")
                        try:
                            elevation_data = extract_elevation_section_data(arch_dwg_data, arch_pdf_images)
                        except Exception as e:
                            elevation_data = dict(_EMPTY_ELEVATION_SECTION, notes=[f"입면/단면 검토 중 오류: {str(e)[:200]}"])
                        _report_progress("elevation", 1, 1, "입면/단면 검토 완료")
                elif architectural_zip_bytes or architectural_pdf_bytes:
                    # 파일은 올렸지만 arch_dwg_data/arch_pdf_images 둘 다 비어서(PDF 변환 실패 등)
                    # 분석을 아예 시작도 못한 경우 — 이 사유를 architectural.missing_info에 남겨서
                    # 프론트엔드가 결과가 0건일 때 "왜"인지 보여줄 수 있게 한다 (그냥 조용히
                    # results["architectural"] 키 자체가 안 생기면 원인을 알 방법이 없어진다).
                    pdf_err = results.get("architectural_pdf_error")
                    reason = (
                        f"건축도면 PDF를 이미지로 변환하는 중 오류가 발생했습니다: {pdf_err}"
                        if pdf_err else
                        "건축도면 파일에서 읽을 수 있는 데이터가 없습니다 — ZIP/PDF가 비어있거나 손상됐을 수 있습니다."
                    )
                    results["architectural"] = {"items": [], "missing_info": [reason], "warnings": []}

            # ── 구조 수량산출 ──
            if structural_zip_bytes or structural_pdf_bytes:
                dwg_data = {}

                if structural_zip_bytes:
                    # 2026-07-27: 여기서 REQUIRED_STRUCTURAL의 임의 코드(S-001~S-501 등)로
                    # 필터링하던 게 실제 프로젝트(부천 현장 등)의 실제 도면번호 규칙과 전혀
                    # 안 맞아 항상 매칭 0건 -> dwg_data가 매번 빈 딕셔너리로 나왔을 가능성이
                    # 크다(구조 PDF 기반 Gemini 추출은 그대로 진행되지만, ZIP으로 올린 DWG의
                    # 기하 데이터가 보조 자료로 전혀 반영되지 않고 조용히 버려진 것). 도면번호
                    # 목록에 의존하지 않고 ZIP 안의 모든 dwg/dxf를 대상으로 바꾼다.
                    dwg_data = parse_dwg_from_zip(structural_zip_bytes)

                # structural_pdf_bytes는 이미 파라미터로 받은 원본 바이트 그대로 사용한다 —
                # 여기서 미리 전부 이미지로 렌더링하지 않는다(대형 도면집을 한꺼번에
                # 렌더링하면 메모리를 너무 많이 써서 서버가 죽을 수 있다). extract_structural_
                # members()가 배치 단위로 필요한 페이지만 그때그때 렌더링한다.

                if _cancelled() and (dwg_data or structural_pdf_bytes):
                    results["structural"] = {
                        "items": [],
                        "missing_info": [_note_cancelled_once() or "사용자가 취소를 요청했습니다."],
                        "warnings": [],
                    }
                elif dwg_data or structural_pdf_bytes:
                    try:
                        # 1단계: Gemini는 '읽기'만 (구조화된 부재 리스트). progress_cb로 배치별
                        # 진행상황("배치 3/6 처리 중")을 폴링 저장소에 실시간으로 남긴다. cancel_cb는
                        # 취소 요청이 들어오면 아직 시작 안 한 배치들을 건너뛰게 한다(비용 절약).
                        def _structural_progress_cb(idx, total):
                            _report_progress("structural", idx, total, f"구조 부재 추출 배치 {idx}/{total} 처리 중")

                        members = extract_structural_members(
                            dwg_data, structural_pdf_bytes, progress_cb=_structural_progress_cb, cancel_cb=_cancelled,
                            confirmed_general_spec=confirmed_general_spec,
                        )
                        # 색칠된 도면 미리보기 — Gemini가 채운 bbox(부재 위치)를 원본 페이지 위에
                        # 색상 박스로 그려서, 의뢰인이 채팅창에서 놓친 벽/보가 있는지 직접 확인할
                        # 수 있게 한다. bbox 데이터가 없거나 실패해도 예외를 던지지 않는 보조
                        # 기능이라, 여기서 문제가 생겨도 이후 흐름에는 영향이 없다.
                        try:
                            annotated_pages = build_annotated_pages(structural_pdf_bytes, members)
                        except Exception as e:
                            annotated_pages = []
                            results.setdefault("structural_warnings_pre", []).append(
                                f"색칠된 도면 미리보기 생성 중 오류가 발생했습니다: {str(e)[:200]}"
                            )

                        if job_id and not _cancelled():
                            # ── 실제 물량 계산 전에 사용자 확인을 거치게 한다 ──
                            # 여기서 바로 계산하지 않고 members를 job_id로 저장해두고, 프론트가
                            # "도면 확인" 팝업에서 사용자 확인/수정을 받은 뒤 /api/quantity/confirm-review/
                            # 로 다시 요청해야 실제 계산(compute_structural_quantities)이 실행된다.
                            _extraction_store_set(job_id, members, elevation_data)
                            incomplete_reasons = _extraction_incomplete_reasons(members)
                            results["structural"] = {
                                "review_required": True,
                                "job_id": job_id,
                                "annotated_pages": annotated_pages,
                                "checklist": _build_review_checklist(members),
                                "warnings": results.pop("structural_warnings_pre", []),
                                # 응답 잘림/배치 실패/취소 등으로 이 추출 결과가 온전하지 않을 때, 검토
                                # 팝업이 "체크리스트만 보면 끝"이 아니라 눈에 띄게 경고하고 명시적 확인을
                                # 받게 하기 위한 플래그 — 조용히 부분 결과로 확정 계산까지 가는 걸 막는다.
                                "extraction_incomplete": bool(incomplete_reasons),
                                "incomplete_reasons": incomplete_reasons[:5],
                            }
                        else:
                            # job_id가 없는 경우를 위한 방어적 fallback — 리뷰 단계 없이 바로 계산.
                            # 검토 팝업을 거치지 않으므로 _apply_member_corrections의 검증도 못 받는다 —
                            # 그래서 여기서 직접 _sanitize_raw_members를 거쳐 원본값(count=0/음수 등)을
                            # 계산에 넣기 전에 한 번 더 걸러낸다.
                            incomplete_reasons = _extraction_incomplete_reasons(members)
                            sanitized_members = _sanitize_raw_members(members)
                            results["structural"] = compute_structural_quantities(sanitized_members, elevation_data)
                            results["structural"]["_raw_members"] = sanitized_members  # 디버그/검증용
                            if elevation_data:
                                results["structural"]["_elevation_section"] = elevation_data  # 디버그/검증용
                            results["structural"]["massing"] = compute_massing_model(sanitized_members, elevation_data)
                            results["structural"]["annotated_pages"] = annotated_pages
                            results["structural"]["extraction_incomplete"] = bool(incomplete_reasons)
                            results["structural"]["incomplete_reasons"] = incomplete_reasons[:5]
                            for w in results.pop("structural_warnings_pre", []):
                                results["structural"].setdefault("warnings", []).append(w)
                    except Exception as e:
                        results["structural"] = {
                            "error": str(e), "items": [],
                            "missing_info": [f"구조 수량산출 처리 중 오류가 발생했습니다: {str(e)[:300]}"],
                            "warnings": [],
                        }
                else:
                    # 파일은 올렸지만 dwg_data/structural_pdf_bytes 둘 다 비어서(빈 ZIP, 빈 PDF 등)
                    # 추출을 아예 시작도 못한 경우 — 이전에는 이 경우 results에 "structural" 키
                    # 자체가 생기지 않아, 프론트엔드에서 "왜 실패했는지" 전혀 알 수 없는 상태로
                    # 그냥 "추출하지 못했어요"만 표시됐다. 실제 원인을 missing_info에 남겨서
                    # 화면에 보이도록 한다.
                    results["structural"] = {
                        "items": [],
                        "missing_info": ["구조도면 파일에서 읽을 수 있는 데이터가 없습니다 — ZIP/PDF가 비어있거나 손상됐을 수 있습니다."],
                        "warnings": [],
                    }
        finally:
            # 작업이 끝나면(성공/실패 무관) 진행률/취소 항목을 정리한다. 결과는 그 앞에
            # _result_set으로 먼저 저장해둬야, 정리와 폴링이 같은 순간에 겹쳐도 프론트가
            # "진행 중도 아니고 결과도 없음" 틈새를 보지 않는다.
            _result_set(job_id, {"ok": True, "results": results})
            _progress_clear(job_id)
            _cancel_clear(job_id)
    except Exception as e:
        # 위 로직 내부의 각 단계는 이미 자체적으로 try/except를 갖고 있어 개별 오류를
        # results 안에 담아 처리하지만, 혹시 그 바깥에서 예상 못한 예외가 튀어나오는
        # 경우까지의 안전망이다 — 이게 없으면 스레드가 조용히 죽어버리고 프론트는
        # 영원히 "진행 중"만 보게 된다(그리고 그 시점까지 쓴 Gemini 비용은 그냥 날아간다).
        _result_set(job_id, {"ok": False, "error": f"수량산출 처리 중 예상치 못한 오류가 발생했습니다: {str(e)[:300]}"})
        _progress_clear(job_id)
        _cancel_clear(job_id)


@require_POST
@_admin_only_json
def api_quantity_confirm_review(request):
    """도면 확인 팝업에서 "진행" 버튼을 눌렀을 때 호출되는 엔드포인트.
    api_run_quantity가 job_id로 임시 저장해둔 추출 결과(members)를 꺼내와서,
    사용자가 체크리스트에서 지정한 corrections(제거/수정)을 반영한 뒤에야
    실제 물량 계산(compute_structural_quantities)을 실행한다.
    corrections 형식(JSON 문자열, POST의 "corrections" 필드):
      [{"review_id": "columns:3", "action": "remove"},
       {"review_id": "walls:5", "action": "edit", "fields": {"count": 4, "length_m": 3.6}}]
    """
    job_id = request.POST.get("job_id") or None
    if not job_id:
        return JsonResponse({"error": "job_id가 없습니다."}, status=400)

    # pop이 아니라 get(읽기 전용)으로 꺼낸다 — 여기서 바로 지워버리면, 아래 계산 중
    # 예외가 나서 500을 반환했을 때 저장 데이터가 이미 사라져서 사용자가 재시도할
    # 방법이 없어(Gemini 추출부터 다시 해야 함) 비용 낭비가 컸다. 계산이 실제로
    # 성공했을 때만 아래에서 별도로 pop해서 지운다.
    stored = _extraction_store_get(job_id)
    if not stored:
        return JsonResponse({
            "error": "검토할 추출 결과를 찾을 수 없습니다 — 시간이 너무 지났거나(2시간 초과) "
                     "이미 처리된 job_id일 수 있습니다. 도면을 다시 업로드해서 분석해 주세요.",
        }, status=404)

    raw_corrections = request.POST.get("corrections") or "[]"
    try:
        corrections = json.loads(raw_corrections)
        if not isinstance(corrections, list):
            corrections = []
    except (TypeError, ValueError):
        corrections = []

    members = stored.get("members") or _EMPTY_MEMBERS
    elevation_data = stored.get("elevation_data")

    try:
        corrected_members, rejected_notes = _apply_member_corrections(members, corrections)
        structural_results = compute_structural_quantities(corrected_members, elevation_data)
        structural_results["_raw_members"] = corrected_members  # 디버그/검증용
        if elevation_data:
            structural_results["_elevation_section"] = elevation_data  # 디버그/검증용
        structural_results["massing"] = compute_massing_model(corrected_members, elevation_data)
        applied_count = len(corrections) - len(rejected_notes)
        if applied_count > 0:
            structural_results.setdefault("warnings", []).insert(
                0, f"도면 확인 단계에서 {applied_count}건의 수정사항이 반영됐습니다."
            )
        for note in rejected_notes:
            structural_results.setdefault("warnings", []).append(note)

        # 추출이 애초에 불완전했다면(응답 잘림/배치 실패/취소) 검토 팝업 단계에서 이미
        # 경고했지만, 확정 계산 결과(엑셀/최종 요약)에도 이 사실이 묻히지 않게 계속 남긴다.
        incomplete_reasons = _extraction_incomplete_reasons(members)
        structural_results["extraction_incomplete"] = bool(incomplete_reasons)
        if incomplete_reasons:
            structural_results.setdefault("warnings", []).insert(
                0, "⚠ 이 결과는 도면 인식이 불완전한 상태(응답 잘림/배치 실패 등)에서 확정됐습니다 — "
                   "전문가 검토 없이 확정 물량으로 사용하지 마세요."
            )
    except Exception as e:
        # 계산 실패 — 저장 데이터를 지우지 않았으니(_extraction_store_get을 썼음) 같은
        # job_id/corrections로 다시 POST하면 Gemini 재추출 없이 재시도할 수 있다.
        return JsonResponse({
            "error": f"수정사항 반영 후 물량 계산 중 오류가 발생했습니다: {str(e)[:300]} "
                     "— 도면을 다시 분석하지 않아도 같은 화면에서 다시 시도할 수 있습니다.",
        }, status=500)

    # 계산이 성공했을 때만 저장 데이터를 지운다.
    _extraction_store_pop(job_id)
    return JsonResponse({"ok": True, "results": {"structural": structural_results}})


@require_POST
@_admin_only_json
def api_quantity_cancel(request):
    """취소 요청 엔드포인트. 프론트엔드가 "취소" 버튼을 누르면 job_id를 담아 호출한다.
    api_run_quantity가 아직 처리 중이면(같은 job_id로 폴링 저장소에 진행상황이 있으면)
    다음 체크포인트(다음 배치/다음 단계 시작 전)에서 이 요청을 확인하고 이후 단계를
    건너뛴다. 이미 시작된 Gemini 호출 1개는 중간에 끊을 방법이 없어 끝까지 처리된다."""
    job_id = request.POST.get("job_id") or None
    if not job_id:
        return JsonResponse({"error": "job_id가 없습니다."}, status=400)
    _cancel_request(job_id)
    return JsonResponse({"ok": True})


@require_POST
@_admin_only_json
def api_quantity_progress(request):
    """진행률 폴링 엔드포인트. 프론트엔드가 api_run_quantity 요청을 보낸 직후부터
    주기적으로(예: 1.5초마다) job_id를 담아 호출해서 지금 어느 단계/배치까지 됐는지 읽는다.

    api_run_quantity가 백그라운드 스레드로 바뀌면서, 이 엔드포인트가 진행상황뿐 아니라
    "다 끝났을 때 최종 결과"까지 같이 실어 나르는 역할을 겸한다 — 별도의 결과 조회
    엔드포인트를 새로 안 만들고, 프론트가 이미 돌리고 있던 폴링 루프 하나로 충분하게
    했다. 결과 저장소(_RESULT_STORE)를 먼저 확인해서 있으면 done=true와 함께 그대로
    돌려주고, 없으면 진행 중 저장소(_PROGRESS_STORE)를 본다. 둘 다 없으면(아직 시작
    전이거나 TTL로 이미 정리된 경우) found=false로 응답한다 — 프론트는 이 경우 "진행
    중"으로만 표시하고 굳이 오류 취급하지 않는다."""
    job_id = request.POST.get("job_id") or request.GET.get("job_id") or None
    if not job_id:
        return JsonResponse({"found": False})

    result = _result_get(job_id)
    if result is not None:
        payload = {"found": True, "done": True}
        payload.update({k: v for k, v in result.items() if k != "_created_at"})
        return JsonResponse(payload)

    progress = _progress_get(job_id)
    if not progress:
        return JsonResponse({"found": False, "done": False})
    return JsonResponse({
        "found": True,
        "done": False,
        "stage": progress.get("stage"),
        "current": progress.get("current"),
        "total": progress.get("total"),
        "label": progress.get("label"),
        "stage_index": progress.get("stage_index"),
        "total_stages": progress.get("total_stages"),
    })


# ─────────────────────────────────────────────
#  API: 엑셀 다운로드
# ─────────────────────────────────────────────
_XLSX_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _xlsx_injection_guard(value):
    """엑셀/구글시트 수식 인젝션(CSV/XLSX Injection) 방지.

    openpyxl은 셀 값이 "="로 시작하는 문자열이면 자동으로 그 셀을 수식(formula) 타입으로
    인식해서 저장한다 — 즉 Gemini가 도면에서 읽어온 mark/note/zone 같은 텍스트에 우연히
    (또는 악의적으로 조작된 도면 파일을 통해 의도적으로) "=1+1", "=CMD(...)" 같은 문자열이
    들어있으면, 그게 그대로 실행 가능한 엑셀 수식이 돼서 파일을 여는 사람의 PC에서
    임의 동작(외부 프로그램 실행 등, 구버전 Excel의 DDE 취약점 포함)으로 이어질 수 있다.
    "+", "-", "@"로 시작하는 값도 일부 스프레드시트 프로그램에서 수식으로 해석될 수 있어
    함께 막는다. 이런 문자로 시작하면 앞에 작은따옴표를 붙여 순수 텍스트로 강제한다."""
    if isinstance(value, str) and value[:1] in _XLSX_FORMULA_TRIGGER_CHARS:
        return "'" + value
    return value


def _sanitize_for_excel(obj):
    """results(=Gemini 추출 결과가 섞여 있는 JSON)를 엑셀로 옮기기 전에 재귀적으로 훑어서
    모든 문자열값에 _xlsx_injection_guard를 적용한다. write_sheet를 비롯한 모든 시트
    작성 코드가 어디서 값을 꺼내 쓰든 이 한 번의 전처리로 전부 보호되게 하기 위함이다
    (각 ws.cell(value=...) 호출부마다 개별적으로 방어하는 것보다 훨씬 빠뜨리기 어렵다)."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_excel(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_excel(v) for v in obj]
    return _xlsx_injection_guard(obj)


def _build_excel_response(request):
    """
    수량산출 결과 JSON을 받아 엑셀 파일로 반환
    Body: { "results": { "structural": {...}, "architectural": {...} } }
    """
    try:
        body = json.loads(request.body)
        results = _sanitize_for_excel(body.get("results", {}))
    except (json.JSONDecodeError, AttributeError):
        return HttpResponse("잘못된 요청입니다.", status=400)

    wb = openpyxl.Workbook()

    # 스타일 정의
    navy_fill = PatternFill("solid", fgColor="073A5B")
    yellow_fill = PatternFill("solid", fgColor="F9C20A")
    green_fill = PatternFill("solid", fgColor="41a86b")
    light_fill = PatternFill("solid", fgColor="F0FAF4")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    sub_header_font = Font(name="맑은 고딕", bold=True, color="073A5B", size=10)
    body_font = Font(name="맑은 고딕", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_sheet(ws, data, sheet_title):
        ws.title = sheet_title

        # 제목 행
        ws.merge_cells("A1:G1")
        ws["A1"] = f"AI 수량산출 결과 — {sheet_title}"
        ws["A1"].font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
        ws["A1"].fill = navy_fill
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 32

        # 요약
        summary = data.get("summary", "")
        ws.merge_cells("A2:G2")
        ws["A2"] = f"요약: {summary}"
        ws["A2"].font = Font(name="맑은 고딕", italic=True, color="073A5B", size=10)
        ws["A2"].alignment = left
        ws.row_dimensions[2].height = 20

        # 컬럼 헤더
        headers = ["No.", "공종", "세부 항목", "수량", "단위", "신뢰도", "비고"]
        col_widths = [6, 18, 30, 12, 8, 10, 35]
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="1a5276") if i % 2 == 0 else navy_fill
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[3].height = 22

        # 데이터 행
        items = data.get("items", []) or []
        for idx, item in enumerate(items, 1):
            row = idx + 3
            confidence_color = {
                "high": "D5F5E3",
                "medium": "FEF9E7",
                "low": "FADBD8",
            }.get(item.get("confidence", "medium"), "FFFFFF")

            values = [
                idx,
                item.get("category", ""),
                item.get("sub_category", ""),
                item.get("quantity", ""),
                item.get("unit", ""),
                {"high": "높음", "medium": "보통", "low": "낮음"}.get(item.get("confidence", ""), ""),
                item.get("note", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = body_font
                cell.fill = PatternFill("solid", fgColor=confidence_color)
                cell.alignment = center if col in (1, 4, 5, 6) else left
                cell.border = thin
            ws.row_dimensions[row].height = 18

        # 합계 행 (콘크리트 m³ / 철근 ton)
        last_row = len(items) + 4
        ws.merge_cells(f"A{last_row}:C{last_row}")
        ws[f"A{last_row}"] = "합계 (참고)"
        ws[f"A{last_row}"].font = sub_header_font
        ws[f"A{last_row}"].fill = light_fill
        ws[f"A{last_row}"].alignment = center
        ws[f"A{last_row}"].border = thin

        concrete_total = sum(
            item.get("quantity", 0) or 0
            for item in items if item.get("unit") == "m³"
        )
        rebar_total = sum(
            item.get("quantity", 0) or 0
            for item in items if item.get("unit") == "ton"
        )
        ws[f"D{last_row}"] = round(concrete_total, 2)
        ws[f"D{last_row}"].font = sub_header_font
        ws[f"D{last_row}"].fill = light_fill
        ws[f"D{last_row}"].border = thin
        ws[f"E{last_row}"] = f"m³(콘크리트) / {round(rebar_total,2)}ton(철근)"
        ws[f"E{last_row}"].font = body_font
        ws[f"E{last_row}"].fill = light_fill
        ws.merge_cells(f"E{last_row}:G{last_row}")
        ws[f"E{last_row}"].border = thin

        # 경고 사항
        warnings = data.get("warnings", [])
        if warnings:
            warn_row = last_row + 2
            ws[f"A{warn_row}"] = "⚠ 주의사항"
            ws[f"A{warn_row}"].font = Font(name="맑은 고딕", bold=True, color="E74C3C")
            for i, w in enumerate(warnings):
                ws[f"A{warn_row+1+i}"] = f"• {w}"
                ws[f"A{warn_row+1+i}"].font = Font(name="맑은 고딕", size=9, color="922B21")
                ws.merge_cells(f"A{warn_row+1+i}:G{warn_row+1+i}")

        # 누락 정보
        missing = data.get("missing_info", [])
        if missing:
            m_row = last_row + 2 + len(warnings) + 2
            ws[f"A{m_row}"] = "확인 필요 항목"
            ws[f"A{m_row}"].font = Font(name="맑은 고딕", bold=True, color="1A5276")
            for i, m in enumerate(missing):
                ws[f"A{m_row+1+i}"] = f"• {m}"
                ws[f"A{m_row+1+i}"].font = Font(name="맑은 고딕", size=9, color="1A5276")
                ws.merge_cells(f"A{m_row+1+i}:G{m_row+1+i}")

        ws.freeze_panes = "A4"

    def write_floor_breakdown_sheet(ws, floor_breakdown):
        """구조 부재를 구간(동/섹션) x 층(zone) 기준 물량 소계로 펼쳐서 보여주는 시트.
        compute_floor_breakdown()이 만든 [{"section", "concrete_m3", "formwork_m2", "rebar_kg",
        "floors": [{"zone", "concrete_m3", "formwork_m2", "rebar_kg", "categories":[...]}]}]
        중첩 리스트를 그대로 받아 그린다. 슬래브/보(수평재)는 이미 시공순서 기준(물리적 바로
        위층)으로 재배정된 상태로 들어온다 — compute_floor_breakdown() 참고."""
        ws.title = "층별 수량산출"

        ws.merge_cells("A1:E1")
        ws["A1"] = "AI 수량산출 결과 — 구간(동)x층별 소계 (시공순서 기준, 기준층 반복·철근 할증 적용)"
        ws["A1"].font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
        ws["A1"].fill = navy_fill
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 32

        headers = ["구간/층", "부재종류", "콘크리트(m³)", "거푸집(m²)", "철근(ton)"]
        col_widths = [22, 18, 14, 14, 14]
        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[2].height = 22

        row = 3
        grand_concrete = 0.0
        grand_formwork = 0.0
        grand_rebar_kg = 0.0

        for section in floor_breakdown:
            sec_row = row
            ws.cell(row=sec_row, column=1, value=f"[{section.get('section', '미상')}] 구간 합계")
            ws.cell(row=sec_row, column=2, value="")
            ws.cell(row=sec_row, column=3, value=section.get("concrete_m3", 0))
            ws.cell(row=sec_row, column=4, value=section.get("formwork_m2", 0))
            ws.cell(row=sec_row, column=5, value=round((section.get("rebar_kg", 0) or 0) / 1000, 3))
            for col in range(1, 6):
                c = ws.cell(row=sec_row, column=col)
                c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
                c.fill = navy_fill
                c.alignment = center if col != 1 else left
                c.border = thin
            ws.merge_cells(start_row=sec_row, start_column=1, end_row=sec_row, end_column=2)
            ws.row_dimensions[sec_row].height = 20
            row += 1

            for floor in section.get("floors", []):
                zone_row = row
                ws.cell(row=zone_row, column=1, value="  " + floor.get("zone", "미상"))
                ws.cell(row=zone_row, column=2, value="(층 소계)")
                ws.cell(row=zone_row, column=3, value=floor.get("concrete_m3", 0))
                ws.cell(row=zone_row, column=4, value=floor.get("formwork_m2", 0))
                ws.cell(row=zone_row, column=5, value=round((floor.get("rebar_kg", 0) or 0) / 1000, 3))
                for col in range(1, 6):
                    c = ws.cell(row=zone_row, column=col)
                    c.font = sub_header_font
                    c.fill = yellow_fill if col == 1 else light_fill
                    c.alignment = center if col != 1 else left
                    c.border = thin
                ws.row_dimensions[zone_row].height = 20
                row += 1

                for cat in floor.get("categories", []):
                    ws.cell(row=row, column=1, value="")
                    ws.cell(row=row, column=2, value="    " + cat.get("category", ""))
                    ws.cell(row=row, column=3, value=cat.get("concrete_m3", 0))
                    ws.cell(row=row, column=4, value=cat.get("formwork_m2", 0))
                    ws.cell(row=row, column=5, value=round((cat.get("rebar_kg", 0) or 0) / 1000, 3))
                    for col in range(1, 6):
                        c = ws.cell(row=row, column=col)
                        c.font = body_font
                        c.alignment = center if col != 2 else left
                        c.border = thin
                    ws.row_dimensions[row].height = 18
                    row += 1

            grand_concrete += section.get("concrete_m3", 0) or 0
            grand_formwork += section.get("formwork_m2", 0) or 0
            grand_rebar_kg += section.get("rebar_kg", 0) or 0
            row += 1  # 구간 사이 여백

        ws.cell(row=row, column=1, value="전체 합계")
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value=round(grand_concrete, 3))
        ws.cell(row=row, column=4, value=round(grand_formwork, 2))
        ws.cell(row=row, column=5, value=round(grand_rebar_kg / 1000, 3))
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
            c.fill = green_fill
            c.alignment = center if col != 1 else left
            c.border = thin
        ws.row_dimensions[row].height = 20

        ws.freeze_panes = "A3"

    # 구조 시트
    if "structural" in results and results["structural"].get("items"):
        ws_s = wb.active
        write_sheet(ws_s, results["structural"], "구조 수량산출")
    else:
        wb.active.title = "구조 수량산출"
        wb.active["A1"] = "구조도면 데이터 없음"

    # 층별 수량산출 시트 (zone/floor_repeat_count 판독이 되어 소계가 있을 때만 생성)
    struct_floor_breakdown = (results.get("structural") or {}).get("floor_breakdown") or []
    if struct_floor_breakdown:
        ws_floor = wb.create_sheet("층별 수량산출")
        write_floor_breakdown_sheet(ws_floor, struct_floor_breakdown)

    # 건축 시트
    if "architectural" in results and results["architectural"].get("items"):
        ws_a = wb.create_sheet("건축 수량산출")
        write_sheet(ws_a, results["architectural"], "건축 수량산출")

    # 통합 시트
    ws_total = wb.create_sheet("통합 요약")
    ws_total["A1"] = "AI 구조산출 자동화 — 통합 요약"
    ws_total["A1"].font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)
    ws_total["A1"].fill = navy_fill
    ws_total.merge_cells("A1:E1")
    ws_total["A1"].alignment = center
    ws_total.row_dimensions[1].height = 32

    row = 3
    for type_key, type_name in [("structural", "구조"), ("architectural", "건축")]:
        data = results.get(type_key, {})
        items = data.get("items", []) or []
        if not items:
            continue

        ws_total[f"A{row}"] = f"[{type_name} 수량산출 요약]"
        ws_total[f"A{row}"].font = Font(name="맑은 고딕", bold=True, color="073A5B", size=11)
        ws_total.merge_cells(f"A{row}:E{row}")
        row += 1

        categories = {}
        for item in items:
            cat = item.get("category", "기타")
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)

        for cat, cat_items in categories.items():
            ws_total[f"A{row}"] = cat
            ws_total[f"A{row}"].font = Font(name="맑은 고딕", bold=True, size=10)
            ws_total[f"A{row}"].fill = light_fill
            ws_total.merge_cells(f"A{row}:E{row}")
            row += 1
            for item in cat_items:
                ws_total[f"B{row}"] = item.get("sub_category", "")
                ws_total[f"C{row}"] = item.get("quantity", "")
                ws_total[f"D{row}"] = item.get("unit", "")
                ws_total[f"E{row}"] = {"high": "●", "medium": "◐", "low": "○"}.get(
                    item.get("confidence", ""), "?"
                )
                for col in "BCDE":
                    ws_total[f"{col}{row}"].font = body_font
                    ws_total[f"{col}{row}"].border = thin
                row += 1
        row += 1

    for col_letter, width in zip("ABCDE", [30, 35, 12, 8, 8]):
        ws_total.column_dimensions[col_letter].width = width

    # 응답
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="CBL_수량산출결과.xlsx"'
    return response


@require_POST
@_admin_only_json
def api_download_excel(request):
    """엑셀 다운로드 뷰. 실제 생성 작업은 _build_excel_response()가 담당하고, 여기서는
    그 과정에서 나는 모든 예외를 잡아서 사용자에게 실패 사유를 그대로 보여준다 — 예전에는
    (예: results 구조가 예상과 달라 openpyxl에서 KeyError/TypeError가 나는 경우) 예외가
    Django 500 에러로 그대로 터져서, 사용자는 이유를 알 수 없는 "엑셀 다운로드에 실패했어요"
    메시지만 봐야 했다."""
    try:
        return _build_excel_response(request)
    except Exception as e:
        return HttpResponse(f"엑셀 생성 중 오류가 발생했습니다: {str(e)[:300]}", status=500)

# CBL_QUANTITY_GENERAL_NOTES_SOURCE_GROUNDED_V5
# 후보 도면의 실제 구조재료 적용문을 전사하고, 적용값은 원문 구역에서 로컬 구조화한다.
GENERAL_NOTES_OVERVIEW_TRANSCRIPTION_PROMPT_V5 = """이 이미지는 구조일반사항 후보 도면입니다.
도면번호·도면명과 구조재료 적용문이 보이는 영역을 우선 전사하세요. 특정 도면번호나
목차번호를 전제로 하지 마세요. 값을 해석하거나 요약하지 말고 보이는 원문만 전사하세요.
특히 6. 구조재료 및 강도의 콘크리트와 철근 문장은 숫자, 이하/이상, 강종, fy,
괄호 안 적용조건까지 원문 그대로 보존하세요. 보이지 않는 내용을 만들지 마세요.
JSON 객체 하나만 반환하세요.
{"drawing_number":"실제 보이는 도면번호 또는 null","drawing_title":"실제 도면명",
 "source_text":"구조재료 적용문을 포함한 실제 원문 전사"}"""


def _cbl_v5_evidence(section_path, quote, pdf_page=None, drawing_number=None,
                     drawing_title="구조일반사항", file_type="구조 PDF",
                     method="pdf_transcription"):
    return {
        "file_type": file_type, "pdf_page": pdf_page,
        "drawing_number": drawing_number, "drawing_title": drawing_title,
        "section_path": section_path, "quote": quote, "method": method,
        "confidence": 0.9,
    }


def _cbl_v5_clean_line(value):
    line = re.sub(r"^[\s>*·•-]+", "", str(value or "").strip())
    return re.sub(r"\s+", " ", line).strip()


def _cbl_v5_section(source_text, start_pattern, end_pattern=None):
    text = str(source_text or "")
    start = re.search(start_pattern, text, flags=re.I)
    if not start:
        return ""
    tail = text[start.start():]
    if end_pattern:
        start_length = start.end() - start.start()
        end = re.search(end_pattern, tail[start_length:], flags=re.I)
        if end:
            return tail[:start_length + end.start()]
    return tail


def _cbl_v5_crop_overview(image):
    if not isinstance(image, Image.Image):
        return image
    width, height = image.size
    # 고정된 도면 좌측 패널을 가정하지 않는다. 전체 페이지를 먼저 판독해
    # 구조재료 적용문과 표 영역을 의미적으로 구분하도록 전달한다.
    return image.copy()


def _cbl_v5_parse_source(source_text, pdf_page=None, drawing_number=None,
                         drawing_title="구조일반사항", file_type="구조 PDF",
                         method="pdf_transcription"):
    source_text = str(source_text or "").strip()
    data = {
        "source_text": source_text, "basic_info": {},
        "concrete_materials": [], "rebar_materials": [],
        "cover_requirements": [], "anchorage_splice_requirements": [],
        "quantity_notes": [], "conflicts": [], "unconfirmed_items": [],
    }
    material = _cbl_v5_section(
        source_text,
        r"(?:^|\n)\s*(?:\d+(?:\.\d+)?\s*[.)]?\s*)?(?:구조재료\s*및\s*강도|구조재료|구조재료표|STRUCTURAL\s+MATERIALS?)",
        r"(?:^|\n)\s*(?:\d+(?:\.\d+)?\s*[.)]?\s*)?(?:특기사항|GENERAL\s+NOTES|피복|정착|이음)",
    )
    if not material:
        data["unconfirmed_items"].append("구조재료 및 강도 원문 구역")
        return data

    all_lines = [_cbl_v5_clean_line(v) for v in source_text.splitlines()
                 if _cbl_v5_clean_line(v)]
    lines = [_cbl_v5_clean_line(v) for v in material.splitlines()
             if _cbl_v5_clean_line(v)]
    heading = next((line for line in all_lines if re.search(
        r"구조재료\s*및\s*강도|구조재료|STRUCTURAL\s+MATERIALS?", line,
        re.I,
    )), "구조재료 및 강도")
    base = heading
    concrete_path = base + " > 콘크리트 적용"
    rebar_path = base + " > 철근 적용"

    def basic(key, value, quote, section, unit=None, scope=None):
        if value in (None, ""):
            return
        data["basic_info"][key] = {
            "value": value, "unit": unit, "scope": scope,
            "evidence": _cbl_v5_evidence(
                section, quote, pdf_page, drawing_number, drawing_title,
                file_type, method,
            ),
        }

    for line in all_lines:
        numbered = re.sub(r"^\s*\d+(?:\.\d+)?\s*", "", line).strip()
        matches = (
            (r"아파트\s*[:：]\s*(.+)", "structure_system_apartment",
             "구조형식", "아파트"),
            (r"지하주차장\s*[:：]\s*(.+)", "structure_system_parking",
             "구조형식", "지하주차장"),
            (r"지진력저항시스템\s*[:：]\s*(.+)", "seismic_force_resisting_system",
             "내진형식", None),
            (r"내진설계범주\s*[:：]\s*(.+)", "seismic_design_category",
             "내진형식", None),
            (r"아파트\s*기초\s*[:：]\s*(.+)", "foundation_type_apartment",
             "기초형식", "아파트"),
            (r"지하주차장\s*기초\s*[:：]\s*(.+)", "foundation_type_parking",
             "기초형식", "지하주차장"),
        )
        for pattern, key, section, scope in matches:
            if re.search(pattern, line):
                basic(key, re.split(r"[:：]", line, maxsplit=1)[1].strip(),
                      line, section, scope=scope)
                break
        else:
            if "건축구조기준" in numbered and "(" in numbered:
                basic("design_code_building", numbered, line, base + " > 2. 설계기준")
            elif "콘크리트구조 설계" in numbered and "(" in numbered:
                basic("design_code_concrete", numbered, line, base + " > 2. 설계기준")

    foundation_scope = None
    for line in all_lines:
        if re.search(r"아파트\s*기초\s*[:：]", line):
            foundation_scope = "아파트 기초"
            continue
        if re.search(r"지하주차장\s*기초\s*[:：]", line):
            foundation_scope = "지하주차장 기초"
            continue
        bearing = re.search(
            r"설계요구지내력\s*=\s*(\d+(?:\.\d+)?)\s*kN\s*/\s*m(?:2|²)",
            line, flags=re.I,
        )
        if bearing and foundation_scope:
            key = ("soil_bearing_capacity_apartment"
                   if foundation_scope == "아파트 기초"
                   else "soil_bearing_capacity_parking")
            raw = bearing.group(1)
            basic(key, float(raw) if "." in raw else int(raw), line,
                  "기초형식", unit="kN/m²",
                  scope=foundation_scope)
            foundation_scope = None
        groundwater = re.search(
            r"(지하외벽|기초\s*및\s*내수압슬래브)\s*[:：]\s*(.+)", line,
        )
        if groundwater:
            scope = re.sub(r"\s+", " ", groundwater.group(1)).strip()
            key = ("groundwater_external_wall" if "지하외벽" in scope
                   else "groundwater_foundation_slab")
            basic(key, groundwater.group(2).strip(), line,
                  "기초형식 > 지하수위", scope=scope)

    for line in lines:
        fck = re.search(r"\bF\s*CK\s*=\s*(\d+(?:\.\d+)?)\s*MPA\b",
                        line, flags=re.I)
        if not fck:
            continue
        prefix = re.sub(r"^\s*\d+\s*[.)]\s*", "",
                        line[:fck.start()].strip(" :：,")).strip()
        member = re.search(r"(전\s*부재(?:\s*\([^)]*\))?)", prefix)
        member_type = (re.sub(r"\s+", " ", member.group(1)).strip()
                       if member else prefix)
        location = prefix[:member.start()].strip(" ,:：") if member else prefix
        raw = fck.group(1)
        data["concrete_materials"].append({
            "location": location or None, "member_type": member_type or None,
            "floor_scope": None,
            "fck_mpa": float(raw) if "." in raw else int(raw),
            "slump_mm": None, "aggregate_mm": None,
            "exposure_condition": None,
            "evidence": _cbl_v5_evidence(
                concrete_path, line, pdf_page, drawing_number, drawing_title,
                file_type, method,
            ),
        })

    for line in lines:
        match = re.search(
            r"D\s*(\d+)\s*(이하|이상)\s*[:：]?\s*"
            r"(SD\s*\d+\s*S?)\s*,?\s*FY\s*=\s*(\d+(?:\.\d+)?)\s*MPA",
            line, flags=re.I,
        )
        if not match:
            continue
        diameter = int(match.group(1))
        comparison = match.group(2)
        grade = re.sub(r"\s+", "", match.group(3)).upper()
        raw = match.group(4)
        data["rebar_materials"].append({
            "diameter_min_mm": diameter if comparison == "이상" else None,
            "diameter_max_mm": diameter if comparison == "이하" else None,
            "diameter_rule": f"D{diameter} {comparison}",
            "grade": grade,
            "fy_mpa": float(raw) if "." in raw else int(raw),
            "member_scope": None,
            "material_type": "내진용 철근" if grade.endswith("S") else "일반철근",
            "evidence": _cbl_v5_evidence(
                rebar_path, line, pdf_page, drawing_number, drawing_title,
                file_type, method,
            ),
        })

    for line in lines:
        scope = re.search(r"내진용\s*철근\s*[:：]\s*(.+)", line)
        if scope and "적용" in scope.group(1):
            data["quantity_notes"].append({
                "category": "내진용 철근 적용범위", "text": line,
                "affected_work": scope.group(1).strip(),
                "evidence": _cbl_v5_evidence(
                    rebar_path, line, pdf_page, drawing_number, drawing_title,
                    file_type, method,
                ),
            })
    for line in all_lines:
        if re.search(r"(?:시공이음|끊어치기|보강근|버림콘크리트|무근콘크리트|개구부)", line):
            data["quantity_notes"].append({
                "category": "구조 특기사항", "text": line,
                "affected_work": "기초·터파기·배수",
                "evidence": _cbl_v5_evidence(
                    "특기사항", line, pdf_page,
                    drawing_number, drawing_title, file_type, method,
                ),
            })
    return data


_cbl_v5_original_material_statement = _general_notes_is_material_strength_statement
def _general_notes_is_material_strength_statement(evidence):
    quote = re.sub(r"\s+", "", str((evidence or {}).get("quote") or "")).upper()
    section = re.sub(r"\s+", "", str((evidence or {}).get("section_path") or "")).upper()
    return "구조재료및강도" in (section or quote) and not any(
        marker in quote for marker in ("이음길이표", "정착길이표", "피복두께표")
    )


_cbl_v5_original_classifier = _classify_general_notes_image_pages
def _cbl_v5_classify_general_notes(pdf_bytes, page_numbers, job_id=None):
    decisions = []
    size = max(1, int(GENERAL_NOTES_LOCATOR_BATCH_SIZE))
    for offset in range(0, len(page_numbers), size):
        batch = page_numbers[offset:offset + size]
        decisions.extend(_cbl_v5_original_classifier(
            pdf_bytes, batch, job_id=job_id,
        ) or [])
        material_markers = ("구조재료", "콘크리트", "fck", "mpa", "철근", "sd", "fy")
        if any(
            row.get("is_general_notes") and any(
                marker in str(term).lower()
                for term in (row.get("evidence_terms") or [])
                for marker in material_markers
            )
            for row in decisions
        ):
            _general_notes_log(
                "vision_locator_candidate_found", job_id=job_id,
                scanned_through_page=max(batch),
            )
            break
    return decisions

def _cbl_v5_select_material_page(scan, decisions, selected, mapping):
    scan_by_page = {int(row.get("page")): row for row in (scan or {}).get("pages") or []
                    if row.get("page") is not None}
    candidates = []
    for row in decisions or []:
        if not row.get("is_general_notes") or row.get("page_type") == "drawing_list":
            continue
        terms = " ".join(str(v) for v in (row.get("evidence_terms") or []))
        scan_row = scan_by_page.get(int(row.get("pdf_page") or 0), {})
        haystack = (terms + " " + " ".join(str(v) for v in scan_row.get("reasons") or [])).lower()
        score = sum(5 for marker in ("구조재료", "콘크리트", "fck", "mpa", "철근", "sd", "fy")
                    if marker in haystack)
        candidates.append((score, int(row["pdf_page"]), row))
    candidates.sort(key=lambda item: (-item[0], item[1]))
    for score, page, row in candidates:
        _general_notes_log("material_page_score", page=page, score=score,
                           drawing_number=row.get("drawing_number"),
                           evidence_terms=row.get("evidence_terms") or [])
    page = candidates[0][1] if candidates and candidates[0][0] > 0 else None
    if page is None:
        page = min(selected) if selected else None
    drawing_list_pages = set((scan or {}).get("drawing_list_pages") or [])
    return None if page in drawing_list_pages else page


def _cbl_v5_finish_result(result):
    result = result or _empty_general_notes_result(
        "구조일반사항 내용을 확인하지 못했습니다."
    )
    result["unconfirmed_items"] = [
        item for item in (result.get("unconfirmed_items") or [])
        if item not in ("피복두께", "정착·이음 기준")
    ]
    result["deferred_items"] = [
        "피복두께는 부재별 배근 확인 단계에서 해당 부재 조건표를 적용합니다.",
        "정착·이음 길이는 실제 부재·철근직경·콘크리트강도 확인 후 해당 표 행을 적용합니다.",
    ]
    result["lap_splice_class"] = "B"
    notes = result.setdefault("quantity_notes", [])
    if not any(row.get("source_type") == "user_confirmed"
               and "B급" in str(row.get("text") or "") for row in notes):
        notes.append({
            "category": "이음 정책",
            "text": "전 부재 B급 인장이음 적용",
            "affected_work": "전 부재",
            "source_type": "user_confirmed",
            "provenance": "사용자 확정조건",
            "evidence": None,
        })
    return result


def extract_general_notes(structural_pdf_bytes, page_hints=None, overview=None,
                          job_id=None, timeout_seconds=GENERAL_NOTES_TIMEOUT_SEC,
                          cad_context=None):
    started = time.monotonic()
    total_pages = (int(pdfinfo_from_bytes(structural_pdf_bytes).get("Pages", 0) or 0)
                   if structural_pdf_bytes else 0)
    scan = (_general_notes_page_candidates(
        structural_pdf_bytes, total_pages, page_hints,
    ) if structural_pdf_bytes else {
        "pages": [], "scan_range": [], "selected_pages": [], "text_used": False,
        "image_fallback": False, "drawing_list_pages": [],
        "expected_drawing_numbers": [],
    })
    for row in scan.get("pages") or []:
        _general_notes_log("page_candidate", job_id=job_id, **row)

    vision_pages = []
    if structural_pdf_bytes and total_pages:
        start_page = (max(scan.get("drawing_list_pages") or [0]) + 1
                      if scan.get("drawing_list_pages") else 1)
        end_page = min(total_pages, GENERAL_NOTES_LOCATOR_MAX_PAGE)
        text_by_page = {row["page"]: row for row in scan.get("pages") or []}
        vision_pages = [
            page for page in range(start_page, end_page + 1)
            if page not in set(scan.get("drawing_list_pages") or [])
            and not (text_by_page.get(page) or {}).get("text_available")
        ]
    decisions = (_cbl_v5_classify_general_notes(
        structural_pdf_bytes, vision_pages, job_id=job_id,
    ) if vision_pages else [])
    selected, mapping = _merge_general_notes_page_candidates(scan, decisions)
    page = _cbl_v5_select_material_page(scan, decisions, selected, mapping)
    selected_pages = [page] if isinstance(page, int) else []

    cad_context = [
        row for row in (cad_context or [])
        if str(row.get("text") or "").strip()
    ]
    result = None
    if selected_pages:
        remaining = timeout_seconds - (time.monotonic() - started)
        if remaining <= 0:
            raise TimeoutError("구조일반사항 페이지 렌더링 시간 초과")
        images = _render_pdf_page_range(
            structural_pdf_bytes, page, page, dpi=GENERAL_NOTES_RENDER_DPI,
            timeout=max(10, int(remaining)),
        )
        if images:
            panel = _cbl_v5_crop_overview(images[0])
            _general_notes_log(
                "overview_panel_selected", job_id=job_id, pdf_page=page,
                source_size=list(images[0].size), crop_size=list(panel.size),
                included_drawing_number=None,
                excluded_sections=[
                    "B. 설계도서 일반사항", "C. 철근콘크리트 일반사항",
                ],
            )
            client = get_gemini_client()
            if client is None:
                raise ValueError("GEMINI_API_KEY가 설정되지 않았습니다.")
            response = client.models.generate_content(
                model=GEMINI_QUANTITY_MODEL,
                contents=[
                    f"[구조 PDF 실제 {page}페이지 · 도면번호는 이미지에서 확인]",
                    types.Part.from_bytes(
                        data=image_to_jpeg_bytes(panel, max_size=(1600, 2600)),
                        mime_type="image/jpeg",
                    ),
                ],
                config=types.GenerateContentConfig(
                    system_instruction=GENERAL_NOTES_OVERVIEW_TRANSCRIPTION_PROMPT_V5,
                    response_mime_type="application/json", temperature=0.0,
                    max_output_tokens=4096,
                    thinking_config=types.ThinkingConfig(thinking_budget=256),
                ),
            )
            raw = _extract_text_from_gemini_response(response)
            diagnostics = _gemini_response_diagnostics(response)
            first_candidate = (diagnostics.get("candidates") or [{}])[0]
            usage = diagnostics.get("usage") or {}
            _general_notes_log(
                "overview_transcription_response", job_id=job_id,
                pdf_page=page,
                finish_reason=first_candidate.get("finish_reason"),
                prompt_token_count=usage.get("prompt_token_count"),
                candidates_token_count=usage.get("candidates_token_count"),
                thoughts_token_count=usage.get("thoughts_token_count"),
                total_token_count=usage.get("total_token_count"),
                raw_length=len(raw),
            )
            cleaned = raw.replace("```json", "").replace("```", "").strip()
            try:
                transcript = json.loads(cleaned)
            except json.JSONDecodeError:
                transcript = _try_repair_truncated_json(cleaned)
            if not isinstance(transcript, dict):
                raise ValueError("구조개요 전사 JSON 파싱 실패")
            source_text = str(transcript.get("source_text") or "").strip()
            actual_number = str(transcript.get("drawing_number") or "").strip() or None
            actual_title = str(transcript.get("drawing_title") or "구조일반사항").strip()
            parsed = _cbl_v5_parse_source(
                source_text, pdf_page=page, drawing_number=actual_number,
                drawing_title=str(
                    actual_title
                ),
            )
            result = _validate_general_notes_result(
                parsed, [page], overview, source_text=source_text,
            )
            if actual_number:
                mapping[actual_number] = page

    fallback_reason = None
    if cad_context and not _general_notes_has_values(result):
        fallback_reason = ("no_pdf_content_candidate" if not selected_pages
                           else "empty_pdf_extraction")
        source_text = "\n".join(str(row.get("text") or "")
                                for row in cad_context)
        primary = cad_context[0]
        cad_drawing_number = re.search(
            r"\b([SA]-\s*\d{3}(?:\s*[~,\-/]\s*\d{3})?)\b",
            str(primary.get("filename") or primary.get("path") or ""), re.I,
        )
        parsed = _cbl_v5_parse_source(
            source_text, pdf_page=None,
            drawing_number=(cad_drawing_number.group(1).replace(" ", "")
                            if cad_drawing_number else "CAD-구조일반사항"),
            drawing_title=str(primary.get("filename") or "구조일반사항"),
            file_type="구조 CAD", method="cad_text",
        )
        result = _validate_general_notes_result(
            parsed, [], overview, source_text=source_text,
        )

    result = _cbl_v5_finish_result(result)
    result["page_candidates"] = scan.get("pages") or []
    result["selected_pages"] = selected_pages
    result["diagnostics"] = {
        "total_pages": total_pages,
        "scan_range": scan.get("scan_range") or [],
        "pdf_text_used": scan.get("text_used", False),
        "image_fallback": bool(vision_pages),
        "drawing_number_page_mapping": mapping,
        "cad_fallback_reason": fallback_reason,
        "material_source_page": page,
        "deferred_drawing_numbers": [],
        "elapsed_seconds": round(time.monotonic() - started, 3),
    }
    _general_notes_log(
        "extraction_complete", job_id=job_id,
        extracted_counts={
            key: len(result.get(key) or []) for key in (
                "concrete_materials", "rebar_materials",
                "cover_requirements", "anchorage_splice_requirements",
                "quantity_notes",
            )
        },
        validation_rejections=result.get("validation_rejections") or [],
        missing_required=result.get("unconfirmed_items") or [],
        conflict_count=len(result.get("conflicts") or []),
        elapsed_seconds=result["diagnostics"]["elapsed_seconds"],
    )
    return result


def _cbl_v5_compact_spec(spec):
    spec = spec or {}
    return {
        "basic_info": {
            key: {
                "value": item.get("value"), "unit": item.get("unit"),
                "scope": item.get("scope"),
            }
            for key, item in (spec.get("basic_info") or {}).items()
            if isinstance(item, dict) and item.get("value") not in (None, "")
        },
        "concrete_materials": [
            {key: row.get(key) for key in (
                "location", "member_type", "floor_scope", "fck_mpa",
            )} for row in spec.get("concrete_materials") or []
        ],
        "rebar_materials": [
            {key: row.get(key) for key in (
                "diameter_min_mm", "diameter_max_mm", "diameter_rule",
                "grade", "fy_mpa", "material_type",
            )} for row in spec.get("rebar_materials") or []
        ],
        "lap_splice_class": spec.get("lap_splice_class") or "B",
        "quantity_policies": [
            {"text": row.get("text"), "source_type": row.get("source_type")}
            for row in spec.get("quantity_notes") or []
            if row.get("source_type") == "user_confirmed"
        ],
    }


_cbl_v5_original_member_batch = _extract_structural_members_one_batch
def _extract_structural_members_one_batch(
    client, dwg_data, image_batch, batch_idx, total_batches,
    page_numbers=None, confirmed_general_spec=None,
):
    compact = (_cbl_v5_compact_spec(confirmed_general_spec)
               if confirmed_general_spec else None)
    return _cbl_v5_original_member_batch(
        client, dwg_data, image_batch, batch_idx, total_batches,
        page_numbers=page_numbers, confirmed_general_spec=compact,
    )


# CBL_DRAWING_COORDINATION_V1
# 구조·건축 PDF의 실제 타이틀블록과 대표 상세를 대조한다. 물량/3D는 만들지 않는다.
DRAWING_COORDINATION_LOCATOR_BATCH = 12
DRAWING_COORDINATION_DETAIL_BATCH = 6
DRAWING_COORDINATION_MAX_DETAIL_PAGES = 54
DRAWING_COORDINATION_IMAGE_PROBE_PAGES = {
    "구조": 48,
    "건축": 72,
}

DRAWING_COORDINATION_LOCATOR_PROMPT = """각 이미지의 타이틀블록을 읽어 JSON만 반환하세요.
{"pages":[{"pdf_page":1,"drawing_number":"S-101","drawing_title":"...",
"drawing_type":"structural_plan|foundation_plan|architectural_plan|elevation|section|
core_plan|core_section|parking_plan|parking_section|ramp|amenity|window_schedule|other",
"building_scope":"101동/102동/주동/지하주차장/공통 또는 null",
"floor_scope":"지하2층/1층/기준층/옥탑층/전체 또는 null",
"confidence":0.0,"evidence_terms":["실제로 보이는 짧은 표기"]}]}
페이지 순서나 파일명으로 도면번호를 추정하지 마세요. 도면목록은 other입니다."""

DRAWING_COORDINATION_DETAIL_PROMPT = """구조·건축 도면의 동·층·레벨·코어·램프·개구부만
실제로 보이는 범위에서 JSON으로 옮기세요. 물량 계산과 3D/매스 생성은 금지합니다.
{"sheets":[{"pdf_page":1,"drawing_number":"A-401","drawing_title":"...",
"building_scope":"101동","floor_scope":"1층",
"levels":[{"label":"1FL","elevation_m":0.0,"floor_height_m":3.2,"quote":"실제 표기"}],
"structural_scope":["벽체","기둥","보","슬래브","기초"],
"cores":[{"type":"엘리베이터|계단","label":"...","location":"...","quote":"..."}],
"openings":[{"type":"창호|출입문|슬래브 오픈구|설비 샤프트|전기 샤프트|배관 샤프트|
PIT|VOID|램프|지하외벽 개구부|내력벽 개구부","label":"...","location":"...",
"floor_scope":"...","width_m":null,"height_m":null,"quote":"실제 표기"}],
"ramps":[{"label":"...","slope":null,"level_from":null,"level_to":null,"quote":"..."}],
"unconfirmed":[],"confidence":0.0}]}
보이지 않는 값은 null로 두고 근거를 만들지 마세요."""


def _drawing_coordination_log(event, **payload):
    logger.info("quantity_drawing_coordination %s", json.dumps(
        {"event": event, **payload}, ensure_ascii=False, sort_keys=True, default=str,
    ))


def _coordination_title_crop(image):
    if not isinstance(image, Image.Image):
        return image
    width, height = image.size
    if width >= height:
        return image.crop((int(width * .60), 0, width, height))
    return image.crop((int(width * .50), int(height * .52), width, height))


def _coordination_json(raw):
    cleaned = str(raw or "").replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        return _try_repair_truncated_json(cleaned) or {}


def _coordination_probe_pages(pdf_bytes, discipline, total):
    """텍스트 레이어가 있는 후보와 초기 이미지 구간만 locator에 전달한다.

    도면집 전체를 매 페이지 이미지로 보내면 200페이지 이상에서 시간 제한을
    소진한다. pdftotext는 한 번만 실행하고, 도면번호/도면명 표식이 있는 페이지와
    도면목록 뒤의 초기 이미지 구간을 합쳐 실제 후보를 놓치지 않게 한다.
    """
    limit = min(int(total or 0), int(DRAWING_COORDINATION_IMAGE_PROBE_PAGES.get(
        discipline, 48)))
    pages = set(range(1, limit + 1))
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as handle:
            handle.write(pdf_bytes or b"")
            tmp_path = handle.name
        proc = subprocess.run(
            ["pdftotext", "-layout", tmp_path, "-"],
            capture_output=True, timeout=90,
        )
        text = proc.stdout.decode("utf-8", errors="ignore")
        for page, chunk in enumerate(text.split("\f"), 1):
            if page > total:
                break
            if re.search(r"\b[SA]-\s*\d{3}", chunk, re.I) or any(
                marker in chunk for marker in (
                    "도면목록", "도면번호", "구조평면", "기초", "주차장",
                    "평면도", "입면도", "단면도", "코어", "경사로", "창호",
                )
            ):
                pages.add(page)
    except Exception:
        pass
    finally:
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
    selected = sorted(page for page in pages if 1 <= page <= total)
    return selected, limit


def _locate_coordination_pages(pdf_bytes, discipline, job_id=None):
    total = int((pdfinfo_from_bytes(pdf_bytes) or {}).get("Pages", 0) or 0)
    client = get_gemini_client()
    if not client:
        raise ValueError("GEMINI_API_KEY가 설정되지 않았습니다.")
    rows = {}
    probe_pages, probe_limit = _coordination_probe_pages(pdf_bytes, discipline, total)
    _drawing_coordination_log(
        "locator_probe", job_id=job_id, discipline=discipline,
        total_pages=total, probe_limit=probe_limit,
        probe_page_count=len(probe_pages), probe_pages=probe_pages,
    )
    for offset in range(0, len(probe_pages), DRAWING_COORDINATION_LOCATOR_BATCH):
        requested = probe_pages[offset:offset + DRAWING_COORDINATION_LOCATOR_BATCH]
        contents, rendered = [f"분야={discipline}, PDF 페이지={requested}"], []
        for page in requested:
            images = _render_pdf_page_range(pdf_bytes, page, page, dpi=90, timeout=25)
            if not images:
                continue
            rendered.append(page)
            contents.extend([
                f"[PDF 실제 {page}페이지]",
                types.Part.from_bytes(data=image_to_jpeg_bytes(
                    _coordination_title_crop(images[0]), max_size=(850, 1200),
                ), mime_type="image/jpeg"),
            ])
        if not rendered:
            continue
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL, contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=DRAWING_COORDINATION_LOCATOR_PROMPT,
                response_mime_type="application/json", temperature=0,
                max_output_tokens=4096,
                thinking_config=types.ThinkingConfig(thinking_budget=128),
            ),
        )
        parsed = _coordination_json(_extract_text_from_gemini_response(response))
        items = parsed.get("pages") if isinstance(parsed, dict) else parsed
        for item in items or []:
            if not isinstance(item, dict):
                continue
            try:
                page = int(item.get("pdf_page"))
            except (TypeError, ValueError):
                continue
            if page in rendered:
                rows[page] = item
        diagnostics = _gemini_response_diagnostics(response)
        _drawing_coordination_log(
            "locator_batch", job_id=job_id, discipline=discipline,
            requested_pages=rendered,
            parsed_pages=sorted(page for page in rows if page in rendered),
            finish_reason=((diagnostics.get("candidates") or [{}])[0]).get("finish_reason"),
            usage=diagnostics.get("usage_metadata"),
        )
    return [rows[page] for page in sorted(rows)], total


def _coordination_target_pages(locator_rows):
    wanted = {
        "structural_plan", "foundation_plan", "architectural_plan", "elevation",
        "section", "core_plan", "core_section", "parking_plan",
        "parking_section", "ramp", "amenity", "window_schedule",
    }
    groups = {}
    for row in locator_rows:
        kind = str(row.get("drawing_type") or "other")
        if kind in wanted:
            groups.setdefault(kind, []).append(int(row["pdf_page"]))
    # 각 도면 종류를 먼저 한 장씩 포함한 뒤 남은 상한을 채운다.  단순히
    # locator 응답 순서로 자르면 앞쪽 평면도만 남고 램프/단면/창호표가
    # 조용히 누락될 수 있다.
    normalized = {kind: sorted(set(pages)) for kind, pages in groups.items()}
    # 각 종류의 첫/마지막 도면은 범위 확인에 중요하므로 상한 안에 우선 예약한다.
    tails = {kind: pages[-1] for kind, pages in normalized.items() if pages}
    selected = []
    selection_limit = max(0, DRAWING_COORDINATION_MAX_DETAIL_PAGES - len(tails))
    index = 0
    kinds = list(normalized)
    while kinds and len(selected) < selection_limit:
        next_kinds = []
        for kind in kinds:
            pages = normalized[kind]
            usable = pages[:-1] if len(pages) > 1 else pages
            if index < len(usable):
                selected.append(usable[index])
            if index + 1 < len(usable):
                next_kinds.append(kind)
            if len(selected) >= selection_limit:
                break
        kinds = next_kinds
        index += 1
    selected.extend(tails.values())
    return sorted(set(selected))[:DRAWING_COORDINATION_MAX_DETAIL_PAGES]


def _extract_coordination_details(pdf_bytes, discipline, locator_rows, job_id=None):
    client = get_gemini_client()
    selected = _coordination_target_pages(locator_rows)
    locator = {int(row["pdf_page"]): row for row in locator_rows}
    details = {}
    empty_pages = []
    for offset in range(0, len(selected), DRAWING_COORDINATION_DETAIL_BATCH):
        batch = selected[offset:offset + DRAWING_COORDINATION_DETAIL_BATCH]
        contents, rendered = [
            "locator=" + json.dumps([locator[p] for p in batch], ensure_ascii=False),
        ], []
        for page in batch:
            images = _render_pdf_page_range(pdf_bytes, page, page, dpi=150, timeout=30)
            if not images:
                continue
            rendered.append(page)
            contents.extend([
                f"[PDF 실제 {page}페이지]",
                types.Part.from_bytes(data=image_to_jpeg_bytes(
                    images[0], max_size=(1800, 1800),
                ), mime_type="image/jpeg"),
            ])
        if not rendered:
            continue
        response = client.models.generate_content(
            model=GEMINI_QUANTITY_MODEL, contents=contents,
            config=types.GenerateContentConfig(
                system_instruction=DRAWING_COORDINATION_DETAIL_PROMPT,
                response_mime_type="application/json", temperature=0,
                max_output_tokens=8192,
                thinking_config=types.ThinkingConfig(thinking_budget=512),
            ),
        )
        parsed = _coordination_json(_extract_text_from_gemini_response(response))
        for item in (parsed.get("sheets") if isinstance(parsed, dict) else []) or []:
            if not isinstance(item, dict):
                continue
            try:
                page = int(item.get("pdf_page"))
            except (TypeError, ValueError):
                continue
            if page not in rendered:
                continue
            source = locator.get(page) or {}
            item["drawing_number"] = item.get("drawing_number") or source.get("drawing_number")
            item["drawing_title"] = item.get("drawing_title") or source.get("drawing_title")
            item["discipline"] = discipline
            details[page] = item
        _drawing_coordination_log(
            "detail_batch", job_id=job_id, discipline=discipline,
            pages=rendered, extracted_pages=sorted(
                page for page in details if page in rendered
            ),
        )
        empty_pages.extend(page for page in rendered if page not in details)

    # 여러 장을 한 번에 보낸 응답이 일부 페이지를 생략할 수 있으므로,
    # 해당 페이지에 한해서만 1회 단일 페이지 재시도를 한다. 다른 도면을
    # 다시 호출하지 않으며, 추출 실패를 확인된 값으로 승격하지 않는다.
    for page in empty_pages:
        source = locator.get(page) or {}
        images = _render_pdf_page_range(pdf_bytes, page, page, dpi=150, timeout=30)
        if not images:
            continue
        try:
            response = client.models.generate_content(
                model=GEMINI_QUANTITY_MODEL,
                contents=[
                    "이 한 장의 도면에서 실제로 보이는 레벨, 구조범위, 코어, 개구부, 램프만 JSON으로 추출하세요.",
                    "locator=" + json.dumps(source, ensure_ascii=False),
                    f"[PDF 실제 {page}페이지]",
                    types.Part.from_bytes(data=image_to_jpeg_bytes(
                        images[0], max_size=(1800, 1800),
                    ), mime_type="image/jpeg"),
                ],
                config=types.GenerateContentConfig(
                    system_instruction=DRAWING_COORDINATION_DETAIL_PROMPT,
                    response_mime_type="application/json", temperature=0,
                    max_output_tokens=4096,
                    thinking_config=types.ThinkingConfig(thinking_budget=256),
                ),
            )
            parsed = _coordination_json(_extract_text_from_gemini_response(response))
            candidates = (parsed.get("sheets") if isinstance(parsed, dict) else []) or []
            for item in candidates:
                if not isinstance(item, dict):
                    continue
                try:
                    item_page = int(item.get("pdf_page"))
                except (TypeError, ValueError):
                    continue
                if item_page != page:
                    continue
                item["drawing_number"] = item.get("drawing_number") or source.get("drawing_number")
                item["drawing_title"] = item.get("drawing_title") or source.get("drawing_title")
                item["discipline"] = discipline
                details[page] = item
                break
            _drawing_coordination_log(
                "detail_retry", job_id=job_id, discipline=discipline,
                page=page, extracted=page in details,
                finish_reason=((_gemini_response_diagnostics(response).get("candidates") or [{}])[0]).get("finish_reason"),
            )
        except Exception as exc:
            _drawing_coordination_log(
                "detail_retry_failed", job_id=job_id, discipline=discipline,
                page=page, error=str(exc)[:240],
            )
    return [details[p] for p in sorted(details)], selected


def _coordination_cross_check(structural_sheets, architectural_sheets):
    conflicts, unconfirmed, values = [], [], {}
    for sheet in structural_sheets + architectural_sheets:
        scope = (sheet.get("building_scope"), sheet.get("floor_scope"))
        for level in sheet.get("levels") or []:
            if level.get("elevation_m") is None:
                continue
            values.setdefault((scope, level.get("label")), []).append({
                "value": level.get("elevation_m"), "discipline": sheet.get("discipline"),
                "page": sheet.get("pdf_page"), "drawing_number": sheet.get("drawing_number"),
                "quote": level.get("quote"),
            })
        if not all(scope):
            unconfirmed.append({
                "pdf_page": sheet.get("pdf_page"),
                "drawing_number": sheet.get("drawing_number"),
                "reason": "동 또는 적용 층 확인 불가",
            })
    for key, evidence in values.items():
        if len({str(row["value"]) for row in evidence}) > 1:
            conflicts.append({
                "type": "도면 간 레벨 충돌", "building_floor": key[0],
                "level": key[1], "evidence": evidence,
            })
    return conflicts, unconfirmed


def extract_drawing_coordination(structural_pdf_bytes, architectural_pdf_bytes,
                                 *, structural_cad_bytes=None,
                                 architectural_cad_bytes=None, job_id=None):
    started = time.monotonic()
    struct_map, struct_total = _locate_coordination_pages(
        structural_pdf_bytes, "구조", job_id,
    )
    arch_map, arch_total = _locate_coordination_pages(
        architectural_pdf_bytes, "건축", job_id,
    )
    struct_sheets, struct_selected = _extract_coordination_details(
        structural_pdf_bytes, "구조", struct_map, job_id,
    )
    arch_sheets, arch_selected = _extract_coordination_details(
        architectural_pdf_bytes, "건축", arch_map, job_id,
    )
    conflicts, unconfirmed = _coordination_cross_check(struct_sheets, arch_sheets)
    cad_diagnostics = {}
    for label, cad_bytes in (("structural", structural_cad_bytes),
                             ("architectural", architectural_cad_bytes)):
        if not cad_bytes:
            cad_diagnostics[label] = {"provided": False, "parsed": 0}
            continue
        try:
            parsed_cad = parse_dwg_from_zip(cad_bytes)
            cad_diagnostics[label] = {
                "provided": True,
                "parsed": sum(1 for value in parsed_cad.values() if "error" not in value),
                "errors": {key: value.get("error") for key, value in parsed_cad.items()
                           if isinstance(value, dict) and value.get("error")},
                "paths": list(parsed_cad),
            }
        except Exception as exc:
            cad_diagnostics[label] = {
                "provided": True, "parsed": 0,
                "errors": {"inventory": str(exc)[:240]}, "paths": [],
            }
    result = {
        "page_maps": {"structural": struct_map, "architectural": arch_map},
        "selected_pages": {
            "structural": struct_selected, "architectural": arch_selected,
        },
        "structural_sheets": struct_sheets,
        "architectural_sheets": arch_sheets,
        "conflicts": conflicts, "unconfirmed_items": unconfirmed,
        "diagnostics": {
            "structural_total_pages": struct_total,
            "architectural_total_pages": arch_total,
            "elapsed_seconds": round(time.monotonic() - started, 3),
            "quantities_generated": False, "geometry_generated": False,
            "cad_diagnostics": cad_diagnostics,
        },
    }
    _drawing_coordination_log(
        "complete", job_id=job_id, structural_map_count=len(struct_map),
        architectural_map_count=len(arch_map),
        structural_detail_count=len(struct_sheets),
        architectural_detail_count=len(arch_sheets),
        conflict_count=len(conflicts), unconfirmed_count=len(unconfirmed),
        elapsed_seconds=result["diagnostics"]["elapsed_seconds"],
    )
    return result


def _run_drawing_coordination_job(job_id, review_id, structural_bytes,
                                  architectural_bytes, structural_cad_bytes=None,
                                  architectural_cad_bytes=None):
    try:
        _progress_set(job_id, "drawing_coordination", 0, 1,
                      "구조평면도와 건축 평·입·단면도를 대조하고 있어요...",
                      stage_index=1, total_stages=1)
        result = extract_drawing_coordination(
            structural_bytes, architectural_bytes,
            structural_cad_bytes=structural_cad_bytes,
            architectural_cad_bytes=architectural_cad_bytes,
            job_id=job_id,
        )
        _review_update(review_id, drawing_coordination=result)
        _review_reset_confirmations_from(review_id, "drawing_coordination")
        _result_set(job_id, {"ok": True, "results": {"drawing_coordination": result}})
    except Exception as exc:
        _result_set(job_id, {"ok": False, "error":
                    f"구조·건축 도면 대조 중 오류가 발생했습니다: {str(exc)[:300]}"})
    finally:
        _progress_clear(job_id)


@require_POST
@_admin_only_json
def api_quantity_drawing_coordination_check(request):
    job_id, review_id = request.POST.get("job_id"), request.POST.get("review_id")
    if not job_id or not review_id:
        return JsonResponse({"error": "job_id와 review_id가 필요합니다."}, status=400)
    rec, err = _review_require_stage(review_id, "general_spec_confirmed", "구조일반사항")
    if err:
        return err
    if rec.get("_user_id") != str(request.user.pk):
        return JsonResponse({"error": "다른 사용자의 확인 세션입니다."}, status=403)
    structural_files = request.FILES.getlist("structural_pdf")
    architectural_files = request.FILES.getlist("architectural_pdf")
    if not structural_files or not architectural_files:
        return JsonResponse({"error": "구조 PDF와 건축 PDF가 모두 필요합니다."}, status=400)
    structural_bytes, _ = _merge_uploaded_pdfs(structural_files)
    architectural_bytes, _ = _merge_uploaded_pdfs(architectural_files)
    structural_cad_files = request.FILES.getlist("structural_cad")
    architectural_cad_files = request.FILES.getlist("architectural_cad")
    structural_cad_bytes, _, _ = _merge_uploaded_cad_sets(structural_cad_files)
    architectural_cad_bytes, _, _ = _merge_uploaded_cad_sets(architectural_cad_files)
    hashes = _review_file_hashes(
        structural_pdf_bytes=structural_bytes,
        architectural_pdf_bytes=architectural_bytes,
        structural_zip_bytes=structural_cad_bytes,
        architectural_zip_bytes=architectural_cad_bytes,
    )
    if not _matching_uploaded_files(rec, hashes):
        return JsonResponse({"error":
            "업로드 파일이 바뀌었습니다. 프로젝트 개요부터 다시 확인해 주세요."}, status=409)
    _progress_set(job_id, "queued", 0, 1, "대기열에 등록됨",
                  stage_index=1, total_stages=1)
    threading.Thread(
        target=_run_drawing_coordination_job,
        args=(job_id, review_id, structural_bytes, architectural_bytes,
              structural_cad_bytes, architectural_cad_bytes),
        daemon=True,
    ).start()
    return JsonResponse({"accepted": True, "job_id": job_id})
